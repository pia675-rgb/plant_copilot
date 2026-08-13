#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
lists.py — IO List 와 계기 리스트 조인 + P&ID TAG 자동 부여

키 역할
------
· **P&ID TAG**  : 앱 기준 키 (알람·매뉴얼·인터락·계기·태그 목록)
· **TAG (IO)**  : 판넬/카드/채널/배선 전용

P&ID TAG 자동 채움 (로드 시)
---------------------------
IO List 의 P&ID TAG 가 비어 있으면 Instrument / Interlock 후보로 채운다.
  1) 우선순위: Instrument → Interlock
  2) 숫자(≥3자리) 겹침 + description 유사도 ≥ 0.7
  3) 태그 문자열 포함 관계( POU_AIT_P4113 ⊃ AIT-P4113 ) 가산
  4) 다중 매칭 → 비움 / 못 찾음 → IO TAG 폴백
별도 fill_pid_tag 스크립트를 돌릴 필요 없다.

Instrument / Interlock 멀티시트
-------------------------------
시트 여러 개여도 전부 읽는다.
"""

from __future__ import annotations

import os
import re
from collections import OrderedDict
from difflib import SequenceMatcher

from openpyxl import load_workbook

COLUMN_ALIASES = {"SERVICE": "DESCRIPTION"}

INSTRUMENT_MAP = {
    "TAG NO.": "TAG",
    "DESCRIPTION": "SERVICE",
    "SENSOR TYPE": "MEAS TYPE",
    "SCALE RANGE|MIN": "RANGE MIN",
    "SCALE RANGE|MAX": "RANGE MAX",
    "SCALE RANGE|UNIT": "UNIT",
    "OUTPUT SIGNAL": "SIGNAL",
    "MODEL": "MODEL",
    "MAKER": "MAKER",
    "LOCATION": "INST LOCATION",
    "REMARKS": "REMARK",
}

SIM_THRESHOLD = 0.7
NUM_RE = re.compile(r"\d{3,}")
# 6019A / 6019 처럼 번호+옵션접미
NUM_LETTER_RE = re.compile(r"(\d{3,})([A-Z]?)(?:~[A-Z0-9]+)?", re.I)
IL_EQUIP_RE = re.compile(r"^([A-Za-z0-9\-_/~]+)\s*:\s*(.+)$")
# F05-POL-A-UPWP-DI-6019A~F → F05-POL-A-UPWP-DI-6019A
RANGE_SUFFIX_RE = re.compile(r"~[A-Z0-9]+$", re.I)
SKIP_INST_TAGS = {
    "CALIBRATION", "PORTABLE", "TEST", "TAG NO.", "TAG", "-",
}


def _s(v):
    return str(v).strip() if v is not None else ""


def _h(v):
    """머리글 정규화 — 실물 양식은 머리 항목이 두 줄로 나뉘어 있다.

    'SENSOR\\nTYPE' 은 눈으로는 'SENSOR TYPE' 이지만 문자열로는 다르다.
    이것 때문에 INSTRUMENT_MAP 이 통째로 빗나가 SENSOR TYPE 을 한 건도
    읽지 못했고, TYPE 사양이 빈 채로 조회됐다. 줄바꿈·연속 공백을
    공백 하나로 접어서 맞춘다.
    """
    return re.sub(r"\s+", " ", _s(v))

def _is_spare(tag, desc=""):
    """SPARE / 빈 점 — 기능 태그 목록에서 제외."""
    t = _s(tag).upper()
    d = _s(desc).upper()
    if not t or t in ("-", "0"):
        return True
    if t == "SPARE" or t.startswith("SPARE") or "_SPARE" in t or t.endswith("_SP"):
        return True
    if d == "SPARE" or d.startswith("SPARE "):
        return True
    return False




def _pid_of(rec):
    pid = _s(rec.get("P&ID TAG")) or _s(rec.get("PID_TAG"))
    if pid in ("", "-", "0", "N/A"):
        pid = ""
    return pid or _s(rec.get("TAG"))


def _norm_text(s):
    s = str(s or "").strip().lower()
    s = re.sub(r"[^a-z0-9가-힣\s]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def _tag_alnum(s):
    return re.sub(r"[^A-Z0-9]", "", str(s or "").upper())


def _numbers(tag):
    return set(NUM_RE.findall(str(tag or "")))


def _num_letters(tag):
    """{'6019A','6019'} 형태 집합 — 펌프 A/B/C 구분용."""
    out = set()
    for m in NUM_LETTER_RE.finditer(str(tag or "").upper()):
        n, let = m.group(1), m.group(2) or ""
        out.add(n + let)
        out.add(n)
    return out


def normalize_pid(pid):
    """F05-...-6019A~F → F05-...-6019A (범위 접미 제거)."""
    pid = _s(pid)
    if not pid:
        return pid
    # A/B 복합 장비 표기 LCV-P1501/LCV-P1502 는 그대로 둠
    if "/" in pid and not pid.upper().startswith("F"):
        return pid
    return RANGE_SUFFIX_RE.sub("", pid)


def _range_letters(pid):
    """6019A~F / 6019A/B → 허용 접미 집합 {'A','B',...,'F'}."""
    s = _s(pid).upper()
    out = set()
    # 6019A~F
    m = re.search(r"(\d{3,})([A-Z])~([A-Z])\b", s)
    if m:
        a, b = ord(m.group(2)), ord(m.group(3))
        if a <= b:
            for o in range(a, b + 1):
                out.add(chr(o))
        return out
    # 6019A/B or trailing A/B on last segment
    m = re.search(r"(\d{3,})([A-Z])(?:/([A-Z]))+", s)
    if m:
        out.add(m.group(2))
        for g in re.findall(r"/([A-Z])", s[m.start():]):
            out.add(g)
        return out
    m = re.search(r"(\d{3,})([A-Z])\b", s)
    if m:
        out.add(m.group(2))
    return out


def _token_set(tag):
    return set(re.findall(r"[A-Z]{2,}", str(tag or "").upper()))


def _sim(a, b):
    na, nb = _norm_text(a), _norm_text(b)
    if not na or not nb:
        return 0.0
    return SequenceMatcher(None, na, nb).ratio()


def read_rows(path, sheet=None):
    if not path or not os.path.isfile(path):
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    hi = next((i for i, r in enumerate(rows)
               if r and "TAG" in [_s(c).upper() for c in r]), None)
    if hi is None:
        return []
    hdr = [_s(c).upper() for c in rows[hi]]
    ti = hdr.index("TAG")
    out = []
    for r in rows[hi + 1:]:
        if not r or ti >= len(r) or not r[ti]:
            continue
        tag = _s(r[ti])
        if tag in ("0", "-"):
            continue
        out.append({hdr[i]: (r[i] if i < len(r) else None)
                    for i in range(len(hdr)) if hdr[i]})
    return out


def _parse_instrument_sheet(rows):
    hi = next((i for i, r in enumerate(rows)
               if r and any(_h(c).upper() == "TAG NO." for c in r)), None)
    if hi is None:
        return []
    grp, cur = [], ""
    for c in rows[hi]:
        v = _h(c)
        if v:
            cur = v
        grp.append(cur)
    sub = [_h(c) for c in rows[hi + 1]] if hi + 1 < len(rows) else []
    names = []
    for i, g in enumerate(grp):
        sc = sub[i] if i < len(sub) else ""
        flat = "%s|%s" % (g, sc) if sc else g
        names.append(INSTRUMENT_MAP.get(flat, flat))
    if "TAG" not in names:
        return []
    ti = names.index("TAG")
    out = []
    for r in rows[hi + 2:]:
        if not r or ti >= len(r) or not r[ti]:
            continue
        tag = _s(r[ti])
        # 실물 셀은 폭 때문에 태그를 두 줄로 끊어 적는다 — 'AIT-P1801\nA~F'.
        # 사람 눈에는 한 태그지만 문자열로는 IO List 의 'AIT-P1801A~F' 와
        # 다르다. 380행 중 186행이 이 형태여서 조인이 절반 가까이 샜다.
        # 태그에 공백이 의미를 갖는 경우는 없으므로 전부 지운다.
        tag = re.sub(r"\s+", "", tag)
        if not tag or tag.upper() in SKIP_INST_TAGS:
            continue
        if not re.search(r"[A-Za-z]", tag):
            continue
        # 값 셀도 폭 때문에 줄바꿈이 들어간다 — MODEL 이 'ISE20C-Y-\nM-N02L'
        # 로 저장되면 매뉴얼 대조도, 화면 표시도 어긋난다. 공백 하나로 접는다.
        rec = {}
        for i in range(len(names)):
            if not names[i]:
                continue
            v = r[i] if i < len(r) else None
            rec[names[i]] = _h(v) if isinstance(v, str) else v
        rec["TAG"] = tag
        out.append(rec)
    return out


def _instrument_files(path):
    """파일 하나 · 파일 목록 · 폴더 어느 쪽이 와도 파일 목록으로 편다.

    실물 계기 리스트는 계기 종류별로 열몇 통으로 나뉘어 들어온다
    (Flow Transmitter · Level Switch · Pressure Gauge …). 한 통만 읽으면
    나머지 계기의 사양이 통째로 비고, 그 상태로도 조회는 성립하는 것처럼
    보여서 누락을 알아채기 어렵다.
    """
    if not path:
        return []
    if isinstance(path, (list, tuple, set)):
        cands = list(path)
    elif os.path.isdir(path):
        cands = [os.path.join(path, f) for f in sorted(os.listdir(path))
                 if re.search(r"instrument\s*list.*\.xlsx?$", f, re.I)
                 and not f.startswith("~$")]
    else:
        cands = [path]
    return [p for p in cands if p and os.path.isfile(p)]


def read_instrument_rows(path):
    """계기 리스트 — 파일·폴더·목록의 모든 시트에서 TAG NO. 수집."""
    files = _instrument_files(path)
    if not files:
        return []
    if len(files) > 1:
        out = []
        seen = set()
        for p in files:
            for rec in read_instrument_rows(p):
                tag = _s(rec.get("TAG")).upper()
                if not tag or tag in seen:
                    continue
                seen.add(tag)
                rec.setdefault("_file", os.path.basename(p))
                out.append(rec)
        return out
    path = files[0]
    wb = load_workbook(path, read_only=True, data_only=True)
    out = []
    seen = set()
    for sn in wb.sheetnames:
        try:
            rows = list(wb[sn].iter_rows(values_only=True))
        except Exception:
            continue
        for rec in _parse_instrument_sheet(rows):
            tag = _s(rec.get("TAG")).upper()
            if not tag or tag in seen:
                continue
            seen.add(tag)
            rec["_sheet"] = sn
            rec["_file"] = os.path.basename(path)
            out.append(rec)
    return out


def read_interlock_equip_candidates(path_or_dir):
    """Interlock 폴더/파일의 모든 시트에서 TAG : DESC 장비 헤더 추출."""
    files = []
    if not path_or_dir:
        return []
    if os.path.isdir(path_or_dir):
        for f in sorted(os.listdir(path_or_dir)):
            if f.lower().endswith((".xlsx", ".xlsm")) and not f.startswith("~$"):
                files.append(os.path.join(path_or_dir, f))
    elif os.path.isfile(path_or_dir):
        files = [path_or_dir]
    out = []
    seen = set()
    for path in files:
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
        except Exception:
            continue
        for sn in wb.sheetnames:
            try:
                rows = wb[sn].iter_rows(values_only=True)
            except Exception:
                continue
            for r in rows:
                if not r or not r[0]:
                    continue
                cell = _s(r[0])
                if cell.startswith(":"):
                    continue
                m = IL_EQUIP_RE.match(cell)
                if not m:
                    continue
                pid = m.group(1).strip()
                desc = re.sub(r"^[→\-\s]+", "", m.group(2).strip()).strip()
                key = pid.upper()
                if not pid or key in seen:
                    continue
                seen.add(key)
                letters = _range_letters(pid)
                out.append({"pid": normalize_pid(pid), "desc": desc, "src": "interlock",
                            "raw_pid": pid, "letters": letters,
                            "file": os.path.basename(path), "sheet": sn})
    return out


def _score_candidate(io_tag, io_desc, cand_pid, cand_desc, prefer_letter=True,
                     allowed_letters=None):
    """숫자 겹침 + 종류 토큰(공유 또는 IO태그 부분문자열) or 태그 포함."""
    io_nums = _numbers(io_tag)
    cand_nums = _numbers(cand_pid)
    if not (io_nums & cand_nums):
        return None

    io_nl = {x for x in _num_letters(io_tag) if len(x) > 3 and x[-1:].isalpha()}
    cand_nl = {x for x in _num_letters(cand_pid) if len(x) > 3 and x[-1:].isalpha()}
    # allowed_letters: 인터락 6019A~F → {A..F}
    if allowed_letters:
        io_lets = {x[-1] for x in io_nl}
        if io_lets and not (io_lets & set(allowed_letters)):
            return None
        # 범위 안이면 접미 충돌 검사 스킵
    elif io_nl and cand_nl and not (io_nl & cand_nl):
        return None

    score = _sim(io_desc, cand_desc)
    noise = {"TRAIN", "UNIT", "LOOP", "FROM", "WITH", "POU", "FAB", "THE", "AND",
             "LINE", "RETURN", "WATER", "TANK"}
    shared = (_token_set(io_tag) & _token_set(cand_pid)) - noise
    ia = _tag_alnum(io_tag)
    ca = _tag_alnum(cand_pid)
    # PCV ⊂ POUPCVC…, LCV ⊂ UPWTLCV…
    for tok in _token_set(cand_pid) - noise:
        if len(tok) >= 3 and tok in ia:
            shared.add(tok)
    for tok in _token_set(io_tag) - noise:
        if len(tok) >= 3 and tok in ca:
            shared.add(tok)
    # 설명에 후보 종류 토큰 (PCV VALVE …)
    da = _tag_alnum(io_desc)
    for tok in _token_set(cand_pid) - noise:
        if len(tok) >= 3 and tok in da:
            shared.add(tok)

    contained = bool(ca and len(ca) >= 5 and (ca in ia or ia.endswith(ca)))
    if contained:
        return 1.0 if ca == ia else 0.95

    if not shared:
        # 설명 매우 유사 + 숫자만으로 허용하지 않음 (오매칭 방지)
        return None

    score = max(score, 0.55 + 0.1 * min(3, len(shared)))
    if io_nl & cand_nl:
        score = max(score, 0.85)
    if score < 0.55:
        return None
    return score


def resolve_pid_tag(io_tag, io_desc, inst_cands, il_cands):
    def pick(pool):
        hits = {}
        for c in pool:
            raw = c.get("raw_pid") or c["pid"]
            # 범위 접미 허용 시 점수 함수의 A/B 거절을 우회하기 위해
            # 후보 pid 를 IO 접미에 맞게 가상 조정
            letters = c.get("letters") or _range_letters(raw)
            io_letters = {x[-1] for x in _num_letters(io_tag)
                          if len(x) > 3 and x[-1:].isalpha()}
            score_pid = c["pid"]
            if letters and io_letters and not (io_letters & letters):
                # 예: 후보 A~F, IO 는 B → 허용
                if not letters:
                    continue
            if letters and io_letters and (io_letters & letters):
                # 접미 충돌 검사 통과 강제: score 용 pid 에 IO 접미 반영
                pass
            sc = _score_candidate(io_tag, io_desc, score_pid, c["desc"],
                                  allowed_letters=letters or None)
            if sc is None:
                continue
            pid = normalize_pid(c["pid"])
            if pid not in hits or sc > hits[pid]:
                hits[pid] = sc
        if len(hits) == 1:
            pid = next(iter(hits))
            return pid, hits[pid]
        if hits:
            best = max(hits.values())
            top = [p for p, s in hits.items() if abs(s - best) < 1e-6]
            if len(top) == 1:
                return top[0], best
        return None, 0.0

    pid, sc = pick(inst_cands)
    if pid:
        return pid, "instrument", sc
    pid, sc = pick(il_cands)
    if pid:
        return pid, "interlock", sc
    return "", "", 0.0


# ── IO 태그 ↔ P&ID 태그 규칙 매핑 ──────────────────────────
#
# 두 문서는 이름 공간이 아예 다르다. IO List 는 PLC 심볼명
# (DWP_LS_P5401A_H), 계기 리스트는 P&ID 태그(LS-P5401A/B) 를 쓴다.
# 실물에서 두 목록의 교집합은 **0** 이다. 그래서 지금까지는 유사도
# 점수로 이어 붙였는데, 그 방식은 호기를 구별하지 못했다 —
# UPWT_LS_P3601C 가 LS-P3601A 에 붙는 식이다. C호기를 물었는데
# A호기 사양이 나오고, 화면에는 그럴듯하게 보인다.
#
# 여기서는 표기 규칙으로 후보를 만들고, **계기 리스트에 그 태그가
# 실재할 때만** 채택한다. 규칙이 제안하고 문서가 승인하는 구조라
# 없는 것을 지어내지 않는다. 규칙이 못 만들면 기존 유사도로 넘긴다.

_RANGE_RE = re.compile(r"^(.*?)([A-Z]\d?)\s*[~\-/]\s*([A-Z]\d?)$")
_PAIR_RE = re.compile(r"^(.*?)([A-Z])/([A-Z])$")
# 한계값 접미 — 같은 계기의 상/하한 접점이라 계기 자체는 하나다
_LIMIT_SUF_RE = re.compile(r"_(H{1,2}|L{1,2}|PLS)$", re.I)
_CORE_RE = re.compile(r"([A-Z]{1,5})_(P\d{3,5})([A-Z]\d?)?$", re.I)


def expand_range_tag(tag):
    """범위 표기를 구성원으로 편다.

    계기 리스트는 같은 사양의 계기를 한 줄에 묶어 적는다.
      LS-P5401A/B   → LS-P5401A, LS-P5401B
      LS-P3601A~C   → LS-P3601A, LS-P3601B, LS-P3601C
      FIS-P2801A1~A4 → FIS-P2801A1 … A4
    묶인 채로 두면 IO 의 개별 호기와 짝이 지어지지 않는다.
    원문도 함께 돌려준다 — 근거로 보여줄 때는 원문이 맞다.
    """
    t = _s(tag)
    if not t:
        return []
    out = [t]
    m = _RANGE_RE.match(t)
    if m:
        base, a, b = m.group(1), m.group(2), m.group(3)
        if len(a) == 1 and len(b) == 1 and a <= b:
            out += [base + chr(c) for c in range(ord(a), ord(b) + 1)]
        elif len(a) == 2 and len(b) == 2 and a[0] == b[0] and a[1] <= b[1]:
            out += [base + a[0] + str(n)
                    for n in range(int(a[1]), int(b[1]) + 1)]
        else:
            out += [base + a, base + b]
    m2 = _PAIR_RE.match(t)
    if m2:
        out += [m2.group(1) + m2.group(2), m2.group(1) + m2.group(3)]
    seen, uniq = set(), []
    for x in out:
        k = x.upper()
        if k not in seen:
            seen.add(k)
            uniq.append(x)
    return uniq


def build_spec_member_index(spec_rows):
    """구성원 태그 → 계기 리스트 원문 태그."""
    idx = {}
    for r in spec_rows:
        raw = _s(r.get("TAG"))
        if not raw:
            continue
        for m in expand_range_tag(raw):
            idx.setdefault(m.upper().replace(" ", ""), raw)
    return idx


def structural_pid_candidates(io_tag):
    """IO 심볼명에서 P&ID 태그 후보를 만든다 (있을 법한 표기만).

    DWP_LS_P5401A_H → LS-P5401A → (호기 없는 형태) LS-P5401
    한계 접미는 벗긴다 — H/L/LL 은 같은 계기의 접점이다.
    """
    t = _s(io_tag)
    for _ in range(3):
        t2 = _LIMIT_SUF_RE.sub("", t)
        if t2 == t:
            break
        t = t2
    m = _CORE_RE.search(t)
    if not m:
        return []
    typ = m.group(1).upper()
    num = m.group(2).upper()
    suf = (m.group(3) or "").upper()
    cands = []
    if suf:
        cands.append("%s-%s%s" % (typ, num, suf))
    cands.append("%s-%s" % (typ, num))
    return cands


def spec_prefixes(member_index):
    """계기 리스트에 실재하는 접두어 집합 (AIT · LS · PIT …).

    후보의 접두어를 여기에 가둔다. 가두지 않으면 FAB_MCC_A_XL_A_P2101
    같은 심볼에서 'A-P2101' 이 후보로 나오는데, 우연히 같은 이름의 계기가
    있으면 램프 접점이 계기 사양을 물고 온다.
    """
    out = set()
    for k in member_index:
        m = re.match(r"^([A-Z]+)-", k)
        if m:
            out.add(m.group(1))
    return out


def resolve_pid_by_rule(io_tag, member_index, prefixes=None):
    """규칙 매핑. 계기 리스트에 실재하는 후보만 돌려준다."""
    for c in structural_pid_candidates(io_tag):
        typ = c.split("-", 1)[0].upper()
        if prefixes is not None and typ not in prefixes:
            continue
        key = c.upper().replace(" ", "")
        if key in member_index:
            return c, member_index[key]
    return None, None


def _build_inst_cands(spec_rows):
    return [{
        "pid": _s(r.get("TAG")),
        "desc": _s(r.get("SERVICE") or r.get("DESCRIPTION")),
        "src": "instrument",
    } for r in spec_rows if _s(r.get("TAG"))]


def load_points(io_path, spec_path=None, attr_path=None, tb_path=None,
                interlock_path=None):
    """IO TAG → 배선 레코드. P&ID 비어 있으면 자동 부여."""
    io_rows = read_rows(io_path)
    spec_rows = read_instrument_rows(spec_path) if spec_path else []
    spec = {_s(r.get("TAG")).upper(): r for r in spec_rows if _s(r.get("TAG"))}
    attrs = {_s(r.get("TAG")): r for r in read_rows(attr_path)} if attr_path else {}

    if interlock_path is None:
        try:
            import config as _cfg
            interlock_path = getattr(_cfg, "INTERLOCK_XLSX", None) or \
                getattr(_cfg, "INTERLOCK_DIR", None)
        except Exception:
            interlock_path = None

    inst_cands = _build_inst_cands(spec_rows)
    # 범위 표기를 편 구성원 색인 — 규칙 매핑과 사양 조인 양쪽에 쓴다.
    member_index = build_spec_member_index(spec_rows)
    member_prefixes = spec_prefixes(member_index)
    spec_by_member = {}
    for r in spec_rows:
        raw = _s(r.get("TAG"))
        if not raw:
            continue
        for m in expand_range_tag(raw):
            spec_by_member.setdefault(m.upper().replace(" ", ""), r)
    il_cands = read_interlock_equip_candidates(interlock_path)

    tb = {}
    if tb_path:
        try:
            from ingest.tb_list import load_terminals
            tb = load_terminals(tb_path)
        except Exception as e:
            print("[lists] TB 리스트를 읽지 못했습니다:", e)

    auto_filled = 0
    out = {}
    for r in io_rows:
        io_tag = _s(r.get("TAG"))
        if not io_tag:
            continue
        desc0 = _s(r.get("DESCRIPTION") or r.get("SERVICE"))
        rec = dict(r)
        rec["IO TAG"] = io_tag
        rec["_spare"] = _is_spare(io_tag, desc0)
        pid = _s(rec.get("P&ID TAG"))
        if pid in ("", "-", "0", "N/A"):
            pid = ""
        if not pid:
            desc = _s(rec.get("DESCRIPTION") or rec.get("SERVICE"))
            # 규칙이 먼저다. 표기 규칙으로 만든 후보가 계기 리스트에
            # 실재하면 그것을 쓴다 — 호기까지 정확히 짚는다.
            pid, raw_tag = resolve_pid_by_rule(io_tag, member_index,
                                               member_prefixes)
            if pid:
                rec["P&ID TAG"] = pid
                rec["_pid_src"] = "rule"
                rec["_pid_score"] = 1.0
                if raw_tag and raw_tag.upper() != pid.upper():
                    # 근거로 보여줄 때는 계기 리스트 원문 표기가 맞다
                    rec["_pid_source_tag"] = raw_tag
                auto_filled += 1
            else:
                pid, src, sc = resolve_pid_tag(io_tag, desc, inst_cands,
                                               il_cands)
                if pid:
                    rec["P&ID TAG"] = pid
                    rec["_pid_src"] = src
                    rec["_pid_score"] = sc
                    auto_filled += 1
                else:
                    rec["P&ID TAG"] = io_tag
                    rec["_pid_src"] = "fallback_io"
        else:
            rec["P&ID TAG"] = pid

        t = tb.get(io_tag)
        if t:
            rec["TERMINAL"] = t.get("TERMINAL") or rec.get("TERMINAL")
            rec["TB"] = t.get("TB")

        # 구성원 태그(LS-P5401A)로 이어졌으면 원문 줄(LS-P5401A/B)에서
        # 사양을 가져와야 한다. 구성원 색인을 먼저 보고, 없으면 원문 키.
        join_key = _s(rec.get("P&ID TAG")).upper()
        _sp = spec_by_member.get(join_key.replace(" ", "")) or spec.get(join_key)
        for src in ((_sp or {}), (attrs.get(join_key) or {})):
            for k, v in src.items():
                if k in ("TAG", "IO TAG", "_sheet"):
                    continue
                if k not in rec or rec.get(k) in (None, ""):
                    rec[k] = v
        sp = _sp or {}
        if sp.get("MEAS TYPE"):
            rec["TYPE"] = sp["MEAS TYPE"]
        _apply_aliases(rec)
        out[io_tag] = rec

    for key, r in spec.items():
        if any(_s(x.get("P&ID TAG")).upper() == key for x in out.values()):
            continue
        if key in out:
            continue
        rec = dict(r)
        rec["TAG"] = _s(r.get("TAG"))
        rec["IO TAG"] = ""
        rec["P&ID TAG"] = _s(r.get("TAG"))
        # IO 점이 없는 계기(로컬 게이지 등)도 TYPE 은 계기 리스트에서 온다.
        # IO 행 쪽에만 이 줄이 있어서, 배선이 없는 계기는 TYPE 이 빈 채로
        # 조회됐다 — 사양 출처 검사가 AIT-P4112 에서 이것을 잡았다.
        if _s(r.get("MEAS TYPE")):
            rec["TYPE"] = _s(r.get("MEAS TYPE"))
        for k, v in (attrs.get(key) or {}).items():
            rec.setdefault(k, v)
        _apply_aliases(rec)
        out[rec["P&ID TAG"]] = rec

    try:
        from ingest.manual_match import manual_for
        manual_for(out)
    except Exception as e:
        print("[lists] 매뉴얼 대조 실패:", e)

    print("[lists] IO %d점 / Instrument 후보 %d / Interlock 후보 %d / P&ID 자동부여 %d"
          % (len(out), len(inst_cands), len(il_cands), auto_filled))
    return out


def build_pid_index(points):
    by_pid = OrderedDict()
    for io_tag, rec in points.items():
        if rec.get("_spare") or _is_spare(io_tag, rec.get("DESCRIPTION") or rec.get("SERVICE")):
            continue
        pid = _pid_of(rec)
        if not pid:
            continue
        if pid not in by_pid:
            base = dict(rec)
            base["TAG"] = pid
            base["P&ID TAG"] = pid
            base["IO TAG"] = _s(rec.get("IO TAG") or rec.get("TAG"))
            base["io_tags"] = []
            base["io_points"] = []
            by_pid[pid] = base
        dev = by_pid[pid]
        io = _s(rec.get("IO TAG") or io_tag)
        if io and io not in dev["io_tags"]:
            dev["io_tags"].append(io)
        dev["io_points"].append({
            "io_tag": io,
            "io_type": _s(rec.get("IO TYPE")),
            "panel": _s(rec.get("PANEL") or rec.get("PANEL NO")),
            "rack": rec.get("RACK"),
            "slot": rec.get("SLOT"),
            "ch": rec.get("CH"),
            "description": _s(rec.get("DESCRIPTION") or rec.get("SERVICE")),
        })
        for k in ("PANEL", "PANEL NO", "LOCATION", "PN(DP)", "RACK", "SLOT"):
            if not dev.get(k) and rec.get(k) not in (None, ""):
                dev[k] = rec.get(k)
        for k in ("MAKER", "MODEL", "TYPE", "MEAS TYPE", "SERVICE",
                  "RANGE MIN", "RANGE MAX", "UNIT", "SIGNAL", "MANUAL FILE"):
            if not dev.get(k) and rec.get(k) not in (None, ""):
                dev[k] = rec.get(k)
        _apply_aliases(dev)
    return dict(by_pid)


def load_pid_devices(io_path, spec_path=None, attr_path=None, tb_path=None,
                     interlock_path=None):
    return build_pid_index(
        load_points(io_path, spec_path, attr_path, tb_path, interlock_path))


def _apply_aliases(rec):
    for want, std in COLUMN_ALIASES.items():
        if not rec.get(want) and rec.get(std):
            rec[want] = rec[std]
    return rec

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
interlock.py — 인터락 리스트 엑셀 파서

실제 서류는 깨끗하지 않다. 두 가지를 처리해야 한다.

  1) 병합/공란 전방채움
     한 인터락 항목이 여러 행에 걸치고, 항목 정보는 첫 행에만 있다.
     아래로 채워 넣지 않으면 조건 행이 고아가 된다.

  2) 자연어 조건 구조화
     "LIT-4003 High High (95 %) 3sec delay" 같은 문장을
     {tag, level, op, setpoint, unit, delay} 로 쪼갠다.

2번은 규칙으로 대부분 잡히지만 전부는 아니다. 못 잡은 것은 parsed=False
로 남기고 raw 를 그대로 보존한다. **추측해서 채우지 않는다** — 인터락은
안전 로직이라 잘못 구조화한 조건이 맞게 구조화한 조건보다 위험하다.

미파싱 조건은 LLM 으로 보강할 수 있으나, 그것은 오프라인 전처리 단계이며
사람 검수를 전제로 한다 (llm_assist.py). 런타임에는 절대 생성하지 않는다.

사용:
    python -m ingest.interlock --stats
    python -m ingest.interlock --unparsed
"""

import argparse
import os
import re
import sys

from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import config  # noqa: E402

# 항목 정보 열 — 공란이면 위에서 채운다
CARRY = ["IL NO", "OUTPUT TAG", "ACTION", "KIND", "LOGIC", "RESET",
         "BYPASS", "PRIORITY", "PLC BLOCK", "DWG No.", "SHEET", "REMARK"]

TAG_RE = re.compile(r"\b([A-Z]{1,4}-\d{3,4}[A-Z]?)\b")

# 알람 레벨 표기 → 정규화
LEVEL = [
    (r"high\s*high|hi\s*hi|\bHH\b|고고", "HH"),
    (r"low\s*low|lo\s*lo|\bLL\b|저저", "LL"),
    (r"\bhigh\b|\bhi\b|고압|고온", "H"),
    (r"\blow\b|\blo\b|저압|저온", "L"),
]

OP_RE = re.compile(r"(>=|<=|=>|=<|>|<|이상|이하|초과|미만)")
OP_MAP = {">=": ">=", "=>": ">=", "<=": "<=", "=<": "<=", ">": ">", "<": "<",
          "이상": ">=", "이하": "<=", "초과": ">", "미만": "<"}

NUM_UNIT_RE = re.compile(
    r"(-?\d+(?:\.\d+)?)\s*"
    r"(%|ppb|ppm|bar|kPa|MPa|degC|℃|m3/h|m³/h|uS/cm|µS/cm|Mohm-cm|MΩ-cm|A|V|Hz)?",
    re.I)

DELAY_RE = re.compile(
    r"(\d+(?:\.\d+)?)\s*(?:sec|s\b|초)\s*(?:delay|지연|연속)?"
    r"|(?:연속|지연)\s*(\d+(?:\.\d+)?)\s*(?:sec|s\b|초)", re.I)

# 상태형 조건 (아날로그 비교가 아닌 것)
STATE_PATTERNS = [
    # E-STOP 이 \bSTOP\b 에 먼저 잡히면 안 되므로 순서가 중요하다.
    (r"E-?STOP|비상\s*정지|EMERGENCY\s*STOP", "ESTOP"),
    (r"운전\s*중|\bRUN\b", "RUN"),
    (r"정지(?!\s*방지)|\bSTOP\b", "STOP"),
    (r"\bOPEN\b|열림", "OPEN"),
    (r"\bCLOSED?\b|닫힘", "CLOSE"),
    (r"\bREADY\b|준비\s*완료", "READY"),
    (r"(?:FAULT|TRIP|OVERLOAD).{0,6}(?:없음|정상|해제)", "NO_FAULT"),
    (r"\bFAULT\b|\bTRIP\b|OVERLOAD", "FAULT"),
    (r"STEP\s*\d+", "SEQ_STEP"),
    (r"수동\s*모드|\bMANUAL\b", "MANUAL_MODE"),
    (r"\bDEMAND\b|수요\s*신호", "DEMAND"),
    (r"키스위치|KEY\s*SW", "KEYSWITCH"),
]

MULTI_AND = re.compile(r"및|그리고|\band\b", re.I)
MULTI_OR = re.compile(r"또는|\bor\b", re.I)


def norm(s):
    return re.sub(r"\s+", " ", str(s or "")).strip()


def parse_condition(text):
    """
    자연어 조건 → 구조화. 확신 없는 항목은 parsed=False 로 남긴다.
    """
    raw = norm(text)
    low = raw.lower()
    out = {"raw": raw, "tags": [], "level": None, "op": None,
           "setpoint": None, "unit": None, "delay_sec": None,
           "state": None, "multi": None, "parsed": False}
    if not raw:
        return out

    out["tags"] = TAG_RE.findall(raw.upper())
    if len(out["tags"]) > 1:
        out["multi"] = "OR" if MULTI_OR.search(raw) else (
            "AND" if MULTI_AND.search(raw) else None)

    for pat, lv in LEVEL:
        if re.search(pat, low, re.I):
            out["level"] = lv
            break

    m = DELAY_RE.search(raw)
    if m:
        out["delay_sec"] = float(m.group(1) or m.group(2))

    m = OP_RE.search(raw)
    if m:
        out["op"] = OP_MAP[m.group(1)]
        rest = raw[m.end():]
    else:
        # "High High (95 %)" 처럼 괄호 안에 설정값만 있는 형태
        rest = raw[raw.find("(") + 1:] if "(" in raw else ""
        if out["level"]:
            out["op"] = ">=" if out["level"] in ("H", "HH") else "<="

    if rest:
        m2 = NUM_UNIT_RE.search(rest)
        if m2 and m2.group(1) is not None:
            # 지연 시간 숫자를 설정값으로 오인하지 않도록 단위를 확인한다
            unit = (m2.group(2) or "").strip()
            if unit or out["op"]:
                out["setpoint"] = float(m2.group(1))
                out["unit"] = unit or None

    for pat, st in STATE_PATTERNS:
        if re.search(pat, raw, re.I):
            out["state"] = st
            break

    # 파싱 성공 판정 — 둘 중 하나는 확정되어야 한다
    analog_ok = out["setpoint"] is not None and out["op"] is not None
    discrete_ok = out["state"] is not None
    out["parsed"] = bool(analog_ok or discrete_ok)
    out["kind"] = "ANALOG" if analog_ok else ("DISCRETE" if discrete_ok else "UNKNOWN")
    return out


_VALVE_ACTION_RE = re.compile(
    r"valve\s*action\s*[:：]\s*([A-Za-z ]+)", re.I)


def _valve_action(remark):
    """REMARK 에서 "* Valve Action : Fail Open" 을 읽는다. 없으면 빈 값."""
    m = _VALVE_ACTION_RE.search(remark or "")
    return m.group(1).strip().upper() if m else ""


def _load_interlocks_demo(path):
    """IL NO 표 형식 (데모) 파서."""
    ws = load_workbook(path, read_only=True, data_only=True).active
    rows = list(ws.iter_rows(values_only=True))
    hi = next(i for i, r in enumerate(rows)
              if r and "IL NO" in [norm(c).upper() for c in r])
    hdr = [norm(c) for c in rows[hi]]
    I = {h: i for i, h in enumerate(hdr) if h}

    carried, items, order = {}, {}, []
    for r in rows[hi + 1:]:
        if not r or all(c is None or norm(c) == "" for c in r):
            continue
        # 전방채움
        for col in CARRY:
            if col in I:
                v = norm(r[I[col]])
                if v:
                    carried[col] = v
        cond_text = norm(r[I["CONDITION"]]) if "CONDITION" in I else ""
        if not cond_text:
            continue

        ilno = carried.get("IL NO")
        if ilno not in items:
            items[ilno] = {
                "il_no": ilno,
                "output_tag": carried.get("OUTPUT TAG"),
                "action": carried.get("ACTION"),
                "kind": carried.get("KIND"),
                "logic": carried.get("LOGIC", "AND"),
                "reset": carried.get("RESET"),
                "bypassable": carried.get("BYPASS") == "Y",
                "priority": int(carried.get("PRIORITY") or 9),
                "plc_block": carried.get("PLC BLOCK"),
                "dwg_no": carried.get("DWG No."),
                "sheet": carried.get("SHEET"),
                "remark": carried.get("REMARK", ""),
                # 고장 위치는 REMARK 의 "* Valve Action : Fail Open" 에서
                # 읽는다. 실물 양식이 그 자리에 적기 때문이다 — 별도 열을
                # 만들면 실물과 모양이 달라진다.
                "fail": _valve_action(carried.get("REMARK", "")),
                "conditions": [],
            }
            order.append(ilno)
        grp = norm(r[I["GROUP"]]) if "GROUP" in I else "G1"
        c = parse_condition(cond_text)
        c["group"] = grp or "G1"
        items[ilno]["conditions"].append(c)

    return [items[k] for k in order]


def load_interlocks(path=None):
    """데모 표 + 실물 양식을 합친다. 한쪽 실패해도 다른 쪽은 살린다."""
    items = []

    def _targets():
        """읽을 파일 목록. 폴더면 그 안의 엑셀 전부."""
        base = path or config.INTERLOCK_XLSX
        if os.path.isdir(base):
            return sorted(os.path.join(base, f) for f in os.listdir(base)
                          if f.lower().endswith((".xlsx", ".xlsm"))
                          and not f.startswith("~$"))
        return [base] if os.path.isfile(base) else []

    # 파일 이름을 여기서 정하지 않는다. 이름이 바뀌면 조회가 통째로 비는데,
    # 인터락 0건은 "인터락이 없다" 로 보여서 원인을 찾기 어렵다. 실제로
    # 파일명을 정리하다 그렇게 됐다. 경로는 config 한 곳에서 온다.
    for _p in _targets():
        try:
            from ingest.interlock_real import (detect_real_format,
                                               load_interlocks_real)
            if detect_real_format(_p):
                items.extend(load_interlocks_real(_p))
            else:
                items.extend(_load_interlocks_demo(_p))
        except Exception as e:                              # noqa: BLE001
            print("인터락 리스트 로드 실패(%s): %s"
                  % (os.path.basename(_p), e))
    return items


def _load_interlocks_legacy(path=None):
    items = []
    demo_path = config.INTERLOCK_XLSX
    real_path = os.path.join(config.DATA_DIR, "INTERLOCK_SAMPLE.xlsx")
    # 명시 path 가 있으면 그것만
    if path:
        try:
            from ingest.interlock_real import detect_real_format, load_interlocks_real
            if detect_real_format(path):
                return load_interlocks_real(path)
        except Exception as e:
            print("실물 파서 실패: %s" % e)
        try:
            return _load_interlocks_demo(path)
        except Exception as e:
            print("데모 파서 실패: %s" % e)
            return []
    # 기본: 둘 다 로드 후 병합
    if os.path.exists(demo_path):
        try:
            from ingest.interlock_real import (detect_real_format,
                                               load_interlocks_real)
            if detect_real_format(demo_path):
                items.extend(load_interlocks_real(demo_path))
            else:
                items.extend(_load_interlocks_demo(demo_path))
        except Exception as e:
            print("인터락 리스트 로드 실패(%s): %s"
                  % (os.path.basename(demo_path), e))
    if os.path.exists(real_path):
        try:
            from ingest.interlock_real import load_interlocks_real
            items.extend(load_interlocks_real(real_path))
        except Exception as e:
            print("실물 인터락 로드 실패: %s" % e)
    # 환경변수로 지정된 추가 파일
    env_path = os.environ.get("COPILOT_INTERLOCK_XLSX")
    if env_path and os.path.exists(env_path):
        env_abs = os.path.abspath(env_path)
        known = {os.path.abspath(demo_path), os.path.abspath(real_path)}
        if env_abs not in known:
            try:
                from ingest.interlock_real import detect_real_format, load_interlocks_real
                if detect_real_format(env_path):
                    items.extend(load_interlocks_real(env_path))
                else:
                    items.extend(_load_interlocks_demo(env_path))
            except Exception as e:
                print("추가 인터락 로드 실패: %s" % e)
    return items


def main():
    ap = argparse.ArgumentParser(description="인터락 리스트 파서")
    ap.add_argument("--stats", action="store_true")
    ap.add_argument("--unparsed", action="store_true")
    args = ap.parse_args()

    items = load_interlocks()
    conds = [c for it in items for c in it["conditions"]]
    ok = [c for c in conds if c["parsed"]]
    print("인터락 항목 %d건 / 조건 %d건" % (len(items), len(conds)))
    print("구조화 성공 %d건 (%.0f%%)  실패 %d건"
          % (len(ok), 100 * len(ok) / max(1, len(conds)), len(conds) - len(ok)))

    if args.stats:
        from collections import Counter
        print("\n조건 유형:", dict(Counter(c["kind"] for c in conds)))
        print("인터락 종류:", dict(Counter(it["kind"] for it in items)))
        print("출력 태그:", sorted({it["output_tag"] for it in items}))

    if args.unparsed:
        print("\n미파싱 조건:")
        for c in conds:
            if not c["parsed"]:
                print("  ·", c["raw"])


if __name__ == "__main__":
    main()

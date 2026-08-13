#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
tb_list.py — mastertool TB 결선표(격자 배치) 리더

TB 리스트는 평면 표가 아니라 **판넬 정면을 흉내 낸 격자 배치도**다.
시트 하나가 판넬 하나이고, TB 블록이 가로·세로로 깔린다.

    ■ PANEL : CUB_1
    ┌ DI-32 (32CH TB) ┐   ┌ DI-32 (32CH TB) ┐   ┌ AI-4 (4CH TB) ┐
    │ TB 1-1        - │   │ TB 1-4        - │   │ TB 2-1      - │
    │ N032  I32.0  태그│   │ N042  I56.0  태그│   │  PIW682+ …    │
    │ N032  I32.1  태그│   …                    │  PIW682-      │
    └─────────────────┘

블록 하나가 다섯 열을 쓰고, 그 안의 열 구실이 카드 종류에 따라 다르다.

    디지털(DI/DO)   [단자블록] [주소 I32.0] [   ] [태그]
    아날로그(AI/AO) [        ] [단자 PIW682+] [주소 PIW682] [태그]
                    아날로그는 한 채널이 +/- 두 행을 쓰고 태그는 + 행에만 있다

그래서 열 위치를 고정으로 잡으면 카드 종류에 따라 어긋난다. **행에서
주소처럼 생긴 값과 태그처럼 생긴 값을 찾아** 잡는다.

여기서 얻는 것은 태그 → (판넬, TB 위치, 단자, PLC 주소) 다. copilot 은
그중 **단자(TERMINAL)** 를 쓴다 — OUTLINE 도면에서 단자대를 하이라이트할
때 필요하다.

    python -m ingest.tb_list [파일]
"""

import os
import re
import sys
from collections import defaultdict

from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import config  # noqa: E402

# 블록 머리 표기가 두 가지다.
#   "DI-32  (32CH TB)"   하이픈·설명 붙은 형태
#   "DI32" / "AI4-4W"    붙여 쓴 형태 (결선 방식이 뒤에 붙기도)
BLOCK_RE = re.compile(r"^(DI|DO|AI|AO)[-\s]?(\d+)")
TB_RE = re.compile(r"^TB\s")
PANEL_RE = re.compile(r"PANEL\s*[:：]\s*(.+)")
# 디지털 주소 I32.0 / Q40.7, 아날로그 주소 PIW682 / PQW300
ADDR_RE = re.compile(r"^(?:[IQ]\d+\.\d+|P?[IQ]W\d+)$", re.I)
# 단자 표기: 아날로그는 주소 뒤에 +/- 가 붙는다
TERM_RE = re.compile(
    r"^(?:P?[IQ]W\d+[+-]|[A-Z]{1,3}\d{2,4}[A-Z]?|TB\d+-\d+)$", re.I)
# 카드 위치 표기 (예: D03-S01) — 어느 랙·슬롯의 카드인지
CARD_REF_RE = re.compile(r"^[A-Z]?\d{1,3}-S\d{1,3}$", re.I)

BLOCK_WIDTH = 5          # 블록 하나가 차지하는 열 수


def _s(v):
    return str(v).strip() if v is not None else ""


def _clean_name(v):
    """
    판넬명 정리.

    생성기가 파이썬 튜플을 그대로 문자열로 쓴 경우가 있다 —
    "('LCP-01',)" · "('CPU 315-2',)". 그대로 두면 화면과 조회 키에 괄호가
    섞인다. 따옴표 안의 값만 꺼낸다.
    """
    t = _s(v)
    m = re.match(r"^\(\s*['\"](.+?)['\"]\s*,?\s*\)$", t)
    return m.group(1).strip() if m else t


def _looks_like_tag(v):
    """
    태그처럼 생겼는가.

    주소·단자 표기와 구분만 하면 된다. 태그 명명 규칙은 프로젝트마다
    다르므로 형태를 좁게 규정하지 않는다 — 좁히면 실물에서 놓친다.
    """
    if not v or len(v) < 3:
        return False
    if ADDR_RE.match(v) or TERM_RE.match(v):
        return False
    return bool(re.search(r"[A-Za-z]", v))


def read_tb_layout(path=None):
    """
    TB 결선표 → 행 목록.

    각 행: TAG · PANEL · TB · TERMINAL · ADDRESS · CARD · SHEET
    """
    path = path or getattr(config, "TB_LIST", "")
    if not path or not os.path.isfile(path):
        return []
    wb = load_workbook(path, read_only=False, data_only=True)

    out = []
    spare = 0
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        panel = _clean_name(sheet)
        for c in range(1, min(ws.max_column, 40) + 1):
            m = PANEL_RE.search(_s(ws.cell(1, c).value))
            if m:
                panel = _clean_name(m.group(1))
                break

        # 블록 머리 찾기
        heads = []
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                v = _s(ws.cell(r, c).value)
                if BLOCK_RE.match(v):
                    # TB 라벨은 다음 행에 있다. 블록 폭 안 어디에 오는지는
                    # 양식에 따라 다르므로 그 범위를 훑는다.
                    tb, ref = "", ""
                    for cc in range(c, c + BLOCK_WIDTH):
                        val = _s(ws.cell(r + 1, cc).value)
                        if TB_RE.match(val):
                            tb = val
                        elif CARD_REF_RE.match(val):
                            ref = val          # 카드 위치 (예: D03-S01)
                    if not tb:
                        # TB 라벨이 없으면 블록으로 보지 않는다
                        continue
                    heads.append((r, c, v, tb, ref))

        # 블록별로 채널 행을 읽는다. 다음 블록 머리(같은 열)까지가 범위.
        by_col = defaultdict(list)
        for r, c, kind, tb, ref in heads:
            by_col[c].append((r, kind, tb, ref))
        for c, items in by_col.items():
            items.sort()
            for i, (r, kind, tb, ref) in enumerate(items):
                stop = items[i + 1][0] if i + 1 < len(items) else ws.max_row + 1
                for rr in range(r + 2, stop):
                    vals = [_s(ws.cell(rr, cc).value)
                            for cc in range(c, c + BLOCK_WIDTH)]
                    if not any(vals):
                        continue
                    addr = next((v for v in vals if ADDR_RE.match(v)), "")
                    # 단자·주소를 먼저 빼고 남은 것에서 태그를 고른다.
                    #
                    # 뒤에서부터 "태그처럼 생긴 것" 을 찾으면 단자 번호를
                    # 태그로 잡는다 — TB9-1 은 단자인데 태그 형태와
                    # 구분되지 않는다. 실제로 그렇게 잘못 읽혔다.
                    used = {addr}
                    # 단자는 두 가지가 섞여 있다.
                    #   현장 단자대 번호   TB9-1 · TB2-37
                    #   카드 단자(주소측)  IW510+ · N011
                    # 도면 하이라이트에 쓰는 것은 앞쪽이므로 그것을
                    # 우선한다. 없으면 카드 단자를 쓴다.
                    term = next((v for v in vals
                                 if v and v not in used
                                 and re.match(r"^TB\d+-\d+$", v, re.I)), "")
                    if not term:
                        term = next((v for v in vals
                                     if v and v not in used
                                     and TERM_RE.match(v)), "")
                    if term:
                        used.add(term)
                    tag = next((v for v in reversed(vals)
                                if v not in used and _looks_like_tag(v)), "")
                    if not tag:
                        continue
                    # 미사용 채널은 주소에서 딴 자리표시자가 태그 칸에
                    # 들어간다 (I32.3 → I32_3). 실제 태그가 아니므로
                    # 세지 않는다 — 세면 단자 조회가 없는 계기를 있다고
                    # 답한다.
                    if addr and tag.upper() == addr.upper().replace(".", "_"):
                        spare += 1
                        continue
                    out.append({
                        "TAG": tag,
                        "PANEL": panel,
                        "TB": tb,
                        "TERMINAL": term or tb,
                        "ADDRESS": addr,
                        "CARD": kind,
                        "CARD REF": ref,
                        "SHEET": sheet,
                    })
    if spare:
        print("[tb] 미사용 채널 %d점 제외 (주소 자리표시자)" % spare)
    return out


def load_terminals(path=None):
    """태그 → 단자 정보. 같은 태그가 여러 번 나오면 첫 번째를 쓴다."""
    out = {}
    dups = []
    for r in read_tb_layout(path):
        t = r["TAG"]
        if t in out:
            dups.append(t)
            continue
        out[t] = r
    if dups:
        print("[tb] 중복 태그 %d건 (첫 항목 사용): %s"
              % (len(dups), ", ".join(sorted(set(dups))[:5])))
    return out


def render(path=None):
    rows = read_tb_layout(path)
    if not rows:
        return "TB 리스트를 읽지 못했습니다: %s" % (path or config.TB_LIST)
    by_panel = defaultdict(list)
    by_card = defaultdict(int)
    for r in rows:
        by_panel[r["PANEL"]].append(r)
        by_card[r["CARD"]] += 1
    L = ["=" * 76, "TB 결선표", "=" * 76,
         "태그 %d점 · 판넬 %d개" % (len(rows), len(by_panel)),
         "카드 종류: " + ", ".join("%s %d" % (k, v)
                                   for k, v in sorted(by_card.items())), ""]
    for p, items in sorted(by_panel.items()):
        tbs = sorted({i["TB"] for i in items})
        L.append("■ %s — %d점 / TB %d개" % (p, len(items), len(tbs)))
        for i in items[:4]:
            L.append("   %-28s %-8s %-10s %s"
                     % (i["TAG"], i["TB"], i["TERMINAL"], i["ADDRESS"]))
        if len(items) > 4:
            L.append("   … 외 %d점" % (len(items) - 4))
    L.append("=" * 76)
    return "\n".join(L)


if __name__ == "__main__":
    print(render(sys.argv[1] if len(sys.argv) > 1 else None))

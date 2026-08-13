# -*- coding: utf-8 -*-
"""
interlock_real.py — 실물 INTERLOCK & ALARM LIST 양식 파서

양식 특징
  · 장비 블록:  "LCV-01 : Raw Water Pond Level Control Valve"
  · 동작 헤더:  "CLOSE BY" / "STOP BY" / "OPEN BY" / "START BY"
  · 조건 행:    번호 | SET CONDITION | (상태문구) | RESET | DELAY | SOURCE | NOTE
  · REMARK:     P&ID PAGE, Fail 위치 등

내부 스키마는 demo 리스트와 동일하게 맞춘다
  → InterlockIndex / UI 를 그대로 재사용.

사용:
    python -m ingest.interlock_real --stats
    python -m ingest.interlock_real --dump LCV-01
"""

from __future__ import annotations

import argparse
import os
import re
import sys

from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import config  # noqa: E402
from ingest.interlock import parse_condition, norm  # noqa: E402

# 장비 헤더: TAG : description
EQUIP_RE = re.compile(
    # 실물: LCV-P1101, TCV-P1201A/B, C01-CS-A-CDP-NAOCL-6025A/B …
    r"^\s*([A-Z0-9][A-Z0-9\-_/~]{1,48})\s*[:：]\s*(.+?)\s*$", re.I)

# 동작 헤더
ACTION_RE = re.compile(
    r"\b(CLOSE|OPEN|STOP|START|ON|OFF)\s+BY\b", re.I)

# 태그 (LIT-01A, LCV-01, DMF-RW-01A, RWP-01 …)
TAG_RE = re.compile(
    r"\b([A-Z]{1,8}(?:-[A-Z]{1,6})?-[A-Z0-9]{1,8}(?:~[A-Z0-9]+)?(?:/[A-Z0-9]+)?)\b",
    re.I)

DELAY_RE = re.compile(r"t\s*\(\s*(\d+(?:\.\d+)?)\s*s?\s*\)|(\d+(?:\.\d+)?)\s*s(?:ec)?", re.I)

PID_RE = re.compile(r"P\s*&\s*I\s*D\s*PAGE[-\s]*(\d+)|PAGE[-\s]*(\d+)", re.I)

# 구역 제목 — 태그가 아닌 블록 머리
#
# 실물 리스트에는 "6th Block Instrument & Analyzer Alarm List" 처럼 **출력
# 장비가 없는 경보 전용 구역**이 섞여 있다. 이 행은 TAG : DESC 형태가
# 아니어서 EQUIP_RE 에 걸리지 않았고, 그래서 블록이 끊기지 않은 채 그 아래
# 18건이 통째로 바로 앞 장비(폐수 펌프)에 붙었다. 펌프 정지 조건을 물으면
# 무관한 계기 경보가 근거로 따라 나오는 오귀속이다.
#
# 판정은 문구가 아니라 구조로 한다 — 셀 하나만 채워진 행이고, 그 아래 몇
# 줄 안에 동작 헤더(STOP BY …)나 SET CONDITION 헤더가 오면 블록 머리다.
# 문구로 판정하면 프로젝트마다 제목이 달라 다시 샌다.
SECTION_SKIP_RE = re.compile(r"^(remark|memo|date|document|no\.?|page)\b", re.I)


def _looks_like_section_head(rows, i):
    """태그가 아닌 블록 머리면 그 제목을, 아니면 None 을 돌려준다."""
    row = rows[i]
    vals = [v for v in (_cell(row, c) for c in range(len(row or ()))) if v]
    if len(vals) != 1:
        return None
    t = vals[0].strip()
    if len(t) < 4 or t.isdigit() or t == "-":
        return None
    if SECTION_SKIP_RE.match(t) or t.rstrip().endswith(":"):
        return None
    if EQUIP_RE.match(norm(t)):
        return None
    seen = 0
    for j in range(i + 1, min(len(rows), i + 8)):
        if _is_empty_row(rows[j]):
            continue
        # 진짜 장비 헤더가 먼저 오면 이 행은 문서 제목일 뿐이다
        if any(EQUIP_RE.match(norm(c) or "") for c in (rows[j] or [])[:5] if c):
            return None
        txt = _row_text(rows[j]).upper()
        if ACTION_RE.search(txt) or "SET CONDITION" in txt:
            return t
        seen += 1
        if seen >= 3:
            break
    return None


def _cell(row, idx):
    if idx is None or idx >= len(row):
        return ""
    return norm(row[idx])


def _is_empty_row(row):
    return not row or all(c is None or norm(c) == "" for c in row)


def _row_text(row, max_cols=35):
    parts = []
    for c in (row or [])[:max_cols]:
        if c is not None and norm(c):
            parts.append(norm(c))
    return " | ".join(parts)


def _interlock_files(path):
    """파일 또는 폴더 → 읽을 xlsx 경로 목록."""
    if not path:
        return []
    if os.path.isdir(path):
        return [
            os.path.join(path, f)
            for f in sorted(os.listdir(path))
            if f.lower().endswith((".xlsx", ".xlsm")) and not f.startswith("~$")
        ]
    if os.path.isfile(path) and path.lower().endswith((".xlsx", ".xlsm")):
        return [path]
    return []


def detect_real_format(path):
    """실물 양식이면 True. 폴더면 안 파일을, 파일은 여러 시트를 본다."""
    files = _interlock_files(path)
    if not files:
        return False
    path = files[0]
    wb = load_workbook(path, read_only=True, data_only=True)
    found_il_no = False
    found_action_by = False
    found_equip = False
    for sn in wb.sheetnames:
        ws = wb[sn]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > 100:
                break
            texts = [norm(c).upper() for c in (row or []) if c is not None]
            joined = " ".join(texts)
            if "IL NO" in joined or "OUTPUT TAG" in joined:
                found_il_no = True
            if ACTION_RE.search(joined):
                found_action_by = True
            if any(EQUIP_RE.match(norm(c) or "") for c in (row or []) if c):
                found_equip = True
        if found_action_by or found_equip:
            break
    if found_il_no and not found_action_by:
        return False
    return found_action_by or found_equip


def load_interlocks_real(path=None):
    path = path or config.INTERLOCK_XLSX
    if os.path.isdir(path):
        # 폴더면 안 xlsx 전부 합산
        items = []
        for f in sorted(os.listdir(path)):
            if f.lower().endswith((".xlsx", ".xlsm")) and not f.startswith("~$"):
                items.extend(load_interlocks_real(os.path.join(path, f)))
        return items
    wb = load_workbook(path, data_only=True)
    items = []
    for sheet_name in wb.sheetnames:
        rows = list(wb[sheet_name].iter_rows(values_only=True))
        items.extend(_parse_interlock_sheet(rows, sheet_name=sheet_name, path=path))
    return items


def _parse_interlock_sheet(rows, sheet_name="", path=""):
    """한 시트의 실물 인터락 행을 items 로 변환."""
    items = []
    current_tag = None
    current_service = ""
    current_action = None
    remark_lines = []
    in_remark = False
    fail_pos = ""
    dwg = ""
    sheet = ""
    used_il_no = set()

    def flush_remark_to_last():
        nonlocal remark_lines, fail_pos, dwg, sheet
        if not items or not remark_lines:
            remark_lines = []
            return
        text = " | ".join(remark_lines)
        it = items[-1]
        # attach to all items of same equipment block — simpler: last item's remark
        # better: all items with same output_tag that are still "open"
        for it2 in reversed(items):
            if it2["output_tag"] != current_tag:
                break
            if not it2.get("remark"):
                it2["remark"] = text
            if dwg and not it2.get("dwg_no"):
                it2["dwg_no"] = dwg
            if sheet and not it2.get("sheet"):
                it2["sheet"] = sheet
            if fail_pos and not it2.get("fail"):
                it2["fail"] = fail_pos
        remark_lines = []

    i = 0
    while i < len(rows):
        row = rows[i]
        if _is_empty_row(row):
            i += 1
            continue

        c0 = _cell(row, 0)
        c1 = _cell(row, 1)
        joined = _row_text(row)

        # ── 장비 헤더 ──────────────────────────────────────
        m = None
        for c in (row or [])[:5]:
            if c and EQUIP_RE.match(norm(c)):
                m = EQUIP_RE.match(norm(c))
                break
        if m:
            flush_remark_to_last()
            in_remark = False
            current_tag = m.group(1).upper()
            current_service = m.group(2).strip()
            current_action = None
            fail_pos, dwg, sheet = "", "", ""
            i += 1
            continue

        # ── 구역 제목(태그 아닌 블록 머리) ──────────────────
        # 여기서 끊지 않으면 아래 조건이 앞 장비에 그대로 붙는다.
        sec = _looks_like_section_head(rows, i)
        if sec:
            flush_remark_to_last()
            in_remark = False
            current_tag = sec.upper()
            current_service = sec
            current_action = None
            fail_pos, dwg, sheet = "", "", ""
            i += 1
            continue

        # ── REMARK 블록 ────────────────────────────────────
        if c0.upper() == "REMARK" or c1.upper() == "REMARK":
            in_remark = True
            i += 1
            continue

        if in_remark:
            # 다음 장비 헤더면 종료
            if any(EQUIP_RE.match(norm(c) or "") for c in (row or [])[:5] if c):
                in_remark = False
                flush_remark_to_last()
                continue  # reprocess same row
            if c0.upper().startswith("EQUIP"):
                in_remark = False
                flush_remark_to_last()
                i += 1
                continue
            line = c1 or c0
            if line:
                remark_lines.append(line)
                if "fail open" in line.lower():
                    fail_pos = "FAIL OPEN"
                elif "fail close" in line.lower():
                    fail_pos = "FAIL CLOSE"
                pm = PID_RE.search(line)
                if pm:
                    sheet = pm.group(1) or pm.group(2)
                    dwg = "P&ID"
            i += 1
            continue

        # ── 동작 헤더 (CLOSE BY / STOP BY) ──────────────────
        am = ACTION_RE.search(joined)
        if am and ("INTERLOCK" in joined.upper() or "SET CONDITION" in joined.upper()
                   or c1.upper().endswith("BY") or " BY" in c1.upper()):
            current_action = am.group(1).upper()
            in_remark = False
            i += 1
            # skip column header row if next is INTERLOCK SET CONDITION
            if i < len(rows):
                nxt = _row_text(rows[i]).upper()
                if "SET CONDITION" in nxt or "RESET CONDITION" in nxt:
                    i += 1
            continue

        # 단독 CLOSE BY 행
        if am and current_tag:
            current_action = am.group(1).upper()
            i += 1
            continue

        # ── 조건 행: 숫자 번호 + SET 조건 ───────────────────
        if current_tag and current_action and c0.isdigit():
            n = int(c0)
            if 1 <= n <= 20:
                # SET condition: col B (index 1) + often status in later cols
                set_parts = []
                if c1 and c1.upper() not in ("INTERLOCK SET CONDITION",):
                    set_parts.append(c1)
                # scan middle columns for status / secondary text (avoid delays/select)
                for ci in range(2, min(len(row), 16)):
                    v = _cell(row, ci)
                    if not v or v == "-":
                        continue
                    up = v.upper()
                    if up in ("RESET CONDITION", "SET DELAY", "RESET DELAY",
                              "ALARM DELAY", "SELECT", "INTERLOCK SET CONDITION"):
                        continue
                    if up.startswith("MODE="):
                        continue
                    if re.match(r"^t\(\d", v, re.I):
                        continue
                    # likely status phrase near SET
                    if ci <= 10:
                        set_parts.append(v)

                reset = ""
                for ci in range(10, min(len(row), 20)):
                    v = _cell(row, ci)
                    if v and ("reset" in v.lower() or "clear" in v.lower()
                              or "run" in v.lower() or "service" in v.lower()):
                        reset = v
                        break
                # typical layout: RESET at col 13 (1-based) → index 12
                if not reset:
                    reset = _cell(row, 12) or ""

                delay_raw = _cell(row, 17) or _cell(row, 18) or ""
                source = _cell(row, 25) or _cell(row, 26) or ""
                note = _cell(row, 28) or _cell(row, 29) or ""

                cond_raw = " ".join(set_parts).strip()
                if not cond_raw or cond_raw == "-":
                    i += 1
                    continue

                # delay from t(7s)
                delay_sec = None
                dm = DELAY_RE.search(delay_raw) or DELAY_RE.search(cond_raw)
                if dm:
                    delay_sec = float(dm.group(1) or dm.group(2))

                c = parse_condition(cond_raw)
                # 실물 태그 패턴 보강
                extra_tags = [t.upper() for t in TAG_RE.findall(cond_raw)]
                for t in extra_tags:
                    if t not in c["tags"]:
                        c["tags"].append(t)
                if delay_sec is not None and c.get("delay_sec") is None:
                    c["delay_sec"] = delay_sec
                    # re-evaluate parsed if we only added delay
                if note:
                    c["raw"] = c["raw"] + (" · " + note if note not in c["raw"] else "")

                # 상태 문구가 있고 state 미검출이면 raw 유지
                if not c["parsed"] and re.search(
                        r"trip|error|stop|select|block|fault|all stop",
                        cond_raw, re.I):
                    # discrete-ish: mark as state free-text but keep parsed=False
                    # unless clear keywords
                    for pat, st, label in [
                        (r"Loop\s*Error", "FAULT", "Loop Error"),
                        (r"EOCR\s*Trip", "FAULT", "EOCR Trip"),
                        (r"All\s*Stop", "STOP", "All Stop"),
                        (r"Block\s*Stop", "STOP", "Block Stop"),
                        (r"Stand-?by\s*Select", "MANUAL_MODE", "Stand-by Select"),
                        (r"\bTrip\b", "FAULT", "Trip"),
                    ]:
                        if re.search(pat, cond_raw, re.I):
                            c["state"] = st
                            c["state_label"] = label
                            c["parsed"] = True
                            c["kind"] = "DISCRETE"
                            break

                # il_no 는 조회 키다. 같은 장비가 시트 안에서 두 번 이상
                # 블록으로 나오면 번호가 블록마다 1부터 다시 시작해 충돌하고,
                # 키로 조회하는 쪽에서 레코드가 조용히 덮인다.
                il_no = "%s-%s-%02d" % (current_tag, current_action, n)
                if il_no in used_il_no:
                    k = 2
                    while "%s#%d" % (il_no, k) in used_il_no:
                        k += 1
                    il_no = "%s#%d" % (il_no, k)
                used_il_no.add(il_no)
                items.append({
                    "il_no": il_no,
                    "output_tag": current_tag,
                    "action": current_action,
                    "kind": "INTERLOCK",
                    "logic": "OR",  # SET conditions typically OR across rows
                    "reset": reset or "MANUAL",
                    "bypassable": False,
                    "priority": 1,
                    "plc_block": source or "",
                    "dwg_no": dwg or "",
                    "sheet": sheet or "",
                    "remark": note or "",
                    "service": current_service,
                    "fail": fail_pos,
                    "conditions": [c],
                })
            i += 1
            continue

        i += 1

    flush_remark_to_last()

    # 같은 장비+동작의 조건을 하나의 OR 항목으로 묶을 수도 있지만
    # 행 단위 IL 번호가 현장과 맞으므로 행=항목 유지
    return items


def outputs_from_real(items):
    """파싱 결과에서 출력 태그 메타 추출."""
    out = {}
    for it in items:
        tag = it["output_tag"]
        if tag not in out:
            out[tag] = {
                "tag": tag,
                "service": it.get("service") or "",
                "type": "",
                "fail": it.get("fail") or "",
            }
        elif it.get("fail") and not out[tag].get("fail"):
            out[tag]["fail"] = it["fail"]
        if it.get("service") and not out[tag].get("service"):
            out[tag]["service"] = it["service"]
    return out



def extract_source_block(tag: str, path=None):
    """장비 태그 구간을 실물 양식에 가깝게 구조화해 반환.

    path 가 폴더(data/interlock) 여도 안 xlsx 를 모두 찾고,
    시트도 전부 검색한다. F05-...-6019A 질의는 ...6019A~F 헤더와도 매칭.
    """
    path = path or config.INTERLOCK_XLSX
    tag_q = (tag or "").strip().upper()
    if not tag_q:
        return None

    def aliases(t):
        t = (t or "").strip().upper()
        out = {t}
        import re as _re
        base = _re.sub(r"~[A-Z0-9]+$", "", t)
        if base:
            out.add(base)
        if "/" in t:
            for p in t.split("/"):
                p = p.strip()
                if p:
                    out.add(p)
                    out.add(_re.sub(r"~[A-Z0-9]+$", "", p))
        return out

    want = aliases(tag_q)
    files = _interlock_files(path)
    if not files:
        return None

    for fpath in files:
        try:
            wb = load_workbook(fpath, data_only=True)
        except Exception as e:
            print("[interlock_real] open fail", fpath, e)
            continue
        for sheet_name in wb.sheetnames:
            rows = list(wb[sheet_name].iter_rows(values_only=True))
            start = None
            title = ""
            for i, row in enumerate(rows):
                for c in (row or [])[:5]:
                    if not c:
                        continue
                    m = EQUIP_RE.match(norm(c))
                    if not m:
                        continue
                    raw = m.group(1).upper()
                    if raw in want or any(a in raw or raw in a for a in want):
                        # 더 엄격: base 일치
                        import re as _re
                        raw_base = _re.sub(r"~[A-Z0-9]+$", "", raw)
                        if raw_base in want or raw in want or any(
                            _re.sub(r"~[A-Z0-9]+$", "", a) == raw_base for a in want
                        ):
                            start = i
                            title = norm(c)
                            break
                if start is not None:
                    break
            if start is None:
                continue

            end = len(rows)
            for j in range(start + 1, len(rows)):
                row = rows[j]
                hit = False
                for c in (row or [])[:5]:
                    if c and EQUIP_RE.match(norm(c)):
                        end = j
                        hit = True
                        break
                if hit:
                    break
                c0 = norm((row or [None])[0]) if row else ""
                if c0.upper().startswith("EQUIP"):
                    end = j
                    break

            def cell(row, idx):
                if not row or idx >= len(row) or row[idx] is None:
                    return ""
                return str(row[idx]).replace("\n", " ").strip()

            action = None
            conditions = []
            remarks = []
            in_remark = False
            col_headers = [
                "No", "INTERLOCK SET CONDITION", "STATUS / 부가",
                "RESET CONDITION", "SET DELAY", "SELECT", "SIGNAL SOURCE", "NOTE",
            ]

            for ridx in range(start + 1, end):
                row = rows[ridx]
                c0 = cell(row, 0)
                c1 = cell(row, 1)
                joined = " ".join(
                    cell(row, i) for i in range(min(30, len(row) if row else 0))
                    if cell(row, i)
                )
                if c0.upper() == "REMARK" or c1.upper() == "REMARK":
                    in_remark = True
                    continue
                if in_remark:
                    line = c1 or c0
                    if line:
                        remarks.append(line)
                    continue
                am = ACTION_RE.search(joined)
                if am and ("BY" in joined.upper()):
                    action = am.group(1).upper() + " BY"
                    continue
                if "SET CONDITION" in joined.upper() and "RESET" in joined.upper():
                    continue
                if c0.isdigit() and 1 <= int(c0) <= 20:
                    set_cond = c1
                    status = cell(row, 7)
                    if not status:
                        for i in range(2, 12):
                            v = cell(row, i)
                            if v and v not in ("-", set_cond) and "RESET" not in v.upper():
                                status = v
                                break
                    reset = cell(row, 12)
                    if not reset:
                        for i in range(10, 18):
                            v = cell(row, i)
                            if v and any(k in v.lower() for k in
                                         ("reset", "clear", "run", "service")):
                                reset = v
                                break
                    delay = cell(row, 17) or cell(row, 18) or "-"
                    select = cell(row, 23) or cell(row, 24) or "-"
                    source = cell(row, 25) or cell(row, 26) or ""
                    note = cell(row, 28) or cell(row, 29) or ""
                    if not set_cond or set_cond == "-":
                        continue
                    conditions.append({
                        "no": c0,
                        "set": set_cond,
                        "status": status,
                        "reset": reset or "-",
                        "delay": delay if delay else "-",
                        "select": select if select else "-",
                        "source": source,
                        "note": note,
                        "excel_row": ridx + 1,
                    })

            grid_rows = []
            for ridx in range(start, end):
                row = rows[ridx]
                cells = []
                for c in (row or [])[:32]:
                    if c is None or (isinstance(c, str) and not str(c).strip()):
                        cells.append("")
                    else:
                        cells.append(str(c).replace("\n", " ").strip())
                while cells and cells[-1] == "":
                    cells.pop()
                if cells:
                    grid_rows.append({"row": ridx + 1, "cells": cells})

            return {
                "tag": tag_q,
                "file": os.path.basename(fpath),
                "path": fpath,
                "sheet": sheet_name,
                "header": title or tag_q,
                "action": action,
                "columns": col_headers,
                "conditions": conditions,
                "remarks": remarks,
                "row_start": start + 1,
                "row_end": end,
                "rows": grid_rows,
            }
    return None



def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--path", default=None)
    ap.add_argument("--stats", action="store_true")
    ap.add_argument("--dump", default=None, help="output tag")
    args = ap.parse_args()
    path = args.path or os.path.join(config.DATA_DIR, "INTERLOCK_SAMPLE.xlsx")
    items = load_interlocks_real(path)
    conds = [c for it in items for c in it["conditions"]]
    ok = sum(1 for c in conds if c["parsed"])
    print("실물 파서: 항목 %d / 조건 %d / 구조화 %d (%.0f%%)"
          % (len(items), len(conds), ok, 100 * ok / max(1, len(conds))))
    tags = sorted({it["output_tag"] for it in items})
    print("출력 태그:", ", ".join(tags))
    if args.dump:
        for it in items:
            if it["output_tag"] == args.dump.upper():
                print("\n[%s] %s → %s" % (it["il_no"], it["kind"], it["action"]))
                for c in it["conditions"]:
                    print("  ·", c["raw"], "| tags=", c["tags"],
                          "| parsed=", c["parsed"], c.get("state"))
                print("  reset:", it["reset"], "remark:", it["remark"])


if __name__ == "__main__":
    main()

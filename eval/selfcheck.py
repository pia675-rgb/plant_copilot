#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
selfcheck.py — 시연 전 전 경로 점검 (고장 주입 포함)

preflight 는 "환경이 갖춰졌는가"를 본다. 이 스크립트는 한 걸음 더
들어가 **각 경로가 실제로 동작하는가, 그리고 실패했을 때 그 사실이
드러나는가**를 본다.

    python -m eval.selfcheck
    python -m eval.selfcheck --skip-llm     # 모델 없이 구조만

## 왜 만들었나

지금까지 세 번 같은 유형으로 깨졌다.

1. 임베딩 캐시가 다른 모델 것이어도 그대로 재사용됨 → 검색 품질만 조용히 저하
2. 조치 생성 LLM 이 없어도 화면은 그럴듯하게 나옴 → 근거 나열인 줄 모름
3. COPILOT_PROVIDER=ollama 가 챗봇에서는 인정되지 않음 → 조용히 규칙 엔진

공통점은 **실패해야 알 수 있는 가정**이었고, 실패해도 화면이 그럴듯해서
알기 어려웠다는 것이다. 그래서 여기서는 정상 동작만 보지 않고, 고장을
주입해서 **검출 장치가 실제로 작동하는지**까지 확인한다.
"""

import argparse
import os
import re
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import config  # noqa: E402

_rows = []


def run(name, fn, critical=True):
    try:
        ok, detail = fn()
    except Exception as e:                                  # noqa: BLE001
        ok, detail = False, "%s: %s" % (type(e).__name__, str(e)[:130])
    _rows.append(("OK  " if ok else ("실패" if critical else "주의"),
                  name, detail))
    return ok


# ── 1. 검색 경로 ────────────────────────────────────────────
def c_search_modes():
    """세 모드가 모두 구성되고 강등 없이 도는가."""
    from retrieval.pipeline import Retriever
    out = []
    for m in ("lexical", "hybrid"):
        r = Retriever(mode=m)
        if r.degraded:
            return False, "%s 강등: %s" % (m, ",".join(r.degraded))
        n = len(r.retrieve("acid residual low", tag="AIT-4002"))
        out.append("%s %d건" % (m, n))
    return True, " / ".join(out)


def c_korean_query():
    """한글 질의가 실제로 영문 매뉴얼에 닿는가 — v2 논지의 핵심."""
    from graph.app_graph import Copilot2
    o = Copilot2(mode="hybrid").answer(
        tag="AIT-4002", alarm="산 잔량이 부족하다고 뜹니다")
    if o["decision"] != "advise":
        return False, "판정 %s (충분성 %.2f) — 한글 질의가 거절됨" % (
            o["decision"], o.get("grade", 0))
    return True, "advise %.2f, 근거 %d건" % (o.get("grade", 0),
                                           len(o.get("evidence", [])))


def c_abstain_works():
    """벤더 문서 없는 태그를 실제로 거절하는가."""
    from graph.app_graph import Copilot2
    o = Copilot2(mode="hybrid").answer(
        tag="PIT-2003", alarm="압력 지시가 이상합니다")
    if o["decision"] != "abstain":
        return False, "판정 %s — 근거 없는 태그에 답하고 있음" % o["decision"]
    # 거절했다고 다 통과시키면 안 된다. 검색이 죽어서 아무것도 못 찾은
    # 경우도 abstain 이 나오기 때문이다. 벤더 문서 게이트가 이유여야 한다.
    parts = o.get("grade_parts") or {}
    if "no_vendor_doc" not in parts:
        return False, ("거절은 했으나 사유가 다릅니다 (%s) — 검색이 죽어서 "
                       "생긴 거절일 수 있습니다" % (o.get("grade_reason") or "")[:50])
    return True, "벤더 문서 게이트로 거절 — %s" % (o.get("grade_reason") or "")[:50]


def c_degrade_visible():
    """
    [주입] 강등이 화면에서 보이는가.

    벡터 검색이 꺼지면 대부분의 질의가 거절로 끝난다. 그런데 거절은
    정상 동작처럼 보여서 "왜 다 거절하지" 만 반복하게 된다. 실제로
    모든 태그가 abstain 인 상태를 며칠 만에 발견했다.

    /api/health 가 강등 사유와 그 결과를 함께 싣는지 확인한다. 강등이
    없는 정상 상태에서는 통과다 — 이 항목이 보는 것은 '강등되었을 때
    드러나는가' 이지 '강등되었는가' 가 아니다.
    """
    from api.server import health
    h = health()
    if not h.get("degraded"):
        return True, "강등 없음 (%s)" % h.get("effective_mode")
    detail = h.get("degrade_detail")
    if not detail:
        return False, ("강등 %s 인데 사유가 실리지 않습니다 — 화면에서 "
                       "원인을 짚을 수 없습니다" % h["degraded"])
    for d in detail:
        if not d.get("reason") or d["reason"] == "사유 미기록":
            return False, "%s 강등 사유가 비어 있습니다" % d.get("component")
    return True, "강등 %s / 사유·영향 노출됨" % ", ".join(h["degraded"])


def c_diversify_off():
    """측정에서 기각된 설정이 켜져 있지 않은가."""
    if config.DIVERSIFY != "off":
        return False, ("COPILOT_DIVERSIFY=%s — 측정 결과 off(33/45)보다 "
                       "낮습니다(kind 29, all 27)" % config.DIVERSIFY)
    return True, "off"


# ── 2. 임베딩 캐시 신원 ─────────────────────────────────────
def c_cache_identity():
    """다른 모델로 구운 캐시를 재사용하지 않는가 (고장 주입)."""
    import json
    if not os.path.isfile(config.EMBED_META):
        return False, "캐시 없음"
    meta = json.load(open(config.EMBED_META, encoding="utf-8"))
    if meta.get("signature") != config.embed_signature():
        return False, "서명 불일치: %s ≠ %s" % (meta.get("signature"),
                                            config.embed_signature())

    # 주입: 서명을 바꾼 척하고 load() 가 거부하는지 본다
    from ingest.build_index import load as load_index
    from retrieval.dense import DenseIndex
    real = config.EMBED_MODEL
    try:
        config.EMBED_MODEL = real + "-FAKE"
        di = DenseIndex(load_index())
        if di.load(verbose=False) is not None:
            return False, "다른 모델 서명인데 캐시를 재사용함 — 검출 실패"
    finally:
        config.EMBED_MODEL = real
    return True, "서명 대조 정상 (%s)" % meta.get("signature")


# ── 3. 조치 생성 ────────────────────────────────────────────
def c_advisor_reachable():
    from graph.advisor import available
    return available()


def c_advisor_rejects_fake():
    """없는 근거를 인용하면 버리는가 (고장 주입)."""
    from graph import advisor
    ev = [{"id": "M9E-401", "title": "Acid container",
           "text": "volume of acid is less than 10%", "cite": "a.pdf p.400"}]
    orig = config.LLM_PROVIDER
    try:
        config.LLM_PROVIDER = "_selfcheck"
        advisor._CHAT["_selfcheck"] = lambda m, t: (
            '{"summary":"x","steps":[{"title":"A","detail":"B",'
            '"evidence":["M9E-9999"]},{"title":"C","detail":"D",'
            '"evidence":[99]}]}')
        try:
            advisor.generate("AIT-4002", "산 잔량 부족", ev)
            return False, "없는 근거를 인용했는데 통과함 — 환각 차단 실패"
        except advisor.AdvisorError:
            pass
        # 정상 인용은 통과해야 한다
        advisor._CHAT["_selfcheck"] = lambda m, t: (
            '{"summary":"x","steps":[{"title":"A","detail":"B",'
            '"evidence":[1]}]}')
        r = advisor.generate("AIT-4002", "산 잔량 부족", ev)
        if r["steps"][0]["evidence_ids"] != ["M9E-401"]:
            return False, "정상 인용이 잘못 되돌려짐"
    finally:
        config.LLM_PROVIDER = orig
        advisor._CHAT.pop("_selfcheck", None)
    return True, "가짜 근거 차단 / 번호 인용 복원 정상"


def c_advisor_real():
    """실제 모델로 조치가 생성되는가."""
    from graph.advisor import generate
    from graph.app_graph import Copilot2
    o = Copilot2(mode="hybrid").answer(
        tag="AIT-4002", alarm="산 잔량이 부족하다고 뜹니다")
    if o["decision"] != "advise":
        return False, "판정 %s — 조치 생성 단계에 도달하지 못함" % o["decision"]
    r = generate("AIT-4002", "산 잔량이 부족하다고 뜹니다", o["evidence"])
    if r["dropped"]:
        return False, "%d개 단계가 근거 검증에서 버려짐" % r["dropped"]
    return True, "%d단계 — %s" % (len(r["steps"]), r["steps"][0]["title"])


# ── 4. 챗봇 ─────────────────────────────────────────────────
def c_chat_gateway():
    """챗봇과 조치 생성이 같은 제공자 설정을 쓰는가."""
    from graph.advisor import _CHAT
    if config.LLM_PROVIDER == "off":
        return True, "off — 규칙 엔진만 사용 (의도된 구성)"
    if config.LLM_PROVIDER not in _CHAT:
        return False, ("COPILOT_PROVIDER=%s 를 대화 게이트웨이가 모릅니다"
                       % config.LLM_PROVIDER)
    return True, "%s — 조치 생성과 동일 경로" % config.LLM_PROVIDER


def c_chat_rule():
    """모델이 없어도 규칙 엔진이 명령을 알아듣는가."""
    from api.server import rule_intent
    cases = [("AIT-4002 산 잔량 알람 조회해줘", "diagnose", "AIT-4002"),
             ("XV-4101 인터락 조회", "interlock", "XV-4101")]
    bad = []
    for t, want, tag in cases:
        r = rule_intent(t) or {}
        if r.get("type") != want or r.get("tag") != tag:
            bad.append(t)
    if bad:
        return False, "인식 실패: %s" % " / ".join(bad)
    return True, "알람·인터락 명령 인식 정상"


def c_tag_notation():
    """현장 표기(공백·하이픈 없음)를 실제 태그로 되돌리는가."""
    from api.server import normalize_tag, rule_intent
    for raw, want in (("LCV 01", "LCV-01"), ("lcv01", "LCV-01"),
                      ("XV 4101", "XV-4101"), ("AIT 4002", "AIT-4002")):
        if normalize_tag(raw) != want:
            return False, "%s -> %s (기대 %s)" % (raw, normalize_tag(raw), want)
    # 없는 태그를 지어내지 않는가
    if normalize_tag("ZZZ-99") is not None:
        return False, "존재하지 않는 태그를 통과시킴"
    r = rule_intent("LCV 01 인터락 보여줘")
    if r.get("type") != "interlock" or r.get("tag") != "LCV-01":
        return False, "'LCV 01 인터락 보여줘' -> %s / %s" % (r.get("type"),
                                                          r.get("tag"))
    if (rule_intent("ZZZ-99 인터락 조회") or {}).get("type") != "chat":
        return False, "없는 태그로 조회를 시도함"
    return True, "공백·무하이픈 표기 복원 / 없는 태그 거절"


# ── 5. 인터락 ───────────────────────────────────────────────
def c_chat_dispatch():
    """
    LLM 이 엉뚱한 응답을 해도 화면이 실행되는가 (고장 주입).

    7B 급 모델은 이 작업에서 규칙보다 못하다. "LCV-01 인터락 조회해줘"
    에 type=chat 과 친절한 문구만 돌려주면 화면은 아무것도 하지 않고,
    사용자는 왜 안 되는지 알 수 없다. 규칙이 우선인지 확인한다.
    """
    from graph import advisor
    from api.server import chat_help, ChatRequest
    orig = config.LLM_PROVIDER
    try:
        config.LLM_PROVIDER = "_selfcheck"
        advisor._CHAT["_selfcheck"] = lambda m, t: (
            '{"type":"chat","reply":"확인하겠습니다."}')
        cases = [("LCV-01 인터락 조회해줘", "interlock", "LCV-01"),
                 ("LCV 01 인터락 보여줘", "interlock", "LCV-01"),
                 ("AIT-4002 산 잔량 알람 조회해줘", "diagnose", "AIT-4002"),
                 ("사용법", "help", None)]
        for msg, want, tag in cases:
            r = chat_help(ChatRequest(message=msg, tab="alarm", use_llm=True))
            if r.get("type") != want or (tag and r.get("tag") != tag):
                return False, "'%s' -> %s / %s (기대 %s / %s)" % (
                    msg, r.get("type"), r.get("tag"), want, tag)
        # 없는 태그는 여전히 실행하지 않아야 한다
        r = chat_help(ChatRequest(message="ZZZ-99 인터락 조회", tab="alarm",
                                  use_llm=True))
        if r.get("type") != "chat":
            return False, "없는 태그로 조회를 시도함"

        # 반대 방향도 본다. 규칙이 못 알아들은 질문은 LLM 답이
        # 살아 있어야 한다 — 규칙의 예시 문구로 덮어쓰면 "어떤 기능이
        # 있니" 에 예시만 반복하게 된다.
        advisor._CHAT["_selfcheck"] = lambda m, t: (
            '{"type":"chat","reply":"__LLM_ANSWER__"}')
        # QA 경로를 끄고 일반 대화만 본다 — 여기서 보려는 것은
        # "규칙의 예시 문구가 LLM 답을 덮지 않는가" 이지 QA 품질이 아니다.
        #
        # 탐침을 "안녕하세요" 에서 바꿨다. 인사는 이제 규칙이 직접
        # 응대하므로(c_smalltalk 참조) 포괄 응답 경로를 지나지 않는다.
        # 규칙에 걸리지 않는 잡담이어야 이 항목이 원래 보려던 것을 본다.
        qa_on = config.CHAT_QA
        try:
            config.CHAT_QA = False
            r = chat_help(ChatRequest(message="그냥 해본 말이야",
                                      tab="alarm", use_llm=True))
            if r.get("reply") != "__LLM_ANSWER__":
                return False, "규칙 예시 문구가 LLM 답변을 덮어씀"
        finally:
            config.CHAT_QA = qa_on
    finally:
        config.LLM_PROVIDER = orig
        advisor._CHAT.pop("_selfcheck", None)
    return True, "규칙 우선 실행 / 대화 질문은 LLM 답 유지"


def c_chat_qa():
    """챗봇 질의응답이 근거 없는 답을 만들지 않는가 (고장 주입)."""
    from graph import advisor, qa
    orig = config.LLM_PROVIDER
    ev = [{"id": "M9E-401", "title": "Acid container",
           "text": "volume of acid is less than 10%", "cite": "m9.pdf p.400"}]

    class Advise:
        def answer(self, **k):
            return {"decision": "advise", "grade": 0.7, "evidence": ev}

    class Abstain:
        def answer(self, **k):
            return {"decision": "abstain", "grade": 0.2, "evidence": [],
                    "grade_reason": "근거 없음"}

    try:
        config.LLM_PROVIDER = "_selfcheck"
        # 없는 근거를 인용한 문장이 섞인 응답
        advisor._CHAT["_selfcheck"] = lambda m, t: (
            '{"answer":"잔량이 10% 미만이면 경보가 뜹니다 [1]. '
            '펌프를 분해하십시오 [9].","used":[1]}')
        r = qa.answer("산 잔량 경보는 언제 뜨나요?", copilot=Advise())
        if not r["ok"] or r.get("dropped", 0) < 1:
            return False, "없는 근거 인용을 버리지 못함"
        if "분해" in r["reply"]:
            return False, "지어낸 문장이 답변에 남음"
        # 근거가 없으면 답을 만들지 않아야 한다
        r2 = qa.answer("커피 냄새가 나요", copilot=Abstain())
        if r2["ok"]:
            return False, "근거가 없는데 답변을 생성함"
    finally:
        config.LLM_PROVIDER = orig
        advisor._CHAT.pop("_selfcheck", None)
    return True, "인용 검증 / 근거 없을 때 거절"


def c_chat_followup():
    """화면 결과를 근거로 후속 질문에 답하는가."""
    from graph import advisor
    from api.server import chat_help, ChatRequest, ChatContext, rule_intent
    # 후속 질문이 명령으로 오인되지 않아야 한다
    for msg in ("조회된 내용을 보고 조치방법을 알려줘", "조치방법 알려줘",
                "방금 결과 설명해줘"):
        if (rule_intent(msg, "AIT-1001") or {}).get("type") != "followup":
            return False, "'%s' 가 명령으로 오인됨" % msg
    # 결과가 없으면 먼저 조회하라고 안내해야 한다
    r = chat_help(ChatRequest(message="조치방법 알려줘", tag="AIT-1001",
                              use_llm=True))
    if "조회" not in (r.get("reply") or ""):
        return False, "조회 결과가 없는데 안내하지 않음"

    orig = config.LLM_PROVIDER
    try:
        config.LLM_PROVIDER = "_selfcheck"
        advisor._CHAT["_selfcheck"] = lambda m, t: (
            '{"summary":"확인 순서","steps":[{"title":"산 용기 잔량 확인",'
            '"detail":"10% 미만이면 주문하십시오.","evidence":[1]}]}')
        ev = [{"id": "M9E-401", "title": "Acid container",
               "text": "volume of acid is less than 10%",
               "cite": "m9.pdf p.400"}]
        ctx = ChatContext(tag="AIT-1001", alarm="acid low",
                          decision="advise", grade=0.71, evidence=ev)
        r = chat_help(ChatRequest(message="조회된 내용을 보고 조치방법을 알려줘",
                                  tag="AIT-1001", context=ctx, use_llm=True))
        if r.get("engine") != "advice" or "산 용기" not in (r.get("reply") or ""):
            return False, "화면 결과로 조치를 만들지 못함: %s" % r.get("engine")
    finally:
        config.LLM_PROVIDER = orig
        advisor._CHAT.pop("_selfcheck", None)
    return True, "후속 질문을 화면 근거로 처리"


def c_interlock_eval():
    """인터락 72문항이 통과하는가."""
    import json
    p = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                     "eval_set_interlock.json")
    if not os.path.isfile(p):
        return False, "평가셋 없음 — python -m eval.make_eval_interlock"
    from eval.run_eval_interlock import check
    from retrieval.interlock_index import InterlockIndex
    ix = InterlockIndex()
    qs = json.load(open(p, encoding="utf-8"))["questions"]
    bad = [q["id"] for q in qs if not check(ix, q)[0]]
    if bad:
        return False, "%d문항 실패: %s" % (len(bad), ",".join(bad[:6]))
    return True, "%d/%d" % (len(qs), len(qs))


# ── 5-2. 판넬 ───────────────────────────────────────────────
def c_panel_eval():
    """판넬 평가셋이 통과하는가."""
    import json
    p = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                     "eval_set_panel.json")
    if not os.path.isfile(p):
        return False, "평가셋 없음 — python -m eval.make_eval_panel"
    from eval.run_eval_panel import check
    from retrieval.panel_index import PanelIndex
    ix = PanelIndex()
    qs = json.load(open(p, encoding="utf-8"))["questions"]
    bad = [q["id"] for q in qs if not check(ix, q)[0]]
    if bad:
        return False, "%d문항 실패: %s" % (len(bad), ",".join(bad[:6]))
    return True, "%d/%d" % (len(qs), len(qs))


def c_card_scope():
    """
    [주입] 카드 단위 상실 계산이 실제로 카드를 구분하고 있는가.

    이중화 구성에서 단일 고장 지점은 카드다. 슬롯 구분이 사라지면
    카드 조회는 '판넬 전체'와 같아지는데, 화면은 여전히 카드라고
    표시한다. 그 상태를 검출한다.
    """
    from retrieval.panel_index import PanelIndex
    ix = PanelIndex()
    cards = ix.cards()
    if len(cards) < 2:
        return False, "카드가 %d장뿐 — 슬롯 정보를 읽지 못했을 수 있습니다" % len(cards)
    cid = cards[0]["card"]
    d = ix.impact(cid)
    panel_points = ix.by_panel(d["panel"])["points"]
    if d["points"] >= panel_points:
        return False, ("카드 상실(%d점)이 판넬 전체(%d점)와 같습니다 — "
                       "슬롯 구분이 동작하지 않습니다"
                       % (d["points"], panel_points))
    return True, "카드 %d장 / %s 는 %d점 (판넬 %d점 중)" % (
        len(cards), cid, d["points"], panel_points)


def c_card_channels_visible():
    """
    [주입] 카드 상실 화면이 잃는 점을 빠짐없이 보여주는가.

    화면에서 판넬 계기 목록을 없앴는데, 인터락 표에는 **조건에 걸린
    태그만** 나온다. 그래서 인터락에 안 걸린 계기가 어디에도 보이지
    않았다. 실제로 CUB-B/R1/S8 은 3점인데 인터락 조건에 걸린 것은
    1점뿐이라, 나머지 2점이 화면에서 사라졌다.

    impact 가 채널 전체를 싣는지, 그리고 인터락 의존 여부가 채널마다
    붙는지 확인한다.
    """
    from retrieval.panel_index import PanelIndex
    ix = PanelIndex()
    partial = None          # 일부 채널만 인터락에 걸린 카드
    for c in ix.cards():
        d = ix.impact(c["card"])
        if not d or not d.get("interlock_loaded"):
            continue
        chs = d.get("channels") or []
        if len(chs) != d["points"]:
            return False, ("%s: 채널 %d개 ≠ 잃는 점 %d — 화면에서 일부 계기가 "
                           "보이지 않습니다" % (c["card"], len(chs), d["points"]))
        hit = [x for x in chs if x.get("interlocks")]
        if hit and len(hit) < len(chs):
            partial = (c["card"], len(chs), len(hit))
    if partial is None:
        return True, "채널 수 일치 (일부만 인터락에 걸린 카드는 없음)"
    card, n, k = partial
    return True, "%s 등: 채널 %d개 전부 노출 (인터락 관련 %d개)" % (card, n, k)


def c_common_cause():
    """
    [주입] 공통원인 점검이 실제로 무언가를 붙잡는가.

    지적 0건은 '설계가 안전하다'가 아니라 '이 리스트에서 걸린 것이
    없다'는 뜻이다. 조건 태그 두 개를 같은 카드로 옮겨 잡히는지 본다.
    """
    from retrieval.panel_index import PanelIndex
    ix = PanelIndex()
    base = ix.common_cause()
    if not base["loaded"]:
        return False, "인터락 리스트 미적재"
    # 조건 태그가 2개 이상인 인터락 하나를 골라 같은 카드로 몰아 본다
    target = None
    for it in ix._interlock().items:
        cond = set()
        for c in it["conditions"]:
            cond.update(c["tags"])
        known = [t for t in cond if t in ix._by_tag]
        if len(known) >= 2:
            target = (it["il_no"], known[:2])
            break
    if not target:
        return True, ("조건 태그가 2개 이상인 인터락이 없어 시험 건너뜀 — "
                      "실물 인터락 리스트는 행마다 조건이 하나(OR)입니다")
    il_no, (a, b) = target
    # 카드 식별에 PN(DP)(스테이션)까지 들어가므로 그 값도 함께 옮겨야
    # 실제로 같은 카드가 된다. 배선 열을 하나라도 빼면 주입이 성립하지
    # 않는데, 검사는 통과한 것처럼 보인다.
    wiring = ("PANEL", "PN(DP)", "RACK", "SLOT")
    src, dst = ix._by_tag[a], ix._by_tag[b]
    saved = {k: dst.get(k) for k in wiring}
    for k in wiring:
        dst[k] = src.get(k)
    got = [f["il_no"] for f in ix.common_cause()["findings"]]
    dst.update(saved)
    if il_no not in got:
        return False, "%s 의 조건을 같은 카드로 몰았는데 잡지 못함" % il_no
    return True, "기준 지적 %d건 / 주입 시 %s 검출" % (
        len(base["findings"]), il_no)


def c_no_panel_impact():
    """
    판넬 단위 '상실' 계산이 되살아나지 않았는가.

    이중화 구성에서는 랙 증설도 스위칭으로 대응하므로 판넬 전체가 죽는
    상황이 성립하지 않는다. 성립하지 않는 시나리오에 숫자를 붙여 보여주면
    현장 판단을 왜곡한다. 편의로 다시 넣기 쉬운 기능이라 잠가 둔다.
    """
    import inspect
    from retrieval.panel_index import PanelIndex
    from api.server import panel_detail
    sig = inspect.signature(PanelIndex.impact)
    if "scope" in sig.parameters:
        return False, "PanelIndex.impact 에 scope 인자가 되살아남"
    if "impact" in inspect.signature(panel_detail).parameters:
        return False, "/api/panel 에 impact 인자가 되살아남"
    ix = PanelIndex()
    if ix.impact(sorted(ix._by_panel)[0]) is not None:
        return False, "판넬명으로 상실 계산이 반환됨"
    return True, "카드 단위만 계산 / 판넬은 구성·위치 조회 전용"


def c_data_dir_clean():
    """
    [주입] data/ 에 업로드 자료 말고 다른 것이 섞이지 않았는가.

    data/ 는 **사용자가 넣는 아홉 가지만** 두는 자리다. 생성물이 같이
    있으면 "무엇을 넣어야 하나" 에 한 줄로 답할 수 없고, 실제로 생성물을
    원본으로 착각해 엑셀에서 손으로 고쳤다가 재생성에서 잃을 뻔했다.
    임시 파일(엑셀을 열어 둔 채 생성기를 돌리면 생긴다)도 여기서 걸린다.
    """
    allowed_files = {"INSTRUMENT_LIST.xlsx", "IO_LIST.xlsx", "TB_LIST.xlsx"}
    # 계기 리스트는 계기 종류별로 여러 통으로 들어오는 것이 실물이다
    # (Flow Transmitter · Level Switch · Pressure Gauge …).
    allowed_re = re.compile(r"instrument\s*list.*\.xlsx?$", re.I)
    allowed_dirs = {"manuals", "drawings", "interlock"}
    generated = {"PANEL_LOCATIONS.csv", "TAG_ATTRIBUTES.xlsx  (폐지)",
                 "drawings_index.csv", "error_codes.json",
                 "maintenance_history.json", "maintenance_history.xlsx"}
    if not os.path.isdir(config.DATA_DIR):
        return False, "data/ 가 없습니다: %s" % config.DATA_DIR
    stray = []
    for name in sorted(os.listdir(config.DATA_DIR)):
        full = os.path.join(config.DATA_DIR, name)
        if os.path.isdir(full):
            if name not in allowed_dirs and name != "__pycache__":
                stray.append(name + "/")
            continue
        if name in allowed_files or allowed_re.search(name):
            continue
        if name in generated:
            stray.append("%s (생성물 — derived/ 로)" % name)
        elif name.startswith("~$") or "." not in name:
            stray.append("%s (임시 파일 — 엑셀을 닫고 생성기를 다시 실행)"
                         % name)
        else:
            stray.append(name)
    if stray:
        return False, "data/ 에 업로드 자료가 아닌 것: " + ", ".join(stray)
    n_man = len([f for f in os.listdir(config.MANUAL_DIR)
                 if f.lower().endswith(".pdf")]) \
        if os.path.isdir(config.MANUAL_DIR) else 0
    n_dwg = len([f for f in os.listdir(config.DRAWING_DIR)
                 if f.lower().endswith(".pdf")]) \
        if os.path.isdir(config.DRAWING_DIR) else 0
    return True, "업로드 자료만 (매뉴얼 %d · 도면 %d)" % (n_man, n_dwg)


def c_io_list_standard():
    """
    [주입] 표준 IO List 가 원천과 어긋나지 않는가.

    IO_LIST.xlsx 는 생성물이다. 원천(v1 계기 리스트·출력 리스트)이 바뀌면
    다시 만들어야 하는데, 안 만들면 조회는 옛 IO List 를, 사람은 새 원천을
    보게 된다. 둘 다 그럴듯해서 어긋났다는 사실이 드러나지 않는다.
    PROVENANCE 시트에 찍힌 원천 행수와 실제 원천을 대조한다.
    """
    if not os.path.isfile(config.IO_LIST):
        return True, "표준 IO List 미사용 (v1 원본으로 동작 중)"
    try:
        import openpyxl
    except ImportError:
        return False, "openpyxl 없음"
    wb = openpyxl.load_workbook(config.IO_LIST, read_only=True, data_only=True)
    if "PROVENANCE" not in wb.sheetnames:
        return True, ("PROVENANCE 시트가 없습니다 — 생성기가 아니라 사용자가 "
                      "직접 넣은 IO List 로 보고 원천 대조는 건너뜁니다")
    stamp = {}
    for r in wb["PROVENANCE"].iter_rows(values_only=True):
        if r and r[0]:
            stamp[str(r[0]).strip()] = r[1]

    def count(path):
        ws = openpyxl.load_workbook(path, read_only=True,
                                    data_only=True).active
        rows = list(ws.iter_rows(values_only=True))
        hi = next(i for i, x in enumerate(rows)
                  if x and "TAG" in [str(c).strip().upper() if c else ""
                                     for c in x])
        ti = [str(c).strip().upper() if c else ""
              for c in rows[hi]].index("TAG")
        return sum(1 for x in rows[hi + 1:] if x and x[ti])

    src_in = os.path.join(config.SOURCE_DIR, "DEMO_INSTRUMENT_LIST.xlsx")
    if not os.path.isfile(src_in):
        return True, "원천 파일이 없어 대조를 건너뜁니다"
    actual = count(src_in)
    src_out = os.path.join(config.SOURCE_DIR, "DEMO_OUTPUT_LIST.xlsx")
    if os.path.isfile(src_out):
        actual += count(src_out)
    try:
        filed = int(stamp.get("SOURCE ROWS",
                              stamp.get("SOURCE INPUT ROWS", -1)))
    except (TypeError, ValueError):
        filed = -1
    if filed != actual:
        return False, ("원천 %d행 ≠ IO List 기록 %s행 — python "
                       "tools/make_io_list.py 를 다시 실행하십시오"
                       % (actual, filed))
    # 계기 리스트도 같은 원천에서 나왔는지
    spec = getattr(config, "INSTRUMENT_SPEC", "")
    if spec and os.path.isfile(spec):
        from ingest.lists import read_instrument_rows
        n = len(read_instrument_rows(spec))
        if n != actual:
            return False, ("계기 리스트 %d행 ≠ IO List %d행 — 두 문서가 "
                           "다른 시점의 원천에서 나왔습니다" % (n, actual))
    else:
        return False, ("INSTRUMENT_LIST.xlsx 가 없습니다 — 제조사·고장모드·"
                       "매뉴얼 연결이 통째로 빕니다")
    return True, "IO List·계기 리스트 %d행 원천 일치" % actual


def c_io_list_header():
    """
    [주입] IO List 헤더가 mastertool/MAXIS 표준 24종과 정확히 같은가.

    표기가 한 글자만 달라도 완전일치 매핑에서 조용히 빗나간다. 실제로
    5번 열은 PN(DP) 인데 구 표기 DP(PN) 이 문서에 남아 있다. MAXIS 는
    별칭으로 받아주지만 표준 매핑은 어긋난다.

    표준 정의는 io_tools_core.py 한 곳에 있고 여기서는 그 목록을 그대로
    복사해 대조만 한다.
    """
    if not os.path.isfile(config.IO_LIST):
        return True, "표준 IO List 미사용 (v1 원본으로 동작 중)"
    try:
        import openpyxl
    except ImportError:
        return False, "openpyxl 없음"
    from tools.make_io_list import STANDARD_ORDER
    ws = openpyxl.load_workbook(config.IO_LIST, read_only=True,
                                data_only=True).active
    rows = list(ws.iter_rows(values_only=True))
    hi = next((i for i, r in enumerate(rows)
               if r and "TAG" in [str(c).strip() if c else "" for c in r]), None)
    if hi is None:
        return False, "헤더 행을 찾지 못했습니다"
    hdr = [str(c).strip() if c else "" for c in rows[hi]]
    got = [h for h in hdr if h]
    front = got[:len(STANDARD_ORDER)]
    if front != STANDARD_ORDER:
        for i, (a, b) in enumerate(zip(front + [""] * 30, STANDARD_ORDER)):
            if a != b:
                return False, ("%d번 열이 표준과 다릅니다: '%s' ≠ '%s' — "
                               "완전일치 매핑에서 빗나갑니다" % (i + 1, a, b))
        return False, "표준 24종 열 수가 맞지 않습니다 (%d개)" % len(front)
    ext = got[len(STANDARD_ORDER):]
    if ext:
        return False, ("IO List 에 표준 밖의 열이 있습니다: %s — 계기 사양은 "
                       "INSTRUMENT_LIST.xlsx 에 두십시오"
                       % ", ".join(ext[:5]))
    return True, "표준 24종만 (확장 열 없음)"


def c_tag_cross_consistency():
    """
    [주입] 세 문서의 태그 교차 대조가 실제로 어긋남을 잡는가.

    IO List·계기 리스트·인터락 리스트는 TAG 로 서로를 가리킨다. 한쪽만
    고치면 조회가 조용히 빈다. 이 항목은 **지적 건수를 판정하지 않는다** —
    지금 데모 데이터에는 실제로 어긋남이 있고 그건 데이터의 문제다.
    여기서 보는 것은 대조기가 살아 있는지다.

    태그 하나의 표기를 바꿔 넣고 '표기 불일치' 로 잡히는지 확인한다.
    """
    from ingest import tag_registry as TR
    base = TR.cross_check()
    base_t = TR.collect()
    if base["counts"]["io"] == 0:
        return False, "IO List 태그를 하나도 읽지 못했습니다"
    if base["counts"]["spec"] == 0:
        return False, "계기 리스트 태그를 하나도 읽지 못했습니다"

    # 주입 대상은 **IO List 와 계기 리스트 양쪽에 있는** 태그여야 한다.
    # 첫 태그를 그냥 집으면 실물에서는 판넬 상태 접점(FAB_CPU_XA_0001)이
    # 걸리는데, 이것은 계기 리스트에 없으므로 바꿔치기해도 대조에 걸릴
    # 것이 없어 시험이 거짓 실패한다.
    spec_tags = set(base_t["spec"]) if base_t else set()
    real = None
    for r in TR.read_rows(config.IO_LIST):
        t0 = str(r.get("TAG") or "").strip()
        if t0 and t0 in spec_tags:
            real = t0
            break
    if not real:
        return True, ("IO List 와 계기 리스트에 함께 있는 태그가 없어 "
                      "주입 시험은 건너뜁니다 — 현재 지적 %d건"
                      % base["total"])

    saved = TR.collect
    mangled = real.replace("-", "").lower()

    def fake():
        t = saved()
        t["spec"] = [mangled if x == real else x for x in t["spec"]]
        return t

    TR.collect = fake
    try:
        got = TR.cross_check()
    finally:
        TR.collect = saved

    hits = [f for f in got["findings"].get("표기 불일치", [])
            if f["tag"] in (real, mangled)]
    if not hits:
        return False, ("'%s' 를 '%s' 로 바꿨는데 표기 불일치로 잡지 "
                       "못했습니다" % (real, mangled))
    return True, ("IO %d · 계기 %d · 인터락 조건 %d / 현재 지적 %d건, "
                  "주입 검출 정상"
                  % (base["counts"]["io"], base["counts"]["spec"],
                     base["counts"]["interlock_input"], base["total"]))


def c_instrument_form():
    """
    [주입] 계기 리스트가 실물 양식(2단 머리)인가.

    한때 이 문서를 제가 임의로 만든 18열 평면 양식으로 두었다. 실물은
    TAG NO. · DESCRIPTION · Q'TY · SENSOR TYPE · MATERIAL(ELEMENT/BODY) ·
    SCALE RANGE(MIN/MAX/UNIT) … 형태의 2단 머리이고, **배선(랙·슬롯·
    채널)은 들어가지 않는다** — IO List 에서 받아간다.

    모양이 다르면 실물을 그대로 갈아 끼울 수 없다.
    """
    spec = getattr(config, "INSTRUMENT_SPEC", "")
    if not spec or not os.path.isfile(spec):
        return False, "INSTRUMENT_LIST.xlsx 가 없습니다"
    from tools.make_io_list import INSTRUMENT_HEAD
    from ingest.lists import read_instrument_rows
    import openpyxl
    ws = openpyxl.load_workbook(spec, read_only=True, data_only=True).active
    rows = list(ws.iter_rows(values_only=True))
    hi = next((i for i, r in enumerate(rows)
               if r and any(str(c).strip().upper() == "TAG NO."
                            for c in r if c)), None)
    if hi is None:
        return False, "TAG NO. 머리 행을 찾지 못했습니다 — 실물 양식이 아닙니다"

    # 실물 머리 항목은 두 줄로 나뉘어 있다('SENSOR\nTYPE'). 눈으로는 같고
    # 문자열로는 다르므로, 비교 전에 공백을 하나로 접는다.
    groups = [re.sub(r"\s+", " ", str(c).strip()) for c in rows[hi] if c]
    want = [g for g, _sub in INSTRUMENT_HEAD]
    # 실물이 표준보다 열을 더 갖고 있는 것은 문제가 아니다 — 실제 문서에는
    # 뒤에 P&ID 열이 하나 더 붙어 있다. 표준 항목이 **앞에서부터 순서대로
    # 전부 있는지**만 본다. 모자라거나 순서가 어긋나면 갈아 끼울 수 없다.
    if groups[:len(want)] != want:
        for a, b in zip(groups + [""] * 30, want):
            if a != b:
                return False, "머리 항목이 다릅니다: '%s' ≠ '%s'" % (a, b)
        return False, "머리 항목 수가 모자랍니다 (%d개)" % len(groups)
    extra = groups[len(want):]

    # 배선 열이 섞여 들어오지 않았는지
    flat = {str(c).strip().upper() for c in rows[hi] if c}
    flat |= {str(c).strip().upper()
             for c in (rows[hi + 1] if hi + 1 < len(rows) else []) if c}
    wired = flat & {"RACK", "SLOT", "CH", "PANEL", "PN(DP)", "ADD"}
    if wired:
        return False, ("계기 리스트에 배선 열이 있습니다: %s — IO List 에서 "
                       "받아가야 합니다" % ", ".join(sorted(wired)))

    n = len(read_instrument_rows(spec))
    if n == 0:
        return False, "행을 하나도 읽지 못했습니다 (2단 머리 해석 실패)"
    tail = (" · 추가 열 %s" % ", ".join(extra)) if extra else ""
    return True, ("실물 양식 %d항목 / %d행, 배선 열 없음%s"
                  % (len(want), n, tail))


def c_pid_rule_mapping():
    """[주입] IO 심볼명 → P&ID 태그 규칙 매핑이 호기를 구별하는가.

    두 문서는 이름 공간이 다르다(IO: DWP_LS_P5401A_H, 계기: LS-P5401A/B).
    유사도로만 잇던 때는 호기를 못 가려 UPWT_LS_P3601C 가 LS-P3601A 에
    붙었다. C호기를 물었는데 A호기 사양이 나오고, 화면에는 그럴듯하게
    보이므로 사람이 알아채지 못한다. 여기서는 두 가지를 본다.
      · 규칙으로 이어진 건이 실제로 있는가
      · 호기 문자가 붙은 IO 태그는 같은 호기의 계기에 붙었는가
    """
    from ingest.lists import load_points, structural_pid_candidates
    pts = load_points(config.IO_LIST,
                      getattr(config, "INSTRUMENT_SPECS", None)
                      or getattr(config, "INSTRUMENT_SPEC", None))
    ruled = [(t, r) for t, r in pts.items() if r.get("_pid_src") == "rule"]
    if not ruled:
        # 자료에 따라 규칙 매핑이 애초에 필요 없다. IO List 와 계기 리스트가
        # 같은 태그 체계를 쓰면(AIT-1001 ↔ AIT-1001) 그냥 이름으로 조인된다.
        # 규칙은 두 이름 공간이 다를 때(DWP_LS_P5401A_H ↔ LS-P5401A/B) 쓰인다.
        direct = sum(1 for _t, r in pts.items()
                     if str(r.get("MODEL") or "").strip()
                     and r.get("_pid_src") in (None, "", "fallback_io"))
        if direct:
            return True, ("IO 태그와 계기 태그가 같은 체계라 규칙 매핑이 "
                          "필요 없습니다 — 이름으로 직접 조인 %d건" % direct)
        return False, ("규칙으로 이어진 태그가 하나도 없습니다 — "
                       "IO 심볼명과 계기 태그가 유사도에만 의존합니다")

    bad = []
    for t, r in ruled:
        cands = structural_pid_candidates(t)
        if not cands:
            bad.append("%s: 후보를 다시 만들지 못함" % t)
            continue
        want = cands[0].upper()
        got = str(r.get("P&ID TAG") or "").upper()
        if got != want:
            bad.append("%s: %s ≠ %s" % (t, got, want))
    if bad:
        return False, "호기 불일치 %d건 — %s" % (len(bad), bad[0])

    # 붙은 사양이 실제로 그 계기의 것인지 (원문 줄과 대조)
    joined = sum(1 for _t, r in ruled if str(r.get("MODEL") or "").strip())
    return True, ("규칙 %d건 · 호기 일치 · 사양 연결 %d건"
                  % (len(ruled), joined))


def c_advice_path_runs():
    """[주입] 조치 생성 경로가 LLM 없이도 끝까지 도는가.

    LLM 연결과는 별개로, 그래프가 예외 없이 완주하는지만 본다.
    device_of 가 기기 키를 여러 개 돌려주도록 바뀌었을 때 retrieve 단계가
    집합 판정에서 터졌고, 화면에는 예외 문자열 한 줄만 뜬 채 조치 생성이
    통째로 멈췄다 — LLM 이 꺼져 있어도 여기까지는 반드시 돌아야 한다.
    """
    from graph.app_graph import Copilot2
    from ingest.lists import load_points
    pts = load_points(config.IO_LIST,
                      getattr(config, "INSTRUMENT_SPECS", None)
                      or getattr(config, "INSTRUMENT_SPEC", None))
    tags = [t for t, r in pts.items() if r.get("MODEL")][:3] or list(pts)[:3]
    if not tags:
        return False, "IO 점을 읽지 못했습니다"
    prev = config.LLM_PROVIDER
    try:
        config.LLM_PROVIDER = "off"
        c = Copilot2(mode="lexical")
        for t in tags:
            out = c.answer(tag=t, alarm="alarm")
            if not isinstance(out, dict) or "decision" not in out:
                return False, "%s: 응답 형태가 아닙니다" % t
    except Exception as e:                                  # noqa: BLE001
        return False, "%s: %s" % (type(e).__name__, e)
    finally:
        config.LLM_PROVIDER = prev
    return True, "태그 %d건 완주 (LLM 없이)" % len(tags)


def c_attr_source_real():
    """
    [주입] TYPE·FAIL POSITION 이 실물 문서에서 오는가.

    두 항목은 한때 내가 만든 부속 데이터에만 있었다. 실물 전환 시 채울
    곳이 없는 값이라, 데모에서는 그럴듯하게 보이고 실물에서는 통째로
    비는 유형이다. 출처를 확인했다.

        TYPE           계기 리스트 SENSOR TYPE
        FAIL POSITION  인터락 리스트 "* Valve Action : Fail Open"

    부속 데이터의 값은 그 문서에서 못 읽었을 때만 쓰는 대체값이므로,
    실물 문서 값이 이겨야 한다.
    """
    from ingest.lists import read_instrument_rows
    from ingest.interlock import load_interlocks
    from retrieval.interlock_index import load_outputs

    # TYPE — 계기 리스트에서 온 값이 그대로 실리는가
    spec = {str(r.get("TAG") or "").strip(): r
            for r in read_instrument_rows(getattr(config, "INSTRUMENT_SPECS", None) or getattr(config, "INSTRUMENT_SPEC", None))}
    outs = load_outputs()
    checked = 0
    for tag, r in spec.items():
        want = (r.get("MEAS TYPE") or "").strip()
        if not want or tag not in outs:
            continue
        got = (outs[tag].get("type") or "").strip()
        if got != want:
            return False, ("%s TYPE 이 계기 리스트와 다릅니다: '%s' ≠ '%s'"
                           % (tag, got, want))
        checked += 1
    if checked == 0:
        return False, "계기 리스트에서 SENSOR TYPE 을 하나도 읽지 못했습니다"

    # FAIL POSITION — 인터락 리스트에 적힌 값이 이기는가
    fails = {it["output_tag"]: it["fail"] for it in load_interlocks()
             if it.get("fail")}
    for tag, want in fails.items():
        got = (outs.get(tag) or {}).get("fail") or ""
        if got.strip().upper() != want.strip().upper():
            return False, ("%s FAIL POSITION 이 인터락 리스트와 다릅니다: "
                           "'%s' ≠ '%s'" % (tag, got, want))
    return True, "TYPE %d건 계기 리스트 / FAIL POSITION %d건 인터락 리스트" % (
        checked, len(fails))


def c_no_attr_file():
    """
    [주입] 부속 데이터(TAG_ATTRIBUTES.xlsx)가 되살아나지 않았는가.

    네 항목이 모두 업로드 자료에서 오게 됐다.

        TERMINAL       TB List
        TYPE           계기 리스트 SENSOR TYPE
        FAIL POSITION  인터락 리스트 "* Valve Action"
        MANUAL FILE    MODEL ↔ 매뉴얼 파일명 대조

    내가 만든 열이 하나라도 되살아나면 데모에서는 채워지고 실물에서는
    비는 항목이 다시 생긴다. 그 차이는 화면에 드러나지 않는다.

    파일이 없는지만 보지 않고, **네 항목이 실제로 채워지는지**까지
    확인한다. 파일만 없애고 값이 비면 고친 것이 아니다.
    """
    for d in (config.DERIVED_DIR, config.DATA_DIR):
        p = os.path.join(d, "TAG_ATTRIBUTES.xlsx")
        if os.path.isfile(p):
            return False, "부속 데이터가 되살아났습니다: %s" % p
    from ingest.lists import load_points
    pts = load_points(config.IO_LIST, getattr(config, "INSTRUMENT_SPECS", None) or getattr(config, "INSTRUMENT_SPEC", None), None,
                      getattr(config, "TB_LIST", None))
    n = len(pts) or 1
    got = {}
    for k in ("TERMINAL", "TYPE", "MANUAL FILE"):
        got[k] = sum(1 for r in pts.values() if str(r.get(k) or "").strip())
    if got["TERMINAL"] == 0:
        return False, ("단자가 하나도 없습니다 — TB List 를 넣거나 "
                       "python -m tools.make_tb_list 로 만드십시오")
    if got["TYPE"] == 0:
        return False, "계기 리스트에서 SENSOR TYPE 을 읽지 못했습니다"
    if got["MANUAL FILE"] == 0:
        return False, ("매뉴얼 연결이 비었습니다 — MODEL 과 매뉴얼 파일명이 "
                       "대조되지 않습니다")
    from retrieval.interlock_index import load_outputs
    fails = sum(1 for v in load_outputs().values() if v.get("fail"))
    if fails == 0:
        return False, ("고장 위치가 비었습니다 — 인터락 리스트에 "
                       "\"* Valve Action\" 표기가 없습니다")
    return True, ("부속 없음 / 단자 %d · 종류 %d · 매뉴얼 %d · 고장위치 %d"
                  % (got["TERMINAL"], got["TYPE"], got["MANUAL FILE"], fails))


def c_tb_list():
    """
    [주입] TB 리스트가 IO List 와 같은 프로젝트 것인가.

    TB 리스트는 격자 배치 결선표라 읽기만 성공해도 태그가 하나도 안
    겹칠 수 있다. 실제로 다른 프로젝트 TB 리스트를 넣었더니 2,254점을
    정상적으로 읽고도 단자 조회가 통째로 비었다. 읽기 성공과 쓸모 있음은
    다르다.

    TB 리스트가 없으면 통과다 — 부속 데이터의 값으로 동작한다.
    """
    tb_path = getattr(config, "TB_LIST", "")
    if not tb_path or not os.path.isfile(tb_path):
        return True, "TB 리스트 없음 — 부속 데이터의 단자 값을 사용"
    from ingest.tb_list import load_terminals
    from ingest.lists import read_rows
    tb = load_terminals(tb_path)
    if not tb:
        return False, "TB 리스트를 읽었으나 태그가 하나도 없습니다"
    io_tags = {str(r.get("TAG") or "").strip()
               for r in read_rows(config.IO_LIST)}
    io_tags.discard("")
    hit = len(set(tb) & io_tags)
    if hit == 0:
        return False, ("TB 리스트 %d점과 IO List %d점이 하나도 겹치지 "
                       "않습니다 — 다른 프로젝트 자료로 보입니다. 단자 "
                       "조회가 통째로 빕니다." % (len(tb), len(io_tags)))
    ratio = hit / max(1, len(io_tags))
    if ratio < 0.5:
        return False, ("IO List %d점 중 %d점만 TB 리스트에 있습니다 "
                       "(%.0f%%) — 나머지는 단자 조회가 빕니다."
                       % (len(io_tags), hit, ratio * 100))
    return True, "TB %d점 / IO List 대조 %d점 일치 (%.0f%%)" % (
        len(tb), hit, ratio * 100)


def c_rack_unique():
    """
    (PLC, RACK, SLOT) 이 카드를 유일하게 가리키는가.

    RACK 이 전 행 같은 값이면 슬롯 번호가 판넬마다 겹쳐 카드를 구분할 수
    없다. 실제로 v1 데모 데이터가 그랬고, 그래서 카드 ID 에 판넬을 섞어
    쓰고 있다. 겹침이 있으면 실패가 아니라 경고로 남긴다 — 카드 ID 는
    판넬을 포함하므로 동작에는 지장이 없다.
    """
    from retrieval.panel_index import PanelIndex
    from collections import defaultdict
    ix = PanelIndex()
    seen = defaultdict(set)
    for r in ix.rows:
        seen[(r["PLC"], str(r.get("PN(DP)", "")), str(r["RACK"]),
              str(r["SLOT"]))].add(r["PANEL"])
    dup = {k: v for k, v in seen.items() if len(v) > 1}
    if dup:
        k, v = sorted(dup.items())[0]
        return False, ("스테이션·랙 번호가 카드를 구분하지 못합니다 — 예: "
                       "%s 가 %s 에 동시 존재 (겹침 %d건). IO List 의 "
                       "PN(DP)·RACK 을 실제 값으로 채우십시오."
                       % ("/".join(k), ", ".join(sorted(v)), len(dup)))
    return True, "카드 %d장이 (PLC, PN(DP), RACK, SLOT) 으로 유일" % len(
        ix.cards())


def c_arrangement_fresh():
    """
    [주입] 배치도의 판넬별 점수가 현재 계기 리스트와 맞는가.

    배치도는 생성물이라 계기 리스트가 바뀌면 다시 만들어야 한다.
    안 만들면 도면에는 옛 점수가, 조회에는 새 점수가 나오는데 둘 다
    그럴듯해서 어긋났다는 사실이 드러나지 않는다. 실제로 계기 4점이
    추가됐을 때 도면만 옛 값에 머물러 있었다.
    """
    from retrieval.panel_index import PanelIndex, load_locations
    ix = PanelIndex()
    locs = load_locations()
    if not locs:
        return False, "PANEL_LOCATIONS.csv 없음 — tools/make_arrangement.py 실행"
    bad = []
    for p in ix.panels():
        rec = locs.get(p["panel"])
        if rec is None:
            bad.append("%s 배치 정보 없음" % p["panel"])
            continue
        try:
            filed = int(rec.get("points") or -1)
        except (TypeError, ValueError):
            filed = -1
        if filed != p["points"]:
            bad.append("%s 도면 %s점 ≠ 리스트 %d점"
                       % (p["panel"], filed, p["points"]))
    if bad:
        return False, ("; ".join(bad)
                       + " — tools/make_arrangement.py 를 다시 실행하십시오")
    return True, "판넬 %d개 점수 일치 (총 %d점)" % (
        len(locs), sum(p["points"] for p in ix.panels()))


def c_panel_drawing():
    """배치도 PDF 가 있고, 판넬명이 그 안에서 실제로 찾아지는가."""
    if not os.path.exists(config.ARRANGEMENT_PDF):
        return False, ("배치도 없음 — python tools/make_arrangement.py "
                       "(판넬 도면 보기가 통째로 404 가 됩니다)")
    try:
        import pymupdf as fitz
    except ImportError:
        import fitz
    from retrieval.panel_index import PanelIndex
    ix = PanelIndex()
    doc = fitz.open(config.ARRANGEMENT_PDF)
    miss = []
    for name, loc in ix.locations.items():
        page = doc[max(0, loc["page"] - 1)]
        if not page.search_for(loc["find"]):
            miss.append(name)
    doc.close()
    if miss:
        return False, ("도면에서 못 찾는 판넬: %s — 하이라이트가 안 걸립니다"
                       % ", ".join(miss))
    return True, "판넬 %d개 모두 도면에서 검색됨" % len(ix.locations)


def c_panel_no_verdict():
    """
    [주입] 판넬 상실에 트립 여부를 단정하지 않는가.

    대체값 정책은 리스트에 없다. 그런데 화면이 '트립됩니다' 라고 쓰면
    사용자는 그걸 근거로 판단한다. 근거 없는 단정이 새로 들어오면
    여기서 걸린다.
    """
    from retrieval.panel_index import PanelIndex
    ix = PanelIndex()
    for c in ix.cards()[:5]:
        txt = ix.render_impact(c["card"])
        bad = [w for w in ("트립됩니다", "트립된다", "정지합니다", "동작합니다")
               if w in txt]
        if bad:
            return False, "%s 서술에 단정 표현: %s" % (c["card"],
                                                      ", ".join(bad))
        if "판정하지 않" not in txt:
            return False, ("%s 서술에서 판정 유보 문구가 사라짐 — 근거 없는 "
                           "결론으로 읽힙니다" % c["card"])
    return True, "카드 서술 의존 관계만, 트립 판정 없음"


def c_manual_vocab_not_blocked():
    """
    [주입] 매뉴얼 용어를 태그 오타로 오인해 질문을 막지 않는가.

    "없는 태그는 거절한다" 를 넓히면 반대쪽이 뚫린다. 태그 정규식은
    "PCS 7", "TB2", "AI 16xI", "bge-m3" 같은 기술 용어도 잡는데,
    이걸 태그 오타로 처리하면 매뉴얼 질문이 통째로 거절된다.
    "AI 16xI 모듈 상태 표시" 는 실제 매뉴얼 6.1 절 제목이다.

    반대 방향도 함께 본다 — 낯선 접두어라고 현재 화면 태그로
    대체해서도 안 된다.
    """
    from api.server import find_tag, rule_intent, known_tag_prefixes
    pres = known_tag_prefixes()
    if not pres:
        return False, "태그 접두어를 뽑지 못했습니다"

    # 1) 매뉴얼 용어가 거절되지 않아야 한다
    for q in ("AI 16xI 모듈 상태 표시 설명해줘", "PCS 7 알람 확인 방법",
              "TB2 단자대가 뭐야", "NAMUR NE43 이 뭐야"):
        _, miss = find_tag(q)
        if miss:
            return False, "'%s' 를 태그 오타(%s)로 오인 — 매뉴얼 질문이 막힙니다" % (q, miss)

    # 2) 낯선 접두어는 현재 태그로 대체하지 않아야 한다
    cur = "AIT-4002"
    got, _ = find_tag("zzt-9999는 어디야", cur)
    if got == cur:
        return False, "낯선 접두어 질의가 현재 태그(%s)로 대체됨" % cur
    r = rule_intent("zzt-9999는 어느 위치에 있어?", cur)
    if cur in (r.get("reply") or ""):
        return False, "없는 태그 위치 질문에 현재 태그 답이 나감"

    # 3) 실재 접두어 오타는 여전히 거절해야 한다
    _, miss = find_tag("AIT-9999 알람 조회해줘", cur)
    if not miss:
        return False, "AIT-9999 를 거절하지 못함 — 없는 태그가 통과합니다"
    return True, "접두어 %d종 기준 / 매뉴얼 용어 통과·없는 태그 거절" % len(pres)


def c_smalltalk():
    """
    [주입] 인사에 예시 목록이나 매뉴얼 거절을 돌려주지 않는가.

    "안녕하세요" 에 "예) AIT-4002 low acid 알람 조회해줘 …" 가 나갔고,
    이어진 "인사 안 해주고 예시를 들어주네?" 는 매뉴얼 검색으로 넘어가
    "근거를 찾지 못했습니다" 가 나왔다. 둘 다 사람 눈에는 명백한데
    지표로는 안 잡힌다.

    인사가 조회 의도를 가로채지 않는지도 함께 본다.
    """
    from api.server import rule_intent
    for greet in ("안녕?", "안녕하세요", "ㅎㅇ", "고마워"):
        r = rule_intent(greet)
        rep = r.get("reply") or ""
        if r.get("type") != "chat":
            return False, "'%s' 가 %s 로 분류됨" % (greet, r.get("type"))
        if "예)" in rep or "예:" in rep:
            return False, "'%s' 에 예시 목록이 나갑니다" % greet
        if r.get("generic"):
            return False, ("'%s' 가 포괄 응답으로 떨어져 매뉴얼 검색으로 "
                           "넘어갑니다" % greet)
    # 인사가 조회를 삼키지 않아야 한다
    r = rule_intent("안녕하세요 그런데 AIT-1001 판넬 어디야")
    if r.get("type") != "panel":
        return False, "인사말이 붙은 조회 요청을 %s 로 삼킴" % r.get("type")
    return True, "인사·감사 4종 응대 / 조회 요청은 그대로 통과"


def c_tag_particle():
    """
    [주입] 조사가 붙은 태그를 읽고, 못 읽었을 때 다른 태그로 대체하지 않는가.

    "ait-1001은 어디야" 처럼 조사가 붙으면 \\b 경계가 깨져 태그를 놓치고,
    그러면 현재 선택된 태그로 조용히 대체되어 **다른 설비의 답**이
    나간다. 화면에는 그럴듯한 문장이 떠서 틀렸다는 사실이 드러나지
    않는다. 실제로 한 번 나갔던 고장이다.
    """
    from api.server import find_tag, rule_intent, get_panel
    px = get_panel()
    if px is None:
        return False, "판넬 인덱스 미적재"
    real = sorted(px._by_tag)[0]
    other = sorted(px._by_tag)[-1]
    for suffix in ("은", "는", "이", "가", "의", "에서"):
        got, _ = find_tag("%s%s 어디야" % (real.lower(), suffix), other)
        if got != real:
            return False, ("'%s%s' 를 %s 로 읽음 — 조사가 붙으면 태그를 "
                           "놓칩니다" % (real, suffix, got))
    # 없는 태그를 물었을 때 현재 태그로 대신 답하지 않는가
    r = rule_intent("zzt-9999는 어느 위치에 있어?", other)
    if other in (r.get("reply") or ""):
        return False, "없는 태그 질문에 현재 태그(%s)로 대신 답함" % other
    return True, "조사 6종 복원 / 미해결 태그는 대체하지 않음"


def c_panel_chat():
    """판넬 명령이 규칙 엔진에서 인식되는가 (표기가 태그와 겹침)."""
    from api.server import rule_intent, get_panel
    px = get_panel()
    if px is None:
        return False, "판넬 인덱스 미적재"
    name = sorted(px._by_panel)[0]
    r = rule_intent("%s 내리면 뭐가 죽어?" % name)
    if r.get("type") != "panel":
        return False, "'%s 내리면' 이 %s 로 잘못 분류됨" % (name, r.get("type"))
    # 태그 의도가 판넬 규칙에 먹히지 않았는지도 함께 본다
    r2 = rule_intent("XV-4101 인터락 조회해줘")
    if r2.get("type") != "interlock":
        return False, "판넬 규칙이 인터락 명령을 가로챔 (%s)" % r2.get("type")
    return True, "판넬 %s / 인터락 명령 각각 정상" % name


# ── 6. 리포트 ───────────────────────────────────────────────
def c_report_font():
    """4D 리포트 한글이 깨지지 않는가."""
    from api.report_4d import _register_fonts
    _register_fonts()
    p = os.environ.get("PMC_KR_FONT")
    if not p:
        return False, ("한글 폰트를 찾지 못해 Helvetica 로 떨어집니다 — "
                       "PDF 한글이 통째로 깨집니다")
    return True, p


# ── 실행 ────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser(description="시연 전 전 경로 점검")
    ap.add_argument("--skip-llm", action="store_true",
                    help="모델 호출 없이 구조만 점검")
    args = ap.parse_args()

    print("\nPlant Maintenance Copilot — 시연 전 점검\n" + "=" * 78)

    run("검색 모드 구성", c_search_modes)
    run("한글 질의 도달", c_korean_query)
    run("근거 없을 때 거절", c_abstain_works)
    run("강등 가시성 [주입]", c_degrade_visible)
    run("다양성 설정", c_diversify_off, critical=False)
    run("임베딩 캐시 신원 [주입]", c_cache_identity)
    run("환각 차단 [주입]", c_advisor_rejects_fake)
    run("챗봇·조치 경로 일치", c_chat_gateway)
    run("규칙 엔진 명령 인식", c_chat_rule)
    run("태그 표기 복원", c_tag_notation)
    run("챗봇 명령 실행 [주입]", c_chat_dispatch)
    run("챗봇 근거 답변 [주입]", c_chat_qa)
    run("챗봇 후속 질문", c_chat_followup)
    run("인터락 72문항", c_interlock_eval)
    run("판넬 평가셋", c_panel_eval)
    run("카드 단위 분리 [주입]", c_card_scope)
    run("카드 채널 노출 [주입]", c_card_channels_visible)
    run("공통원인 점검 [주입]", c_common_cause)
    run("판넬 상실 미제공", c_no_panel_impact)
    run("data 폴더 정결 [주입]", c_data_dir_clean)
    run("표준 IO List 원천 대조 [주입]", c_io_list_standard)
    run("IO List 표준 헤더 [주입]", c_io_list_header)
    run("태그 교차 정합성 [주입]", c_tag_cross_consistency)
    run("계기 리스트 실물 양식 [주입]", c_instrument_form)
    run("사양 출처 실물 문서 [주입]", c_attr_source_real)
    run("P&ID 규칙 매핑 [주입]", c_pid_rule_mapping)
    run("조치 생성 경로 [주입]", c_advice_path_runs)
    run("부속 데이터 폐지 [주입]", c_no_attr_file)
    run("TB 리스트 대조 [주입]", c_tb_list)
    run("스테이션·랙 유일성", c_rack_unique)
    # 도면은 실물 반입이 어려워 데모 도면을 그대로 쓰기로 했다. 실물 IO List
    # 와 데모 배치도를 함께 쓰면 판넬명이 어긋나는 것이 당연하므로, 이 항목은
    # 시연을 막는 실패가 아니라 주의로 둔다 — 데모 세트로 돌리면 그대로 통과한다.
    run("배치도 최신 [주입]", c_arrangement_fresh, critical=False)
    run("판넬 배치도 검색", c_panel_drawing)
    run("카드 트립 단정 금지 [주입]", c_panel_no_verdict)
    run("판넬 명령 인식", c_panel_chat)
    run("인사 응대 [주입]", c_smalltalk)
    run("태그 조사 표기 [주입]", c_tag_particle)
    run("매뉴얼 용어 오인 금지 [주입]", c_manual_vocab_not_blocked)
    run("한글 PDF 폰트", c_report_font, critical=False)
    if not args.skip_llm:
        run("조치 생성 모델 연결", c_advisor_reachable, critical=False)
        run("조치 생성 실행", c_advisor_real, critical=False)

    w = max(len(r[1]) for r in _rows)
    for st, name, detail in _rows:
        print("[%s] %-*s  %s" % (st, w, name, detail))
    print("=" * 78)

    bad = [r for r in _rows if r[0] == "실패"]
    warn = [r for r in _rows if r[0] == "주의"]
    if bad:
        print("실패 %d건 — 이 상태로 시연하면 해당 경로가 조용히 무너집니다." % len(bad))
        return 1
    if warn:
        print("필수 경로 정상. 주의 %d건은 화면 품질에 영향합니다." % len(warn))
    else:
        print("전 경로 정상. 고장 주입 검사도 통과했습니다.")
    return 0


if __name__ == "__main__":
    sys.exit(main())

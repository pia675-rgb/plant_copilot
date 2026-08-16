#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FastAPI 서버 — React UI용 백엔드 (v1 기능 포함)

실행:
    uvicorn api.server:app --reload --port 8000
"""

from __future__ import annotations

import csv
import datetime as dt
import json
import os
import re
import sys
from typing import Any, Dict, List, Optional

from fastapi import FastAPI, HTTPException, Query
from fastapi.responses import Response
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook
from pydantic import BaseModel, Field

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import config  # noqa: E402
from graph.app_graph import Copilot2  # noqa: E402
from retrieval.pipeline import Retriever  # noqa: E402
from retrieval.interlock_index import InterlockIndex  # noqa: E402
from retrieval.panel_index import PanelIndex  # noqa: E402

# 기본 모드. lexical 은 한글 질의를 구조적으로 못 푼다 — 시연 기본값으로 쓰면
# 가장 약한 구성을 심사위원에게 먼저 보여주게 된다. 환경변수로 덮을 수 있다.
DEFAULT_MODE = os.environ.get("COPILOT_UI_MODE", "hybrid")

app = FastAPI(title="Plant Maintenance Copilot v2", version="2.2")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], allow_credentials=True,
    allow_methods=["*"], allow_headers=["*"],
)


@app.on_event("startup")
def _prewarm_llm():
    """모델을 미리 올려둔다. 첫 조회가 로딩 시간을 물지 않게 한다.

    별도 스레드로 돌린다 — 예열이 끝날 때까지 서버 기동을 붙잡으면
    화면이 늦게 뜬다. COPILOT_PREWARM=0 으로 끌 수 있다.
    """
    if os.environ.get("COPILOT_PREWARM", "1").strip() == "0":
        return
    import threading
    from graph import advisor

    def run():
        ok, msg = advisor.prewarm()
        print("[prewarm] %s" % msg)

    threading.Thread(target=run, daemon=True).start()

# ── 싱글톤 ──────────────────────────────────────────────────
_retrievers: Dict[str, Retriever] = {}
_copilots: Dict[str, Copilot2] = {}
_interlock: Optional[InterlockIndex] = None
_panel: Optional[PanelIndex] = None
_panel_error: Optional[str] = None
_instruments: Optional[Dict[str, dict]] = None
_history: Optional[List[dict]] = None
_drawings: Optional[Dict[str, list]] = None


def get_retriever(mode: str = DEFAULT_MODE) -> Retriever:
    if mode not in _retrievers:
        _retrievers[mode] = Retriever(mode=mode)
    return _retrievers[mode]


def get_copilot(mode: str = DEFAULT_MODE) -> Copilot2:
    if mode not in _copilots:
        _copilots[mode] = Copilot2(mode=mode)
    return _copilots[mode]


def get_interlock() -> InterlockIndex:
    global _interlock
    if _interlock is None:
        _interlock = InterlockIndex()
    return _interlock


def get_panel() -> Optional[PanelIndex]:
    """
    계기 리스트의 PANEL 열 + 배치 CSV. 인터락 인덱스를 넘겨 재사용한다.

    실패하면 조용히 None 을 돌려주는 대신 사유를 남기고 /api/health 에
    드러낸다. 판넬 조회가 통째로 안 뜨는데 이유를 알 수 없는 상태가
    이 프로젝트에서 반복해서 나온 고장 유형이다.
    """
    global _panel, _panel_error
    if _panel is None and _panel_error is None:
        try:
            _panel = PanelIndex(interlock=get_interlock())
        except Exception as e:                              # noqa: BLE001
            _panel_error = str(e)
            print("[panel] 적재 실패:", e)
    return _panel


def load_instruments() -> Dict[str, dict]:
    """
    P&ID TAG 기준 기기 딕셔너리.

    · 알람·매뉴얼·태그 목록·계기 상세 → 이 함수 (키 = P&ID TAG)
    · 판넬/카드/채널 배선 → retrieval.panel_index 가 load_points(IO TAG) 사용
    """
    global _instruments
    if _instruments is not None:
        return _instruments
    path = config.INSTRUMENTS
    if not path or not os.path.isfile(path):
        print("[api] IO List 없음:", path)
        _instruments = {}
        return _instruments
    try:
        from ingest.lists import load_pid_devices
        devices = load_pid_devices(
            path,
            getattr(config, "INSTRUMENT_SPECS", None) or getattr(config, "INSTRUMENT_SPEC", None),
            None,
            getattr(config, "TB_LIST", None),
        )
        clean = {}
        for pid, r in devices.items():
            row = {}
            for k, v in r.items():
                if k in ("io_tags", "io_points"):
                    row[k] = v
                else:
                    row[k] = "" if v is None else v
            clean[pid] = row
        _instruments = clean
        n_io = sum(len(r.get("io_tags") or []) for r in clean.values())
        print("[api] P&ID 기기 %d건 / IO 점 %d건 (%s)"
              % (len(_instruments), n_io, path))
        return _instruments
    except Exception as e:
        print("[api] IO List 로드 실패:", e)
        _instruments = {}
        return _instruments


_io_points = None


def load_io_points() -> Dict[str, dict]:
    """IO TAG → 배선 레코드 (알람 선택·판넬용)."""
    global _io_points
    if _io_points is not None:
        return _io_points
    path = config.INSTRUMENTS
    if not path or not os.path.isfile(path):
        _io_points = {}
        return _io_points
    try:
        from ingest.lists import load_points
        pts = load_points(
            path,
            getattr(config, "INSTRUMENT_SPECS", None) or getattr(config, "INSTRUMENT_SPEC", None),
            None,
            getattr(config, "TB_LIST", None),
        )
        clean = {}
        for tag, r in pts.items():
            row = {k: ("" if v is None else v) for k, v in r.items()
                   if k not in ("io_tags", "io_points")}
            # bool 유지
            if r.get("_spare"):
                row["_spare"] = True
            clean[tag] = row
        _io_points = clean
        print("[api] IO 점 %d건" % len(_io_points))
        return _io_points
    except Exception as e:
        print("[api] IO 점 로드 실패:", e)
        _io_points = {}
        return _io_points


def resolve_to_pid(tag: str):
    """IO TAG 또는 P&ID → (pid_tag, io_tag|None).
    도면·매뉴얼 검색은 항상 pid_tag 사용."""
    tag = (tag or "").strip()
    if not tag:
        return "", None
    io = load_io_points().get(tag)
    if io:
        pid = (io.get("P&ID TAG") or tag).strip()
        return pid, tag
    inst = load_instruments().get(tag)
    if inst:
        return tag, None
    for pid, r in load_instruments().items():
        if tag in (r.get("io_tags") or []):
            return pid, tag
    return tag, None


def load_history() -> List[dict]:
    global _history
    if _history is not None:
        return _history
    path = config.HISTORY
    if not os.path.exists(path):
        _history = []
        return _history
    with open(path, encoding="utf-8") as f:
        _history = json.load(f)
    return _history


def save_history(rows: List[dict]):
    global _history
    path = config.HISTORY
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)
    _history = rows



# 인터락 태그 적재 중 삼킨 오류. /api/health 로 드러낸다.
_interlock_tag_error = None


def load_drawings() -> Dict[str, list]:
    global _drawings
    if _drawings is not None:
        return _drawings
    path = config.DRAWINGS_INDEX
    out: Dict[str, list] = {}
    if not os.path.exists(path):
        _drawings = out
        return out
    with open(path, newline="", encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            tag = (r.get("TAG") or "").strip()
            if not tag:
                continue
            out.setdefault(tag, []).append({
                "type": r.get("TYPE") or "",
                "sheet_no": r.get("SHEET_NO") or "",
                "file": r.get("FILE") or "",
                "page": int(r["PAGE"]) if r.get("PAGE") else 0,
                "find": r.get("FIND") or tag,
            })
    for tag in out:
        out[tag].sort(key=lambda d: {"P&ID": 0, "SCHEMATIC": 1, "OUTLINE": 2,
                                     "ARRANGEMENT": 3}.get(d["type"], 9))
    _drawings = out
    return out


def load_output_tags() -> Dict[str, dict]:
    """
    출력 태그: 인터락이 동작시키는 대상.

    별도 output list 문서는 없다. IO List(DO/AO 점) + 계기 리스트 +
    부속 데이터를 합친 결과에서 인터락 출력에 해당하는 태그를 고른다.
    """
    from retrieval.interlock_index import load_outputs
    try:
        return load_outputs()
    except Exception as e:                                  # noqa: BLE001
        print("[api] 출력 태그 로드 실패:", e)
        return {}


# ── 모델 ────────────────────────────────────────────────────
class SearchRequest(BaseModel):
    tag: Optional[str] = None
    alarm: str = ""
    code: str = ""
    mode: str = Field(default=DEFAULT_MODE, pattern="^(lexical|hybrid|full)$")


class DiagnoseRequest(BaseModel):
    tag: Optional[str] = None
    alarm: str = ""
    code: str = ""
    mode: str = Field(default=DEFAULT_MODE, pattern="^(lexical|hybrid|full)$")


class InterlockRequest(BaseModel):
    tag: str
    action: Optional[str] = None
    as_input: bool = False


class FeedbackRequest(BaseModel):
    tag: str
    symptom: str = ""
    root_cause: str
    action_taken: str
    code_ref: str = ""
    manual_match: str = "부분일치"  # 일치 / 부분일치 / 불일치
    duration_min: int = 0
    parts: str = "-"
    tech: str = ""


class ChatContext(BaseModel):
    """화면에 떠 있는 조회 결과. 후속 질문에 답하려면 이것이 필요하다."""
    tag: Optional[str] = None
    alarm: str = ""
    decision: Optional[str] = None
    grade: Optional[float] = None
    evidence: List[dict] = []
    steps: List[dict] = []


class ChatRequest(BaseModel):
    message: str
    tag: str = ""
    tab: str = "alarm"
    use_llm: bool = True
    # 화면에 떠 있는 조회 결과. 후속 질문에 답하려면 필요하다.
    context: Optional[ChatContext] = None


class ReportRequest(BaseModel):
    tag: str
    alarm: str = ""
    code: str = ""
    mode: str = DEFAULT_MODE
    tech: str = ""
    confirmed_cause: str = ""
    final_action: str = ""
    parts: str = "-"
    duration_min: Optional[int] = None


class AdviceRequest(BaseModel):
    tag: Optional[str] = None
    alarm: str = ""
    code: str = ""
    mode: str = DEFAULT_MODE
    # 기본은 LLM 조치 생성. True 를 주면 근거 나열 템플릿만 쓴다.
    mock: bool = False


# ── 엔드포인트 ──────────────────────────────────────────────
@app.get("/api/health")
def health():
    """
    강등 여부를 여기서 드러낸다. 시연 중 Ollama 가 안 떠 있으면 UI 는
    'hybrid' 라고 표시하면서 실제로는 렉시컬로 도는데, 그 상태를 모르고
    "왜 한글이 안 먹지" 를 무대에서 디버깅하게 된다.
    """
    info = {"status": "ok", "version": "2.2", "default_mode": DEFAULT_MODE}
    try:
        r = get_retriever(DEFAULT_MODE)
        info["effective_mode"] = r.effective_mode
        info["degraded"] = r.degraded
        if r.degraded:
            info["degrade_detail"] = r.degrade_detail()
        info["label"] = r.label()
        if _interlock_tag_error:
            info["interlock_tag_error"] = _interlock_tag_error
        known_tags()
        if _known_tags_error:
            info["tag_index_error"] = _known_tags_error
        info["known_tags"] = len(known_tags())
        px = get_panel()
        if _panel_error:
            info["panel_index_error"] = _panel_error
        elif px is not None:
            info["panels"] = len(px.panels())
            info["arrangement_pdf"] = os.path.exists(config.ARRANGEMENT_PDF)
            if px.unlocated:
                info["panels_unlocated"] = px.unlocated
    except Exception as e:                                  # noqa: BLE001
        info["status"] = "degraded"
        info["error"] = str(e)
    return info


@app.get("/api/tags")
def list_tags(system: Optional[str] = None, q: Optional[str] = None,
              kind: Optional[str] = None):
    """kind=instrument|output|all(default).

    · instrument (알람) : **IO TAG** — DI/AI 입력점 (트립·상태·계측)
    · output (인터락)  : **P&ID TAG** — DO/AO 가 있는 출력 장비
    도면·매뉴얼은 항상 pid_tag 로 조회.
    """
    rows = []

    def _is_do_ao(io_type: str) -> bool:
        t = (io_type or "").strip().upper()
        return bool(t) and ("DO" in t or "AO" in t)

    def _is_di_ai(io_type: str) -> bool:
        t = (io_type or "").strip().upper()
        return bool(t) and ("DI" in t or "AI" in t)

    def _looks_like_io_tag(t: str) -> bool:
        import re as _re
        return bool(_re.match(r"^[A-Z][A-Z0-9]*_[A-Z0-9_]+$", (t or "").upper()))

    # ── 알람용: IO TAG (DI/AI) ──────────────────────────────
    if kind in (None, "all", "instrument"):
        for tag, r in sorted(load_io_points().items()):
            if r.get("_spare"):
                continue
            iot = str(r.get("IO TYPE") or "")
            if _is_do_ao(iot):
                continue
            if iot and not _is_di_ai(iot):
                continue
            if not iot:
                continue
            pid = (r.get("P&ID TAG") or tag)
            svc = r.get("DESCRIPTION") or r.get("SERVICE") or ""
            if q:
                hay = " ".join([tag, pid, svc, iot]).lower()
                if q.lower() not in hay:
                    continue
            rows.append({
                "tag": tag,
                "pid_tag": pid,
                "service": svc,
                "maker": r.get("MAKER") or "",
                "model": r.get("MODEL") or r.get("TYPE") or "",
                "signal": r.get("SIGNAL") or iot,
                "io_type": iot,
                "kind": "instrument",
                "io_count": 1,
            })

    # ── 인터락용: P&ID 출력 장비 ────────────────────────────
    if kind in (None, "all", "output"):
        for tag, r in sorted(load_instruments().items()):
            types = [str(r.get("IO TYPE") or "")]
            for p in (r.get("io_points") or []):
                types.append(str(p.get("io_type") or ""))
            is_out = any(_is_do_ao(t) for t in types)
            if not is_out:
                continue
            if _looks_like_io_tag(tag) and "-" not in tag.split("_")[0]:
                continue
            if q:
                hay = " ".join(str(x) for x in [
                    tag, r.get("SERVICE"), r.get("MODEL"),
                    " ".join(r.get("io_tags") or []),
                ]).lower()
                if q.lower() not in hay:
                    continue
            rows.append({
                "tag": tag,
                "pid_tag": r.get("P&ID TAG") or tag,
                "service": r.get("SERVICE") or "",
                "maker": r.get("MAKER") or "",
                "model": r.get("MODEL") or r.get("TYPE") or "",
                "signal": r.get("SIGNAL") or "",
                "kind": "output",
                "io_count": len(r.get("io_tags") or []),
            })

    return {"tags": rows, "systems": [], "count": len(rows)}


@app.get("/api/instrument/{tag:path}")
def instrument_detail(tag: str):
    pid, io_tag = resolve_to_pid(tag)
    try:
        inst = load_instruments().get(pid) or load_instruments().get(tag)
        io_rec = load_io_points().get(tag) or load_io_points().get(io_tag or "")
    except Exception as e:
        return {
            "tag": tag,
            "pid_tag": pid or tag,
            "instrument": {"tag": tag, "pid_tag": pid or tag},
            "drawings": [],
            "history": [],
            "history_count": 0,
            "warning": str(e),
        }
    if not inst and io_rec:
        # IO 만 있는 경우 — P&ID 메타는 최소
        inst = {
            "TAG": pid or tag,
            "P&ID TAG": pid or tag,
            "SERVICE": io_rec.get("DESCRIPTION") or io_rec.get("SERVICE") or "",
            "IO TYPE": io_rec.get("IO TYPE") or "",
            "PANEL": io_rec.get("PANEL") or "",
            "RACK": io_rec.get("RACK"),
            "SLOT": io_rec.get("SLOT"),
            "CH": io_rec.get("CH"),
            "io_tags": [tag],
            "io_points": [],
        }
    if not inst:
        return {
            "tag": tag,
            "pid_tag": pid or tag,
            "instrument": {"tag": tag, "pid_tag": pid or tag},
            "drawings": list(load_drawings().get(pid or tag, [])),
            "history": [h for h in load_history() if h.get("tag") in (tag, pid)],
            "history_count": 0,
        }
    pid = inst.get("P&ID TAG") or pid or tag
    io_tags = list(inst.get("io_tags") or [])
    io_points = list(inst.get("io_points") or [])
    dwgs = list(load_drawings().get(pid, [])) or list(load_drawings().get(tag, []))
    panel_loc = None
    px = get_panel()
    if px is not None:
        # 판넬 위치는 IO TAG 로 조회 (P&ID 가 아님)
        pt = None
        for iot in ([tag] + list(io_tags or [])):
            pt = px.by_tag(iot)
            if pt:
                break
        if pt:
            panel_loc = pt["location"]
            if pt["drawing"] and not any(d.get("type") == "ARRANGEMENT"
                                         for d in dwgs):
                dwgs.append(pt["drawing"])
    hist = [h for h in load_history()
            if h.get("tag") in (tag, pid) or h.get("tag") in io_tags]
    hist.sort(key=lambda x: x.get("date") or "", reverse=True)
    return {
        "tag": tag,
        "pid_tag": pid,
        "io_tags": io_tags,
        "io_points": io_points,
        "instrument": {
            "tag": tag,
            "pid_tag": pid,
            "service": inst.get("SERVICE") or "",
            "maker": inst.get("MAKER") or "",
            "model": inst.get("MODEL") or "",
            "meas_type": inst.get("MEAS TYPE") or "",
            "unit": inst.get("UNIT") or "",
            "signal": inst.get("SIGNAL") or "",
            "panel": inst.get("PANEL") or inst.get("PANEL NO") or "",
            "terminal": inst.get("TERMINAL") or inst.get("TB") or "",
            "plc": inst.get("PLC") or "",
            "slot": inst.get("SLOT") or "",
            "channel": inst.get("CHANNEL") or inst.get("CH") or "",
            "dwg_no": inst.get("DWG NO") or inst.get("DWG_NO") or "",
        },
        "panel_location": panel_loc,
        "drawings": dwgs,
        "history": hist,
        "history_count": len(hist),
    }


@app.get("/api/history/{tag}")
def history_by_tag(tag: str):
    hist = [h for h in load_history() if h.get("tag") == tag]
    hist.sort(key=lambda x: x.get("date") or "", reverse=True)
    return {"tag": tag, "history": hist, "count": len(hist)}


@app.post("/api/feedback")
def add_feedback(req: FeedbackRequest):
    rows = load_history()
    wo = "WO-%s-%03d" % (dt.date.today().strftime("%Y"), len(rows) + 1)
    rec = {
        "wo_no": wo,
        "date": dt.date.today().isoformat(),
        "tag": req.tag,
        "device": (load_instruments().get(req.tag) or {}).get("MODEL") or "",
        "symptom": req.symptom,
        "code_ref": req.code_ref,
        "first_action": "",
        "root_cause": req.root_cause,
        "action_taken": req.action_taken,
        "manual_match": req.manual_match,
        "duration_min": req.duration_min,
        "parts": req.parts or "-",
        "tech": req.tech or "",
    }
    rows.append(rec)
    save_history(rows)
    return {"ok": True, "record": rec}


@app.delete("/api/history/{wo_no}")
def delete_history(wo_no: str):
    """작업번호(wo_no)에 해당하는 보수 이력을 삭제한다. (legacy_v1 기능 복원)"""
    rows = load_history()
    new_rows = [r for r in rows if r.get("wo_no") != wo_no]
    if len(new_rows) == len(rows):
        raise HTTPException(404, "이력 %s 을(를) 찾을 수 없습니다." % wo_no)
    save_history(new_rows)
    return {"ok": True, "deleted": wo_no, "remaining": len(new_rows)}


@app.post("/api/search")
def search(req: SearchRequest):
    try:
        r = get_retriever(req.mode)
        query = (req.alarm + " " + req.code).strip()
        hits = r.retrieve(query, tag=req.tag)
        results = []
        for rec, score, trace in hits:
            src = rec.get("source") or {}
            results.append({
                "id": rec["id"],
                "kind": rec["kind"],
                "title": rec["title"],
                "text": rec["text"][:500],
                "score": round(float(score), 4),
                "trace": trace,
                "source": src,
                "device": rec.get("device", ""),
                "cite": "%s p.%s (%s)" % (
                    src.get("file", ""), src.get("pdf_page", ""), src.get("section", "")),
            })
        return {"query": query, "tag": req.tag, "mode": req.mode,
                "count": len(results), "results": results}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.post("/api/diagnose")
def diagnose(req: DiagnoseRequest):
    try:
        pid, io_tag = resolve_to_pid(req.tag)
        # 매뉴얼·검색은 P&ID, 응답에는 원본 IO TAG 유지
        search_tag = pid or req.tag
        out = get_copilot(req.mode).answer(
            tag=search_tag, alarm=req.alarm, code=req.code)
        evidence = []
        for e in out.get("evidence", [])[: config.FINAL_TOP_K]:
            evidence.append({
                "id": e.get("id"),
                "kind": e.get("kind"),
                "title": e.get("title"),
                "text": (e.get("text") or "")[:500],
                "score": e.get("score"),
                "cite": e.get("cite"),
                "source": e.get("source") or {},
                "summary_ko": "",
            })

        # 매뉴얼 근거 한국어 요약 (표시용). 실패해도 무시.
        #
        # 순차로 부르던 때는 요약 한 건에 15~40초가 걸려 조회 전체가 1~2분씩
        # 걸렸다. 검색·판정은 이미 1ms 안에 끝나 있는데 화면은 요약을 기다린다.
        # 요약끼리는 서로 의존하지 않으므로 동시에 던진다 — 대기 시간이 가장
        # 느린 한 건으로 줄어든다. COPILOT_SUMMARY=off 면 통째로 건너뛴다.
        if getattr(config, "SUMMARY_KO", True):
            try:
                from concurrent.futures import ThreadPoolExecutor
                from graph.advisor import summarize_ko
                targets = [it for it in evidence
                           if it.get("kind") == "manual_text"
                           ][: getattr(config, "SUMMARY_MAX", 3)]
                if targets:
                    with ThreadPoolExecutor(max_workers=len(targets)) as ex:
                        futs = [ex.submit(summarize_ko,
                                          it.get("text") or "",
                                          it.get("title") or "")
                                for it in targets]
                        for it, fu in zip(targets, futs):
                            try:
                                ko = fu.result()
                            except Exception:               # noqa: BLE001
                                ko = ""
                            if ko:
                                it["summary_ko"] = ko
            except Exception as _e:                         # noqa: BLE001
                print("[diagnose] summary_ko 생략:", _e)

        return {
            "decision": out.get("decision"),
            "grade": out.get("grade"),
            "grade_reason": out.get("grade_reason"),
            "trace": out.get("trace", []),
            "attempts": out.get("attempts"),
            "rewrites": out.get("rewrites", []),
            "evidence": evidence,
            "mode": req.mode,
            "tag": req.tag,
            "pid_tag": search_tag,
            "io_tag": io_tag,
            "query": out.get("query") or req.alarm,
        }
    except Exception as e:
        raise HTTPException(500, str(e))


def _template_steps(evidence):
    """LLM 없이 근거를 그대로 보여주는 대체 경로."""
    steps = []
    for i, e in enumerate(evidence[:5], 1):
        steps.append({
            "n": i,
            "title": e.get("title") or f"근거 {i}",
            "detail": (e.get("text") or "")[:220],
            "source": e.get("cite") or "",
            "kind": "manual" if e.get("kind") == "manual_text" else "code",
            "evidence_ids": [e.get("id")],
        })
    return steps


@app.post("/api/advice")
def advice(req: AdviceRequest):
    """
    조치 순서 생성.

    LLM 이 붙어 있으면 근거를 정비원이 따라갈 순서로 바꾼다. 이때
    각 단계는 반드시 검색 결과에 있는 근거 ID 를 인용해야 하며,
    없는 ID 를 인용한 단계는 코드에서 버린다(graph/advisor.py).

    LLM 이 없거나 실패하면 근거를 그대로 나열하는 템플릿으로 되돌아간다.
    빈 화면을 보여주는 것보다 낫고, 근거 자체는 검증된 것이기 때문이다.
    mock 필드로 어느 경로였는지 화면에 알린다.

    req.mock=True 를 주면 LLM 을 건너뛰고 템플릿만 쓴다.
    """
    try:
        diag = get_copilot(req.mode).answer(tag=req.tag, alarm=req.alarm, code=req.code)
        evidence = diag.get("evidence") or []
        decision = diag.get("decision")

        steps, mock, note, dropped = None, True, "", 0

        # 거절 판정에서는 조치를 만들지 않는다. 근거가 부족하다고
        # 판정한 뒤에 조치를 생성하면 판정이 의미가 없어진다.
        if decision == "advise" and evidence and not req.mock:
            try:
                from graph.advisor import generate
                res = generate(req.tag, req.alarm, evidence)
                steps = [dict(s, n=i, kind="llm")
                         for i, s in enumerate(res["steps"], 1)]
                mock, dropped = False, res.get("dropped", 0)
                note = res.get("summary", "")
                if dropped:
                    note += " (근거 검증에서 %d개 단계 제외)" % dropped
            except Exception as e:                          # noqa: BLE001
                note = "조치 생성 미사용 — %s" % str(e)[:120]

        if steps is None:
            steps = _template_steps(evidence)

        if not steps:
            steps = [{
                "n": 1,
                "title": "근거 부족 — 현장 확인 우선",
                "detail": "검색·CRAG 결과가 충분하지 않습니다. 태그 상태와 현장 계측값을 먼저 확인하십시오.",
                "source": "system",
                "kind": "system",
            }]
        summary = (
            f"판정: {decision} (충분성 {diag.get('grade', 0):.2f}). "
            f"{diag.get('grade_reason') or ''}"
        )
        if note:
            summary += " — " + note
        return {
            "mock": mock,
            "dropped_steps": dropped,
            "summary": summary,
            "steps": steps,
            "decision": decision,
            "grade": diag.get("grade"),
            "trace": diag.get("trace", []),
            "evidence": evidence[: config.FINAL_TOP_K],
        }
    except Exception as e:
        raise HTTPException(500, str(e))


@app.post("/api/interlock")
def interlock(req: InterlockRequest):
    try:
        ix = get_interlock()
        if req.as_input:
            res = ix.by_input(req.tag)
            if not res:
                return {"found": False, "tag": req.tag,
                        "message": f"{req.tag} 가 걸린 인터락이 없습니다."}
            hits = []
            for it, c in res["hits"]:
                hits.append({
                    "il_no": it["il_no"], "output_tag": it["output_tag"],
                    "kind": it["kind"], "action": it["action"],
                    "condition": _cond_dict(c),
                    "bypassable": it["bypassable"], "reset": it["reset"],
                    "dwg_no": it["dwg_no"],
                })
            return {"found": True, "as_input": True, "tag": req.tag,
                    "affected_outputs": res["affected_outputs"], "hits": hits}

        res = ix.by_output(req.tag, req.action)
        if not res:
            return {"found": False, "tag": req.tag,
                    "message": f"인터락 리스트에 {req.tag} 항목이 없습니다."}

        def pack(items):
            return [{
                "il_no": it["il_no"], "kind": it["kind"], "action": it["action"],
                "logic": it["logic"], "reset": it["reset"],
                "bypassable": it["bypassable"],
                "conditions": [_cond_dict(c) for c in it["conditions"]],
                "plc_block": it.get("plc_block"), "dwg_no": it.get("dwg_no"),
                "sheet": it.get("sheet"), "remark": it.get("remark"),
            } for it in items]

        return {
            "found": True, "as_input": False, "tag": req.tag,
            "action": res["action"], "output": res["output"],
            "blocking": pack(res["blocking"]),
            "enabling": pack(res["enabling"]),
            "other": pack(res["other"]),
        }
    except Exception as e:
        raise HTTPException(500, str(e))


def _cond_dict(c: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "parsed": c.get("parsed", False),
        "raw": c.get("raw", ""),
        "tags": c.get("tags", []),
        "kind": c.get("kind"),
        "op": c.get("op"),
        "setpoint": c.get("setpoint"),
        "unit": c.get("unit"),
        "level": c.get("level"),
        "state": c.get("state"),
        "state_label": c.get("state_label"),
        "multi": c.get("multi"),
        "delay": c.get("delay_sec") if c.get("delay_sec") is not None else c.get("delay"),
    }




@app.post("/api/report")
def report_4d(req: ReportRequest):
    """4D 트러블 리포트 PDF 다운로드."""
    try:
        from api.report_4d import build_4d_pdf
    except ImportError:
        from report_4d import build_4d_pdf  # type: ignore

    try:
        inst_data = None
        try:
            inst_data = instrument_detail(req.tag)
        except Exception:
            inst_data = {"instrument": {}, "history": [], "drawings": []}

        inst = inst_data.get("instrument") or {}
        history_raw = inst_data.get("history") or []

        # 검색 근거
        try:
            r = get_retriever(req.mode)
            query = (req.alarm + " " + req.code).strip()
            hits = r.retrieve(query, tag=req.tag) if query else []
        except Exception:
            hits = []

        manuals = []
        for rec, score, trace in hits[:6]:
            src = rec.get("source") or {}
            cite = "%s p.%s" % (src.get("file") or "", src.get("pdf_page") or "")
            manuals.append({
                "id": rec.get("id"),
                "name": rec.get("title") or "",
                "description": (rec.get("text") or "")[:120],
                "cite": cite,
                "cite_short": cite,
            })

        history = []
        for h in history_raw[:5]:
            history.append({
                "root_cause": h.get("root_cause") or "",
                "action": h.get("action_taken") or h.get("action") or "",
                "wo_no": h.get("wo_no") or "",
                "match": h.get("manual_match") or h.get("match") or "",
                "duration_min": h.get("duration_min"),
                "date": h.get("date") or "",
            })

        # 조치 순서 — 검색 상위 근거 제목을 단계로 사용
        steps = []
        for rec, score, trace in hits[:5]:
            steps.append(rec.get("title") or rec.get("id") or "점검 항목")

        import datetime as _dt
        now = _dt.datetime.now()
        payload = {
            "doc_no": "TR-%s-%s" % (now.strftime("%Y%m%d"), (req.tag or "TAG")[-4:]),
            "date_str": now.strftime("%Y-%m-%d %H:%M"),
            "tag": req.tag,
            "alarm": req.alarm or "",
            "code": req.code or "",
            "tech": req.tech or "-",
            "maker": inst.get("maker") or "",
            "model": inst.get("model") or "",
            "panel": inst.get("panel") or "-",
            "panel_grid": ((inst_data.get("panel_location") or {}).get("grid")
                           or ""),
            "panel_area": ((inst_data.get("panel_location") or {}).get("area")
                           or ""),
            "terminal": inst.get("terminal") or "-",
            "service": inst.get("service") or "",
            "status": "조회",
            "manual": manuals,
            "history": history,
            "advice_steps": steps,
            "confirmed_cause": req.confirmed_cause or "",
            "final_action": req.final_action or "",
            "parts": req.parts or "-",
            "duration_min": req.duration_min,
        }
        pdf = build_4d_pdf(payload)
        fname = "4D_Report_%s_%s.pdf" % (
            (req.tag or "TAG").replace("/", "-"),
            now.strftime("%Y%m%d_%H%M"),
        )
        return Response(
            content=pdf,
            media_type="application/pdf",
            headers={
                "Content-Disposition": "attachment; filename=%s" % fname,
            },
        )
    except Exception as e:
        raise HTTPException(500, "보고서 생성 실패: %s" % e)



@app.get("/api/interlock-source")
def interlock_source(tag: str = Query(...)):
    """인터락 리스트 엑셀 원본 블록 (사람이 대조할 수 있게)."""
    try:
        from ingest.interlock_real import (
            extract_source_block, detect_real_format, _interlock_files,
        )
        path = config.INTERLOCK_XLSX
        if not os.path.exists(path):
            raise HTTPException(404, "인터락 파일 없음: %s" % path)
        files = _interlock_files(path)
        if not files:
            raise HTTPException(
                404,
                "인터락 xlsx 없음 — data/interlock/ 에 .xlsx 를 넣으십시오 (%s)" % path,
            )
        # 폴더/파일 모두 extract 가 처리 (실물 양식)
        try:
            real = any(detect_real_format(f) for f in files)
        except Exception:
            real = True  # extract 쪽에서 재시도
        if real:
            block = extract_source_block(tag, path)
            if not block:
                raise HTTPException(404, f"{tag} 원본 블록 없음")
            return block
        # 데모 표 형식: 해당 OUTPUT TAG 행만
        from openpyxl import load_workbook
        path = files[0]
        ws = load_workbook(path, data_only=True).active
        rows = list(ws.iter_rows(values_only=True))
        hi = next(i for i, r in enumerate(rows)
                  if r and "IL NO" in [str(c).strip().upper() if c else "" for c in r])
        hdr = [str(c).strip() if c else "" for c in rows[hi]]
        out_rows = [{"row": hi + 1, "cells": hdr}]
        tag_u = tag.strip().upper()
        for i, r in enumerate(rows[hi + 1:], start=hi + 2):
            vals = [str(c).strip() if c is not None else "" for c in r]
            # OUTPUT TAG column
            try:
                ti = next(k for k, h in enumerate(hdr) if h.upper() == "OUTPUT TAG")
            except StopIteration:
                ti = 1
            if len(vals) > ti and vals[ti].upper() == tag_u:
                while vals and vals[-1] == "":
                    vals.pop()
                out_rows.append({"row": i, "cells": vals})
        if len(out_rows) <= 1:
            raise HTTPException(404, f"{tag} 행 없음")
        return {
            "tag": tag_u,
            "file": os.path.basename(path),
            "path": path,
            "header": tag_u,
            "row_start": out_rows[1]["row"] if len(out_rows) > 1 else hi + 1,
            "row_end": out_rows[-1]["row"],
            "rows": out_rows,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))




# 규칙 기반 의도 분석. 모듈 레벨에 둔다 — 모델이 없을 때의 유일한
# 경로이자 챗봇이 지어낸 태그를 대조하는 기준이므로, 다른 모듈에서
# 불러 시험할 수 있어야 한다. 함수 안에 숨겨 두면 점검이 안 된다.
# ── 태그 표기 정규화 ────────────────────────────────────────
_known_tags_cache = None
_known_tags_error = None


# ── 판넬 조회 ───────────────────────────────────────────────
def _need_panel() -> PanelIndex:
    px = get_panel()
    if px is None:
        raise HTTPException(503, "판넬 인덱스를 적재하지 못했습니다: %s"
                            % (_panel_error or "원인 미상"))
    return px


@app.get("/api/panels")
def list_panels():
    px = _need_panel()
    return {"panels": px.panels(), "unlocated": px.unlocated,
            "arrangement_pdf": os.path.basename(config.ARRANGEMENT_PDF),
            "arrangement_exists": os.path.exists(config.ARRANGEMENT_PDF)}


@app.get("/api/panel/{name}")
def panel_detail(name: str):
    """
    판넬 구성·위치 조회.

    판넬 단위 '상실' 은 제공하지 않는다. 이중화 구성에서는 랙 증설도
    스위칭으로 대응하므로 판넬 전체가 죽는 상황이 성립하지 않는다.
    상실 영향은 /api/card 를 쓴다.
    """
    px = _need_panel()
    d = px.by_panel(name)
    if not d:
        raise HTTPException(404, "계기 리스트에 %s 판넬이 없습니다." % name)
    return d


@app.get("/api/cards")
def list_cards(panel: Optional[str] = None):
    """
    IO 카드 목록.

    이중화(S7-400H/410H) 구성에서는 CPU·전원·통신이 이중화되고 카드만
    단일이다. 그래서 실제 단일 고장 단위는 판넬이 아니라 카드다.
    """
    px = _need_panel()
    cards = px.cards()
    if panel:
        cards = [c for c in cards if c["panel"] == panel]
    return {"cards": cards, "count": len(cards)}


@app.get("/api/card")
def card_detail(id: str = Query(..., description="예: CUB-B/R0/S8"),
                impact: int = Query(0, ge=0, le=1)):
    # 카드 ID 에 '/' 가 들어가므로 경로가 아니라 질의 인자로 받는다.
    px = _need_panel()
    d = px.impact(id) if impact else px.by_card(id)
    if not d:
        raise HTTPException(404, "계기 리스트에 %s 카드가 없습니다." % id)
    return d


@app.get("/api/tag-consistency")
def tag_consistency():
    """
    IO List · 계기 리스트 · 인터락 리스트의 태그 교차 대조.

    어느 쪽이 맞는지는 판정하지 않는다 — 두 문서가 다르다는 사실까지만
    말한다. 실물 문서로 갈아 끼울 때 제일 먼저 걸리는 지점이다.
    """
    from ingest.tag_registry import cross_check
    return cross_check()


@app.get("/api/common-cause")
def common_cause():
    """한 인터락의 조건 태그가 같은 카드에 몰려 있는지 — 설계 검토 항목."""
    px = _need_panel()
    return px.common_cause()


@app.get("/api/panel-of/{tag}")
def panel_of_tag(tag: str):
    px = _need_panel()
    d = px.by_tag(tag)
    if not d:
        raise HTTPException(404, "계기 리스트에 %s 가 없습니다." % tag)
    return d


@app.get("/api/cards")
def list_cards(panel: Optional[str] = None):
    """
    IO 카드 목록.

    이중화(S7-400H/410H) 구성에서는 CPU·전원·통신이 이중화되고 카드만
    단일이다. 그래서 실제 단일 고장 단위는 판넬이 아니라 카드다.
    """
    px = _need_panel()
    cards = px.cards()
    if panel:
        cards = [c for c in cards if c["panel"] == panel]
    return {"cards": cards, "count": len(cards)}


@app.get("/api/card")
def card_detail(id: str = Query(..., description="예: CUB-B/R0/S8"),
                impact: int = Query(0, ge=0, le=1)):
    # 카드 ID 에 '/' 가 들어가므로 경로가 아니라 질의 인자로 받는다.
    px = _need_panel()
    d = px.impact(id) if impact else px.by_card(id)
    if not d:
        raise HTTPException(404, "계기 리스트에 %s 카드가 없습니다." % id)
    return d


@app.get("/api/tag-consistency")
def tag_consistency():
    """
    IO List · 계기 리스트 · 인터락 리스트의 태그 교차 대조.

    어느 쪽이 맞는지는 판정하지 않는다 — 두 문서가 다르다는 사실까지만
    말한다. 실물 문서로 갈아 끼울 때 제일 먼저 걸리는 지점이다.
    """
    from ingest.tag_registry import cross_check
    return cross_check()


@app.get("/api/common-cause")
def common_cause():
    """한 인터락의 조건 태그가 같은 카드에 몰려 있는지 — 설계 검토 항목."""
    px = _need_panel()
    return px.common_cause()


@app.get("/api/panel-of/{tag}")
def panel_of_tag(tag: str):
    px = _need_panel()
    d = px.by_tag(tag)
    if not d:
        raise HTTPException(404, "계기 리스트에 %s 가 없습니다." % tag)
    return d


def known_tags():
    """계기 리스트 + 인터락 출력 태그의 합집합. 대문자 기준."""
    global _known_tags_cache
    if _known_tags_cache is None:
        # 여기서 실패를 삼키면 태그 목록이 비고, 그러면 챗봇이 멀쩡한
        # 태그에도 "리스트에 없습니다" 라고 답한다. 원인을 찾기 어려운
        # 종류의 고장이므로 사유를 남긴다.
        global _known_tags_error
        ts, errs = set(), []
        for name, fn in (("출력 태그", load_output_tags),
                         ("계기 리스트", load_instruments)):
            try:
                ts |= {str(t).upper() for t in fn().keys()}
            except Exception as e:                          # noqa: BLE001
                errs.append("%s: %s" % (name, e))
        _known_tags_error = " / ".join(errs) if errs else None
        _known_tags_cache = {t for t in ts if t}
    return _known_tags_cache


def normalize_tag(raw):
    """
    사람이 쓰는 표기를 실제 태그로 되돌린다.

    현장에서는 하이픈을 빼거나 띄어 쓴다. "LCV 01", "lcv01", "LCV_01"
    은 모두 LCV-01 을 말한다. 이전 정규식은 하이픈 형태만 받았고,
    그래서 "LCV 01 인터락 보여줘" 가 태그 없이 넘어가 화면에서 아무
    일도 일어나지 않았다. 실패가 조용해서 원인을 알기 어려웠다.

    실재하는 태그로만 되돌린다. 목록에 없으면 None 을 준다 —
    없는 태그를 만들어 조회하지 않기 위해서다.
    """
    if not raw:
        return None
    key = re.sub(r"[^a-z0-9]", "", str(raw).lower())
    if not key:
        return None
    for t in known_tags():
        if re.sub(r"[^a-z0-9]", "", t.lower()) == key:
            return t
    return None


_panel_names_cache = None


def known_panels():
    """판넬명 집합. 실패하면 빈 집합이 아니라 사유를 남긴다."""
    global _panel_names_cache
    if _panel_names_cache is None:
        px = get_panel()
        _panel_names_cache = ({p["panel"].upper() for p in px.panels()}
                              if px is not None else set())
    return _panel_names_cache


def find_panel(msg):
    """
    문장에서 판넬명을 찾는다.

    판넬명(CUB-B, RIO-01)은 태그와 표기가 겹친다. find_tag 가 먼저 돌면
    'RIO-01' 을 없는 태그로 보고 "리스트에 없습니다" 를 반환한다.
    그래서 판넬 판정을 태그 판정보다 앞에 둔다.
    """
    names = known_panels()
    if not names:
        return None
    up = msg.upper()
    hits = [n for n in names if re.search(r"(?<![A-Z0-9])%s(?![A-Z0-9])"
                                          % re.escape(n), up)]
    if hits:
        return max(hits, key=len)
    # 하이픈 없이 쓴 경우 (CUB B / CUBB)
    for n in names:
        loose = re.escape(n).replace(r"\-", r"[\s_-]*")
        if re.search(r"(?<![A-Z0-9])%s(?![A-Z0-9])" % loose, up):
            return n
    return None


_tag_prefix_cache = None


def known_tag_prefixes():
    """
    실재하는 태그에서 접두어를 뽑아 둔다 (AIT, PIT, XV, LCV, P …).

    하드코딩하지 않고 리스트에서 읽는다. 실물 계기 리스트로 바꾸면
    접두어도 함께 바뀌어야 하기 때문이다.
    """
    global _tag_prefix_cache
    if _tag_prefix_cache is None:
        pres = set()
        for t in known_tags():
            m = re.match(r"^([A-Za-z]+)", str(t))
            if m:
                pres.add(m.group(1).upper())
        _tag_prefix_cache = pres
    return _tag_prefix_cache


def find_tag(msg, cur_tag=None):
    """
    문장에서 태그를 찾는다. 하이픈이 없어도 찾고, 실재 여부를 확인한다.

    반환: (태그 또는 None, 후보였지만 목록에 없던 문자열 또는 None)
    두 번째 값이 있으면 "그런 태그는 없습니다" 라고 말할 수 있다.
    """
    # 경계를 \b 로 잡으면 한글 조사가 붙은 표기를 통째로 놓친다.
    #
    #   "ait-1001은 어느 위치에 있어?"
    #        └ \b 는 '1' 과 '은' 사이를 경계로 보지 않는다 (둘 다 word 문자)
    #
    # 그러면 find_tag 가 실패하고 cur_tag 로 조용히 대체되어, 사용자가
    # 물어본 것과 **다른 태그**의 답이 나간다. 화면에는 그럴듯한 답이
    # 떠서 틀렸다는 사실조차 드러나지 않는다.
    #
    # 영숫자만 경계로 보게 바꾼다 — 한글·공백·문장부호는 모두 구분자다.
    B0, B1 = r"(?<![A-Za-z0-9])", r"(?![A-Za-z0-9])"
    # 밑줄로 여러 마디를 잇는 실물 표기를 **가장 먼저** 본다.
    #   DWP_AMP_6018A · WATER_TANK_LIT_P3601B · FAB_CPU_XA_0001
    # 이것이 없으면 아래 'LCV 01' 규칙이 'DWP_AMP' 까지만 끊어 먹고,
    # 해석에 실패한 뒤 현재 화면 태그로 대체되어 **다른 설비의 답**이
    # 나간다. 화면에는 그럴듯한 문장이 떠서 틀린 줄도 모른다.
    pats = [B0 + r"([A-Za-z][A-Za-z0-9]*(?:_[A-Za-z0-9]+)+)" + B1,  # DWP_AMP_6018A
            B0 + r"([A-Za-z]{1,8}-[A-Za-z0-9]{1,8})" + B1,      # LCV-01
            B0 + r"([A-Za-z]{2,8})[\s_]+([A-Za-z0-9]{1,8})" + B1,  # LCV 01
            B0 + r"([A-Za-z]{2,8})(\d{1,5}[A-Za-z]?)" + B1]       # LCV01
    # 못 푼 후보를 두 등급으로 나눈다. 등급을 안 나누면 양쪽이 번갈아
    # 뚫린다.
    #
    #   강함 : 접두어가 실재 태그 접두어 (AIT-9999, XV-1234)
    #          → 태그를 말한 게 확실하니 거절하고 QA 로도 넘기지 않는다
    #   약함 : 숫자는 있지만 접두어가 낯설다 (PCS 7, TB2, AI 16xI, bge-m3)
    #          → 태그가 아니라 기술 용어일 수 있으니 질문은 막지 않는다
    #            다만 **현재 화면 태그로 대체하지도 않는다**
    #
    # 약한 후보까지 거절하면 "AI 16xI 모듈 상태 표시" 같은 매뉴얼 질문이
    # 통째로 막히고(실제로 매뉴얼 6.1 절 제목이다), 반대로 약한 후보에
    # 대체를 허용하면 "zzt-9999는 어디야" 에 엉뚱한 설비 답이 나간다.
    miss = None
    weak = False
    for p in pats:
        for m in re.finditer(p, msg):
            cand = "-".join(g for g in m.groups() if g)
            t = normalize_tag(cand)
            if t:
                return t, None
            if not re.search(r"\d", cand):
                continue
            head = re.match(r"^([A-Za-z]+)", cand)
            if head and head.group(1).upper() in known_tag_prefixes():
                if miss is None:
                    miss = cand.upper()
            else:
                weak = True
    if miss:
        return (cur_tag or None), miss
    if weak:
        return None, None          # 대체하지 않는다
    return (cur_tag or None), None


def panel_intent(msg, low, cur_tag=None):
    """
    판넬 관련 의도. 해당 없으면 None 을 돌려 기존 규칙으로 넘긴다.

    답변에 실제 수치를 담는다. "판넬 조회 탭으로 이동합니다" 만 하면
    사용자가 챗봇에 물어본 이유가 사라진다.
    """
    px = get_panel()
    if px is None:
        return None

    panel = find_panel(msg)
    wants_impact = re.search(r"내리|끄|차단|정전|단전|셧다운|shutdown|"
                             r"점검|영향|잃|죽|무슨\s*일", low)
    wants_card = re.search(r"카드|모듈|슬롯|채널|card|module", low)
    wants_cc = re.search(r"공통\s*원인|같은\s*카드|한\s*카드|공통\s*고장|"
                         r"분산\s*배치|common\s*cause", low)

    # ── 공통원인 점검 ─────────────────────────────────────
    if wants_cc:
        cc = px.common_cause()
        if not cc["loaded"]:
            return {"type": "chat", "reply": "인터락 리스트를 읽지 못했습니다."}
        if cc["findings"]:
            head = ", ".join("%s(%s)" % (f["il_no"], f["severity"])
                             for f in cc["findings"][:5])
            body = "인터락 %d건 중 %d건이 지적됩니다 — %s" % (
                cc["checked"], len(cc["findings"]), head)
        else:
            body = ("인터락 %d건을 점검했고 지적은 없습니다. 다만 이는 "
                    "'설계가 안전하다'가 아니라 이 리스트 범위에서 걸린 것이 "
                    "없다는 뜻입니다." % cc["checked"])
        return {"type": "panel", "tab": "panel", "reply": body}

    # ── 카드 상실 영향 ────────────────────────────────────
    # 이중화 구성에서는 카드가 유일한 단일 고장 지점이다.
    if wants_card:
        tag, miss = find_tag(msg, cur_tag)
        if miss:
            return {"type": "panel", "tab": "panel",
                    "reply": "%s 는 계기 리스트에 없습니다." % miss}
        if tag:
            cid = px.card_of(tag)
            d = px.impact(cid) if cid else None
            if d:
                if not d["interlock_loaded"]:
                    body = ("계기 %d점을 잃습니다. 인터락 리스트를 읽지 못해 "
                            "의존 관계는 계산하지 못했습니다." % d["points"])
                else:
                    full = [r["il_no"] for r in d["dependencies"]
                            if r["remaining_protection"].startswith("없음")]
                    body = ("같은 카드에 %d점(%s)이 물려 있습니다. "
                            "의존 인터락 %d건 중 안전 인터락 %d건, 남는 보호가 "
                            "없는 항목은 %s, 영향 출력은 %s 입니다."
                            % (d["points"], ", ".join(d["lost_tags"]),
                               len(d["dependencies"]), d["safety_count"],
                               ", ".join(full) or "없습니다",
                               ", ".join(o["tag"]
                                         for o in d["affected_outputs"])
                               or "없습니다"))
                return {"type": "panel", "tab": "panel", "panel": d["panel"],
                        "card": cid, "tag": tag,
                        "reply": "%s 는 %s 카드입니다. %s\n\n※ %s"
                                 % (tag, cid, body, d["caveat"])}
    # "판넬 명 알려줘", "태그가 있는 판넬", "어느 판넬" 등을 모두 위치 조회로 본다.
    # 이전 패턴은 '어디/위치/어느 판넬' 만 받아서 "판넬 명을 알려줘" 를
    # 놓쳤고, 그 결과 QA/LLM 경로로 넘어가 판넬 이름을 지어냈다.
    wants_where = re.search(
        r"어디|위치|"
        r"어느\s*판\s*넬|무슨\s*판\s*넬|어떤\s*판\s*넬|"
        r"판\s*넬\s*(명|이름|어디|위치|번호|은|는|이|가|을|를)|"
        r"있는\s*판\s*넬|속한\s*판\s*넬|물린\s*판\s*넬|"
        r"판\s*넬.*알려|알려.*판\s*넬|"
        r"which\s*panel|what\s*panel|panel\s*name",
        low)

    # ── 판넬 전체 상실 질문은 성립하지 않는다 ──────────────
    #
    # 이중화(S7-400H/410H) 구성에서는 랙 증설도 스위칭으로 대응하므로
    # 판넬 전체가 한 번에 죽는 상황이 없다. 숫자를 만들어 답하면
    # 없는 시나리오에 근거를 준다. 카드 단위로 되돌린다.
    if panel and wants_impact and not wants_card:
        d = px.by_panel(panel)
        cards = [c for c in px.cards() if c["panel"] == panel]
        return {"type": "panel", "tab": "panel", "panel": panel,
                "reply": ("이중화 구성이라 %s 판넬 전체가 한 번에 죽는 "
                          "상황은 다루지 않습니다. 단일 고장 단위는 IO "
                          "카드이고, %s 에는 카드 %d장에 계기 %d점이 "
                          "물려 있습니다. 카드를 고르면 그 카드 상실 "
                          "영향을 보여드립니다."
                          % (panel, panel, len(cards), d["points"]))}

    # ── 태그가 어느 판넬인가 ──────────────────────────────
    if wants_where and not panel:
        tag, miss = find_tag(msg, cur_tag)
        # miss 가 있다는 것은 문장 안의 태그를 하나도 풀지 못했다는 뜻이다.
        # "어디에 있어?" 는 이름 자체가 질문의 주어라서, 현재 선택된
        # 태그로 대신 답하면 다른 설비의 위치를 알려주게 된다.
        if miss:
            return {"type": "panel", "tab": "panel",
                    "reply": "%s 는 계기 리스트와 인터락 리스트에 없습니다. "
                             "태그를 확인해 주세요." % miss}
        if not tag:
            return {"type": "panel", "tab": "panel",
                    "reply": "말씀하신 태그를 계기 리스트에서 찾지 "
                             "못했습니다. 태그를 확인해 주세요."}
        if tag:
            d = px.by_tag(tag)
            if d:
                loc = d["location"]
                where = ("%s / 그리드 %s / %s"
                         % (loc["area"], loc["grid"],
                            "실내" if loc["indoor"] else "옥외")) if loc \
                    else "배치 정보 없음"
                return {"type": "panel", "tab": "panel", "panel": d["panel"],
                        "tag": tag,
                        "card": d.get("card"),
                        "reply": "%s 는 %s 판넬입니다. 위치는 %s, 단자대 %s, "
                                 "%s Rack %s Slot %s Ch %s (카드 %s) 입니다."
                                 % (tag, d["panel"], where, d["terminal"] or "-",
                                    d["plc"], d["rack"], d["slot"], d["ch"],
                                    d.get("card"))}
            # 태그는 알지만 PANEL 배선이 없다.
            # 출력 태그(LCV/XV/펌프 등)는 인터락 리스트에 있어도
            # 계기 리스트 PANEL 열이 없으면 판넬 위치를 말할 수 없다.
            # 없는 판넬 이름을 지어내지 않는다.
            outs = {}
            try:
                outs = load_output_tags()
            except Exception:                               # noqa: BLE001
                outs = {}
            if tag in outs or tag.upper() in {t.upper() for t in outs}:
                rec = outs.get(tag) or outs.get(tag.upper()) or {}
                svc = rec.get("service") or ""
                return {"type": "panel", "tab": "panel", "tag": tag,
                        "reply": (
                            "%s 는 인터락/출력 태그입니다%s. "
                            "판넬 위치는 계기 리스트의 PANEL 열(IO 배선)에서 "
                            "읽는데, 이 태그에는 그 정보가 없습니다. "
                            "판넬 이름을 추정하지 않습니다. "
                            "동작 조건이 필요하면 인터락 조회를 해 주세요."
                            % (tag, (" (%s)" % svc) if svc else "")
                        )}
            return {"type": "chat",
                    "reply": "%s 는 계기 리스트에 PANEL 정보가 없어 "
                             "판넬 위치를 말할 수 없습니다." % tag}

    # ── 판넬명만 나온 경우 ────────────────────────────────
    if panel and re.search(r"판\s*넬|panel|무엇|뭐|물려|계기", low):
        d = px.by_panel(panel)
        loc = d["location"] or {}
        return {"type": "panel", "tab": "panel", "panel": panel,
                "reply": "%s — %s 그리드 %s, 계기 %d점 (%s)"
                         % (panel, loc.get("area", "위치 정보 없음"),
                            loc.get("grid", "-"), d["points"],
                            ", ".join("%s %d" % (k, len(v))
                                      for k, v in d["by_tb"].items()))}
    return None


SMALLTALK = [
    # (정규식, 답변)
    (r"^\s*(안녕|하이|헬로|반가|ㅎㅇ|hello|hi|hey)",
     "안녕하세요. 계기 알람·인터락·판넬 조회를 도와드립니다. "
     "찾으시는 태그나 증상을 알려주세요."),
    (r"(고마워|고맙|감사|thanks|thank you)",
     "도움이 되었다면 다행입니다. 더 찾으실 것이 있으면 말씀해 주세요."),
    (r"(잘가|바이|수고|들어가|bye|끝)",
     "네, 필요하실 때 다시 불러 주세요."),
    (r"(미안|죄송|sorry)",
     "괜찮습니다. 다시 말씀해 주세요."),
    (r"(누구|정체|넌\s*뭐|너는\s*뭐|what are you)",
     "이 화면의 조회를 돕는 도우미입니다. 계기 리스트·인터락 리스트·"
     "장비 매뉴얼에 있는 내용만 답하고, 없는 것은 없다고 말합니다."),
    (r"(뭐\s*할\s*수|뭘\s*할\s*수|무엇을\s*할\s*수|기능|어떤\s*걸\s*도와|"
     r"뭐\s*해줄)",
     "알람 조회(태그+증상), 인터락 조회(출력 태그), 판넬·카드 조회, "
     "도면 보기를 할 수 있습니다. 자세한 사용법은 '사용법' 이라고 "
     "말씀해 주세요."),
]


def smalltalk_intent(msg, low):
    """
    인사·감사·잡담에 예시 목록을 던지지 않게 한다.

    "안녕하세요" 에 "예) AIT-4002 low acid 알람 조회해줘 …" 를 돌려주면
    사람은 무시당했다고 느낀다. 실제로 그렇게 나갔다. 짧은 인사말은
    도메인 질의가 아니므로 매뉴얼 검색으로 넘겨서도 안 된다 — 근거가
    없으니 "매뉴얼에서 근거를 찾지 못했습니다" 가 나온다.

    길이를 제한하는 이유는 "안녕히 계세요, 그런데 AIT-1001 은…" 같은
    문장을 인사로 삼키지 않기 위해서다.
    """
    if len(msg.strip()) > 24:
        return None
    for pat, reply in SMALLTALK:
        if re.search(pat, low):
            return {"type": "chat", "reply": reply}
    return None


def rule_intent(msg: str, cur_tag: str = None):
    """
    규칙 기반 의도 분석.

    모델이 없을 때의 유일한 경로이고, 챗봇이 지어낸 태그를 대조하는
    기준이기도 하다. 이전에는 챗봇 엔드포인트 안에 중첩되어 있어
    바깥의 req 를 직접 참조했고, 그래서 밖에서 불러 시험할 수 없었다.
    점검할 수 없는 코드는 조용히 깨진다.
    """
    low = msg.lower()

    # 인사·감사는 조회 의도가 아니다. 예시 목록도 매뉴얼 검색도 아니다.
    st = smalltalk_intent(msg, low)
    if st:
        return st

    # 판넬 의도를 태그 판정보다 먼저 본다 (표기가 겹치므로)
    pintent = panel_intent(msg, low, cur_tag)
    if pintent:
        return pintent

    tag, miss = find_tag(msg, cur_tag)
    # 문장에 적힌 태그가 목록에 없으면, 현재 화면 태그로 대체하지 않는다.
    # miss 가 있는데 tag=cur_tag 로 넘어가면 다른 설비 답변이 나간다.
    if miss:
        return {"type": "chat",
                "reply": "%s 는 계기 리스트와 인터락 리스트에 없습니다. "
                         "태그를 확인해 주세요." % miss}
    if re.match(r"^(도움|help|사용법|가이드)", low) or msg.strip() == "?":
        return {
            "type": "help",
            "reply": (
                "사용 가이드\n"
                "1) 알람 조회: 태그 선택 → 증상 입력 → 알람 조회\n"
                "2) 원문/도면: 결과에서 원문 보기·도면 보기\n"
                "3) 인터락: 인터락 조회 탭에서 출력 태그\n"
                "4) 자연어 예: AIT-4002 low acid 알람 조회해줘"
            ),
        }
    # 화면에 이미 결과가 떠 있는 상태에서의 후속 질문은 명령이 아니다.
    #
    # "조회된 내용을 보고 조치방법을 알려줘" 는 '조회' 라는 글자 때문에
    # 다시 알람 조회 명령으로 걸렸고, 같은 조회를 반복하며 대화가
    # 제자리를 돌았다. 이런 문장은 명령이 아니라 방금 결과에 대한
    # 질문이므로 규칙에서 빼고 질의응답으로 넘긴다.
    if re.search(r"(조회된|검색된|나온|방금|위의|이|그|저)\s*(내용|결과|것|거)"
                 r"|결과를?\s*(보고|바탕|기반)|앞서|아까", low):
        return {"type": "followup", "tag": tag, "question": msg}

    if re.search(r"도면|p\s*&\s*i\s*d|pid|p&id", low):
        if not tag:
            return {"type": "chat", "reply": "태그를 알려주세요. 예: AIT-1001 P&ID 도면 보여줘"}
        return {"type": "drawing", "tag": tag, "tab": "alarm",
                "reply": "%s 도면을 엽니다." % tag}
    if re.search(r"인터락.*원본|원본.*인터락", low):
        if not tag:
            return {"type": "chat", "reply": "예: LCV-01 인터락 원본 보여줘"}
        act = "CLOSE" if re.search(r"close|닫", low) else (
            "START" if re.search(r"start|기동", low) else (
            "STOP" if re.search(r"stop|정지", low) else "OPEN"))
        return {"type": "interlock_source", "tag": tag, "tab": "interlock",
                "action": act, "openSource": True,
                "reply": "%s 인터락 원본을 엽니다." % tag}
    if re.search(r"인터락|interlock", low):
        if not tag:
            return {"type": "chat", "reply": "예: XV-4101 인터락 조회해줘"}
        act = "CLOSE" if re.search(r"close|닫", low) else (
            "START" if re.search(r"start|기동", low) else (
            "STOP" if re.search(r"stop|정지", low) else "OPEN"))
        return {"type": "interlock", "tag": tag, "tab": "interlock",
                "action": act,
                "reply": "%s %s 인터락을 조회합니다." % (tag, act)}
    if re.search(r"조치|어떻게\s*(해|하나|하면)|뭘\s*해|해결", low):
        # 조치 순서는 조회 결과가 있어야 만들 수 있다. 화면 상태를
        # 아는 챗봇 계층에서 처리하도록 followup 으로 넘긴다.
        return {"type": "followup", "tag": tag, "question": msg,
                "want": "advice"}

    if re.search(r"알람|조회|검색|고장", low) or (tag and re.search(r"해줘|보여", low)):
        alarm = re.sub(r"\b([A-Za-z]{1,8}-[A-Za-z0-9]{1,8})\b", " ", msg)
        alarm = re.sub(r"알람|조회|해줘|해주세요|검색|좀|관련", " ", alarm, flags=re.I)
        alarm = re.sub(r"\s+", " ", alarm).strip() or "alarm"
        if not tag:
            return {"type": "chat", "reply": "예: AIT-4002 acid residual low 알람 조회해줘"}
        return {"type": "diagnose", "tag": tag, "tab": "alarm", "alarm": alarm,
                "reply": "%s 알람 조회를 실행합니다." % tag}
    # 아무 규칙에도 걸리지 않은 포괄 응답. LLM 이 더 나은 답을 낼 수
    # 있으므로 generic 표식을 달아, 챗봇에서 덮어쓰지 않게 한다.
    return {
        "type": "chat",
        "generic": True,
        "reply": ("무엇을 도와드릴까요? 태그와 증상을 함께 말씀하시면 알람 "
                  "조회를 실행합니다.\n"
                  "예) AIT-4002 low acid 알람 조회해줘 · XV-4101 인터락 조회 "
                  "· AIT-1001 은 어느 판넬이야 · 사용법"),
    }



class _FixedEvidence:
    """
    이미 화면에 있는 근거를 그대로 쓰는 어댑터.

    후속 질문에 대해 검색을 다시 돌리면 사용자가 보고 있는 것과 다른
    근거로 답할 수 있다. "조회된 내용을 보고" 라는 요청에는 조회된
    그 내용으로 답해야 한다.
    """

    def __init__(self, evidence, ctx):
        self._ev = evidence
        self._ctx = ctx

    def answer(self, **kw):
        return {"decision": "advise",
                "grade": (self._ctx.grade if self._ctx else None) or 0.7,
                "evidence": self._ev}


@app.post("/api/chat")
def chat_help(req: ChatRequest):
    """도우미 챗봇. LLM API 가 있으면 의도 분석, 없으면 규칙 기반."""
    import re
    import json as _json
    import urllib.request

    text = (req.message or "").strip()
    if not text:
        raise HTTPException(400, "message 필요")

    # 대화 모델 호출은 조치 생성과 같은 게이트웨이를 쓴다.
    #
    # 이전에는 여기서 provider 를 따로 해석했고, 목록이
    # ("openai","azure","llm") 이라 COPILOT_PROVIDER=ollama 를 넣으면
    # 조치 생성은 켜지는데 챗봇만 조용히 규칙 엔진으로 떨어졌다.
    # 같은 환경변수가 두 곳에서 다른 뜻을 갖고 있었고, 그 사실이
    # 화면 어디에도 드러나지 않았다. 게이트웨이를 하나로 합친다.
    from graph.advisor import _CHAT, _parse

    ACTIONABLE = {"diagnose", "drawing", "interlock", "interlock_source",
                  "advice", "help", "navigate", "panel"}
    NEEDS_TAG = {"diagnose", "drawing", "interlock", "interlock_source",
                 "advice"}

    def finalize(data, source):
        """태그를 실재하는 것으로 되돌리고, 없으면 되묻는다."""
        data.setdefault("type", "chat")
        data.setdefault("reply", "요청을 처리합니다.")
        t = normalize_tag(data.get("tag"))
        if t:
            data["tag"] = t
        else:
            data.pop("tag", None)
        if data["type"] in NEEDS_TAG and not data.get("tag"):
            data["type"] = "chat"
            data["reply"] = ("어느 태그인지 알려주세요. "
                             "예: LCV-01 인터락 조회 / AIT-4002 알람 조회")
        data["engine"] = source
        return data

    def looks_like_question(t):
        """UI 명령이 아니라 도메인 질문인가."""
        if len(t) < 6:
            return False
        # 물음표 하나로 도메인 질문이라고 보면 안 된다. "인사 안 해주고
        # 예시를 들어주네?" 같은 말이 매뉴얼 검색으로 넘어가 "근거를 찾지
        # 못했습니다" 가 나온다. 의문사·요청어가 실제로 있어야 한다.
        if re.search(r"(어떻게|왜|무엇|뭐가|뭔|어디|언제|얼마|방법|절차|"
                     r"주기|원인|의미|뜻|차이|기준|규격|사양|알려|설명|"
                     r"인가요|하나요|되나요|일까|인지|점검|조치|확인)", t):
            return True
        return False

    # 규칙 엔진을 먼저 돌린다.
    #
    # 이전에는 LLM 이 우선이었고, 규칙은 LLM 이 없을 때의 대체 경로였다.
    # 그런데 7B 급 모델은 이 작업에서 규칙보다 못하다. "LCV-01 인터락
    # 조회해줘" 에 type=chat 과 "인터락 정보를 조회하겠습니다" 를 돌려주면,
    # 화면은 실행하지 않고 사용자는 왜 안 되는지 알 수 없다.
    #
    # 규칙 엔진은 이 앱이 지원하는 명령을 정확히 덮고 결정적이다.
    # 규칙이 실행 가능한 의도를 뽑아내면 그것을 쓰고, 규칙이 못 알아들은
    # 표현에 한해 LLM 에 물어본다. 이 순서가 시연에서도 안전하다 —
    # 같은 문장에 같은 동작이 나온다.
    rule = rule_intent(text, req.tag) or {}
    if rule.get("type") in ACTIONABLE:
        return finalize(rule, "rule")

    # 규칙이 "목록에 없는 태그" 를 이미 확정했으면 QA/LLM 으로 넘기지 않는다.
    # 넘기면 매뉴얼 검색·대화 모델이 'Main Feed Control Panel' 같은
    # 그럴듯한 판넬 이름을 지어낸다. 없는 것은 없다고 말하는 것이 답이다.
    if (rule.get("type") == "chat" and rule.get("reply")
            and re.search(r"없습니다|없어|목록에\s*없|리스트에\s*없",
                          rule.get("reply", ""))):
        return finalize(rule, "rule")

    # 후속 질문 — 화면에 떠 있는 결과를 근거로 답한다.
    #
    # 이 경로가 없으면 챗봇은 매 메시지를 독립 명령으로 보고 같은
    # 조회를 반복한다. 사용자는 결과를 보며 묻는데 챗봇만 그것을
    # 모르는 상태였다.
    if rule.get("type") == "followup":
        ctx = req.context
        ev = (ctx.evidence if ctx else None) or []
        if not ev:
            return finalize({"type": "chat",
                             "reply": "아직 조회 결과가 없습니다. 먼저 "
                                      "태그와 증상으로 알람을 조회해 주세요. "
                                      "예: AIT-4002 acid residual low 알람 조회해줘"},
                            "rule")
        # "조회된 내용을 보고 조치방법을 알려줘" 처럼 두 규칙에 걸친
        # 문장이 있으므로 질문 본문에서 다시 판별한다.
        want = rule.get("want")
        if not want and re.search(r"조치|어떻게\s*(해|하나|하면)|뭘\s*해|해결",
                                  rule.get("question", "")):
            want = "advice"
        try:
            if want == "advice":
                from graph.advisor import generate
                res = generate(ctx.tag, ctx.alarm or rule["question"], ev)
                lines = [res.get("summary") or "확인 순서입니다."]
                for i, st in enumerate(res["steps"], 1):
                    lines.append("%d. %s — %s" % (i, st["title"], st["detail"]))
                return {"type": "chat", "engine": "advice",
                        "reply": "\n".join(lines),
                        "grounded": True,
                        "citations": [{"id": e["id"],
                                       "title": e.get("title", ""),
                                       "cite": e.get("cite", "")}
                                      for e in ev[:3]]}
            from graph.qa import answer as qa_answer
            res = qa_answer(rule["question"], tag=ctx.tag,
                            copilot=_FixedEvidence(ev, ctx),
                            instruments=load_instruments())
            return {"type": "chat", "engine": "followup",
                    "reply": res["reply"], "grounded": res["ok"],
                    "citations": [{"id": e["id"], "title": e.get("title", ""),
                                   "cite": e.get("cite", "")}
                                  for e in (res.get("evidence") or [])[:3]]}
        except Exception as e:                              # noqa: BLE001
            return finalize({"type": "chat",
                             "reply": "결과를 해석하지 못했습니다: %s"
                                      % str(e)[:110]}, "rule")

    # 명령이 아니고 도메인 질문으로 보이면 매뉴얼 근거로 답한다.
    #
    # 챗봇이 검색·판정·근거 검증을 쓰지 않고 명령 해석만 하고 있었다.
    # 같은 부품을 대화 경로에도 연결한다. 근거가 부족하면 여기서도
    # 지어내지 않고 모른다고 답한다.
    if config.CHAT_QA and looks_like_question(text):
        try:
            from graph.qa import answer as qa_answer
            res = qa_answer(text, tag=normalize_tag(req.tag),
                            mode=DEFAULT_MODE, copilot=get_copilot(DEFAULT_MODE),
                            instruments=load_instruments())
            return {"type": "chat", "engine": "qa",
                    "reply": res["reply"],
                    "grounded": res["ok"],
                    "grade": res.get("grade"),
                    "citations": [{"id": e["id"], "title": e.get("title", ""),
                                   "cite": e.get("cite", "")}
                                  for e in (res.get("evidence") or [])[:3]]}
        except Exception as e:                              # noqa: BLE001
            pass    # 실패하면 아래 일반 대화 경로로 내려간다

    provider = config.LLM_PROVIDER
    if not (req.use_llm and provider in _CHAT):
        out = finalize(rule, "rule")
        out["llm_reason"] = ("COPILOT_PROVIDER=%s" % provider
                             if req.use_llm else "요청에서 비활성")
        return out

    system = (
        "당신은 Plant Maintenance Copilot 의 도우미입니다. "
        "플랜트 정비원이 쓰는 도구이며, 다음을 할 수 있습니다.\n"
        "- 알람 조회: 설비 태그와 증상을 주면 벤더 매뉴얼과 에러코드표에서 "
        "원인·조치를 찾아 출처(문서·페이지)와 함께 보여줍니다. 증상은 "
        "한국어로 써도 되고 매뉴얼이 영문이어도 찾습니다.\n"
        "- 근거가 부족하면 답을 지어내지 않고 근거 부재를 알립니다.\n"
        "- 보수 이력 대조: 같은 설비에서 과거에 있었던 조치를 함께 보여줍니다.\n"
        "- 인터락 조회: 밸브·펌프가 왜 안 움직이는지, 동작 조건을 "
        "인터락·퍼미시브·시퀀스로 나누어 보여주고 엑셀 원본과 대조합니다.\n"
        "- 도면: 태그가 표시된 P&ID 위치와 배선 정보를 보여줍니다.\n"
        "- 4D 리포트: 조회 결과를 PDF 보고서로 출력합니다.\n\n"
        "사용자 메시지를 UI 명령 JSON 으로 해석하십시오. 명령이 아니라 "
        "질문이나 인사이면 type=chat 으로 두고 reply 에 한국어로 자연스럽게 "
        "답하십시오. 기능을 물으면 위 목록을 바탕으로 두세 문장으로 "
        "설명하고, 바로 써 볼 수 있는 예시를 한 줄 덧붙이십시오.\n"
        "Reply language: Korean.\n"
        "Schema:\n"
        '{"type":"diagnose|drawing|interlock|interlock_source|advice|help|chat|navigate",'
        '"tag":"AIT-4002 or null","tab":"alarm|interlock","alarm":"symptom text or null",'
        '"action":"OPEN|CLOSE|START|STOP or null","openSource":false,'
        '"reply":"short Korean confirmation"}\n'
        "Rules: do not invent tags; if tag missing ask in reply with type=chat. "
        "For alarm search type=diagnose. For P&ID type=drawing. For interlock type=interlock."
    )
    user = "current_tab=%s current_tag=%s\nuser: %s" % (req.tab, req.tag, text)

    try:
        content = _CHAT[provider](
            [{"role": "system", "content": system},
             {"role": "user", "content": user}], 45)
        data = _parse(content)
        out = finalize(data, "llm")
        out["provider"] = provider
        # 규칙이 구체적인 안내를 갖고 있으면 그것을 우선한다
        # (없는 태그·태그 누락 등). 다만 포괄 응답(generic)은 예외다 —
        # 그건 "아무것도 못 알아들었다"는 뜻이라, LLM 의 답을 덮으면
        # "너는 어떤 기능이 있니" 같은 질문에 예시만 반복하게 된다.
        if (out["type"] == "chat" and rule.get("reply")
                and not rule.get("generic")):
            out["reply"] = rule["reply"]
        out.pop("generic", None)
        return out
    except Exception as e:                                  # noqa: BLE001
        out = finalize(rule, "rule_fallback")
        out["llm_error"] = str(e)[:160]
        return out


@app.get("/api/chat/status")
def chat_status():
    provider = os.environ.get("COPILOT_PROVIDER", "rule").lower()
    has_key = bool(os.environ.get("OPENAI_API_KEY") or os.environ.get("AZURE_OPENAI_API_KEY"))
    return {
        "provider": provider,
        "llm_ready": provider in ("openai", "azure", "llm") and has_key,
        "model": os.environ.get("COPILOT_CHAT_MODEL") or os.environ.get("COPILOT_MODEL") or "gpt-4o-mini",
        "base_url": os.environ.get("OPENAI_BASE_URL", ""),
    }


# ── PDF / 도면 렌더 (v1 manuals.py 동일 방식) ────────────────
def _norm_name(name):
    import re as _re
    return _re.sub(r"[^a-z0-9]", "", (name or "").lower())


def _list_pdfs(folder):
    """하위 폴더까지 훑는다 — 실물 매뉴얼은 벤더별 폴더로 들어온다."""
    if not os.path.isdir(folder):
        return []
    out = []
    for root, _dirs, files in os.walk(folder):
        for f in files:
            if f.lower().endswith(".pdf") and not f.startswith("~$"):
                rel = os.path.relpath(os.path.join(root, f),
                                      folder).replace("\\", "/")
                # 벤더 폴더 이름이 한글이면 파일시스템 인코딩에 따라
                # 디코딩되지 않은 바이트가 섞여 들어온다. 그대로 JSON 으로
                # 내보내면 응답 전체가 500 으로 죽는다 — 매뉴얼 목록이
                # 통째로 안 뜨는데 이유는 안 보이는 형태가 된다.
                out.append(rel)
    return sorted(out)


def _safe_name(s):
    """JSON 으로 내보낼 수 있는 이름.

    벤더 폴더 이름이 한글인데 파일시스템 인코딩과 어긋나면 디코딩되지
    않은 바이트가 문자열에 남는다. 그대로 응답에 실으면 JSON 인코딩에서
    500 이 나 매뉴얼 목록이 통째로 안 뜬다. 윈도우에서는 대개 그대로
    통과하므로 이 함수는 아무것도 바꾸지 않는다.
    """
    return str(s).encode("utf-8", "replace").decode("utf-8")


def _resolve_pdf(file_name, folder):
    if not file_name or not os.path.isdir(folder):
        return None
    exact = os.path.join(folder, file_name)
    if os.path.exists(exact):
        return exact
    # 목록을 안전한 이름으로 내보냈다면 그 이름으로 되돌아온다 — 원래
    # 파일로 되짚는다. 이 되짚기가 없으면 목록은 뜨는데 클릭하면 안 열린다.
    for f in _list_pdfs(folder):
        if _safe_name(f) == file_name:
            return os.path.join(folder, f)
    want = _norm_name(os.path.splitext(file_name)[0])
    best, best_len = None, 0
    for f in _list_pdfs(folder):
        got = _norm_name(os.path.splitext(os.path.basename(f))[0])
        if want and (want in got or got in want):
            n = min(len(want), len(got))
            if n > best_len:
                best, best_len = os.path.join(folder, f), n
    return best


def _open_pdf(path):
    """PDF 를 **바이트로 읽어서** 연다.

    경로를 그대로 넘기면 벤더 폴더 이름이 비ASCII 일 때 MuPDF 의 C 계층이
    경로를 받지 못해 열기부터 실패한다. 파일은 멀쩡한데 화면에는 '렌더
    실패' 만 뜬다. 바이트로 넘기면 경로 인코딩이 개입하지 않는다.
    """
    import fitz
    with open(path, "rb") as fh:
        return fitz.open(stream=fh.read(), filetype="pdf")


def _render_page(path, page_no, dpi=150):
    try:
        import fitz
    except ImportError:
        return None
    if not path or not os.path.exists(path):
        return None
    try:
        with _open_pdf(path) as d:
            i = max(1, min(int(page_no), len(d))) - 1
            return d[i].get_pixmap(dpi=dpi).tobytes("png")
    except Exception:
        return None


def _render_drawing(path, page_no, tag=None, dpi=140, margin=110, band=False):
    try:
        import fitz
    except ImportError:
        return None, None, 0
    if not path or not os.path.exists(path):
        return None, None, 0
    try:
        with _open_pdf(path) as d:
            i = max(1, min(int(page_no), len(d))) - 1
            pg = d[i]
            terms = [t for t in (tag or "").split("|") if t]
            rects, focus = [], []
            for t in terms:
                got = pg.search_for(t)
                rects.extend(got)
                if got and not focus:
                    focus = got
            if rects:
                sh = pg.new_shape()
                for r in rects:
                    sh.draw_rect(fitz.Rect(r.x0 - 9, r.y0 - 9, r.x1 + 9, r.y1 + 9))
                sh.finish(color=(0.82, 0.0, 0.17), width=1.6)
                sh.commit()
            full = pg.get_pixmap(dpi=dpi).tobytes("png")
            crop = None
            if focus:
                r = focus[0]
                if band:
                    box = fitz.Rect(pg.rect.x0 + 8, r.y0 - margin,
                                    pg.rect.x1 - 8, r.y1 + margin) & pg.rect
                    scale = 1.4
                else:
                    box = fitz.Rect(r.x0 - margin, r.y0 - margin,
                                    r.x1 + margin, r.y1 + margin) & pg.rect
                    scale = 2.0
                crop = pg.get_pixmap(dpi=int(dpi * scale), clip=box).tobytes("png")
            return full, crop, len(rects)
    except Exception:
        return None, None, 0


@app.get("/api/manual-status")
def manual_status():
    try:
        import fitz  # noqa: F401
        have = True
    except ImportError:
        have = False
    files = _list_pdfs(config.MANUAL_DIR)
    demo_pdfs = _list_pdfs(config.DRAWING_DIR) if os.path.isdir(config.DRAWING_DIR) else []
    return {
        "pymupdf": have,
        "manual_dir": config.MANUAL_DIR,
        "manual_count": len(files),
        "manuals": [_safe_name(f) for f in files],
        "demo_dir": config.DRAWING_DIR,
        "demo_pdfs": [_safe_name(f) for f in demo_pdfs],
    }


@app.get("/api/manual-page")
def manual_page(
    file: str = Query(..., description="PDF 파일명"),
    page: int = Query(1, ge=1),
    dpi: int = Query(160, ge=72, le=400),
):
    path = _resolve_pdf(file, config.MANUAL_DIR)
    if not path:
        # demo_data 쪽도 한번 찾아봄
        path = _resolve_pdf(file, config.DRAWING_DIR)
    if not path:
        raise HTTPException(404, f"PDF 없음: {file} (dir={config.MANUAL_DIR})")
    png = _render_page(path, page, dpi=dpi)
    if not png:
        raise HTTPException(500, "페이지 렌더 실패 (pymupdf 확인)")
    # HTTP 헤더는 latin-1 만 담을 수 있다. 벤더 폴더 이름이 한글이면
    # 경로를 그대로 넣는 순간 인코딩에서 터져 응답이 500 이 된다 —
    # 렌더는 멀쩡히 끝났는데 화면에는 PDF 가 안 뜨는 형태가 된다.
    from urllib.parse import quote
    return Response(content=png, media_type="image/png", headers={
        "Cache-Control": "public, max-age=3600",
        "X-PDF-Path": quote(_safe_name(path), safe="/"),
    })


@app.get("/api/drawing-page")
def drawing_page(
    file: str = Query(...),
    page: int = Query(1, ge=1),
    find: str = Query(""),
    crop: int = Query(1, ge=0, le=1),
    dpi: int = Query(160, ge=72, le=400),
):
    # 도면 PDF 는 demo_data 에 있음
    # 자료는 전부 data/ 아래에 있다. 도면 → 매뉴얼 → data 루트 순.
    path = _resolve_pdf(file, config.DRAWING_DIR)
    if not path:
        path = _resolve_pdf(file, config.MANUAL_DIR)
    if not path:
        path = _resolve_pdf(file, config.DATA_DIR)
    if not path:
        path = _resolve_pdf(file, config.DRAWING_DIR)
    if not path:
        raise HTTPException(404, f"도면 PDF 없음: {file}")
    full, cropped, hits = _render_drawing(path, page, tag=find or None, dpi=dpi)
    if not full:
        raise HTTPException(500, "도면 렌더 실패")
    body = cropped if (crop and cropped) else full
    return Response(content=body, media_type="image/png", headers={
        "Cache-Control": "public, max-age=3600",
        "X-Hits": str(hits),
        "X-Has-Crop": "1" if cropped else "0",
    })


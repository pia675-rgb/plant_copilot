#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
make_arrangement.py — 합성 ARRANGEMENT(판넬 위치 배치도) 생성

v1 이 만든 도면 세 종은 계기 하나를 따라가는 도면이다.

    P&ID        공정상 어디에 붙어 있는 계기인가
    SCHEMATIC   IO 카드 채널부터 로컬 계기까지 어떻게 결선되어 있나
    OUTLINE     그 단자대가 판넬 '안'의 어디에 박혀 있나

빠진 층이 하나 있다. **그 판넬 자체가 현장 어디에 있는가.**
정비원이 실제로 이동해야 하는 대상은 계기가 아니라 판넬이다.

이 스크립트는 Instrument List 의 PANEL 열에 실제로 나타나는 판넬만
읽어서 배치도를 그린다. 판넬 위치·구역은 이 파일 상단의 PANELS/AREAS 가
유일한 원천이고, 같은 값이 PANEL_LOCATIONS.csv 로 함께 떨어진다.
도면과 조회 결과가 어긋나지 않게 하려는 것이다.

    python make_arrangement.py --src . --out .

의존성: reportlab, openpyxl

주의: 데모용 가상 도면이다. 실제 플랜트 배치가 아니다.
"""

import argparse
import csv
import os
import sys
from collections import defaultdict

from openpyxl import load_workbook
from reportlab.lib.pagesizes import A3, landscape
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas

PAGE = landscape(A3)
PROJECT_TITLE = "DEMO WATER PLANT Ph1"
PROJECT_NO = "99001D"
MAKER = "DEMO ENGINEERING CO., LTD."
ARR_PREFIX = "DM-PNT1C01-UW-E20-"
ARR_START = 40001

# ── 도면 좌표계 (mm) ──────────────────────────────────────────
# 그리드: 가로 A~H (45mm), 세로 1~5 (40mm). 도면 관례대로 좌하단 원점.
GRID_X0, GRID_Y0 = 30.0, 40.0
GRID_W, GRID_H = 45.0, 40.0
GRID_COLS, GRID_ROWS = 8, 5

# 구역 — (이름, 부제, x, y, 폭, 높이, 실내여부)
AREAS = [
    ("ELECTRICAL ROOM", "INDOOR / AIR CONDITIONED",
     30.0, 158.0, 150.0, 82.0, True),
    ("PROCESS AREA - WEST", "PRETREATMENT / RECLAIM",
     30.0, 40.0, 150.0, 108.0, False),
    ("PROCESS AREA - EAST", "RO / CEDI / POLISHING / DISTRIBUTION",
     190.0, 40.0, 200.0, 200.0, False),
]

# 공정 설비 — (약칭, 중심x, 중심y, 폭, 높이)  배치 맥락용, 조회 대상 아님
EQUIPMENT = [
    ("PRETREATMENT FILTER TRAIN", 70.0, 118.0, 62.0, 16.0),
    ("RECLAIM TANK",             148.0, 118.0, 22.0, 22.0),
    ("RAW WATER PUMP PIT",        70.0,  62.0, 62.0, 16.0),
    ("RO SKID  1st / 2nd PASS",  240.0, 205.0, 80.0, 18.0),
    ("CEDI SKID",                240.0, 165.0, 60.0, 16.0),
    ("UPW TANK",                 345.0, 195.0, 34.0, 34.0),
    ("POLISHING LOOP SKID",      240.0, 100.0, 80.0, 18.0),
    ("DISTRIBUTION HEADER",      330.0, 100.0, 46.0, 14.0),
]

# 판넬 — (이름, 중심x, 중심y, 폭, 깊이, 종류, 설치구역, 비고)
PANELS = [
    ("CUB-A",  62.0, 212.0, 26.0, 14.0, "CONTROL CUBICLE",
     "ELECTRICAL ROOM", "PLC-UPW-01 MAIN RACK"),
    ("CUB-B", 104.0, 212.0, 26.0, 14.0, "CONTROL CUBICLE",
     "ELECTRICAL ROOM", "PLC-UPW-01 EXT RACK"),
    ("CUB-C", 146.0, 212.0, 26.0, 14.0, "CONTROL CUBICLE",
     "ELECTRICAL ROOM", "PLC-UPW-01 EXT RACK"),
    ("CUB-D",  62.0, 182.0, 26.0, 14.0, "OUTPUT CUBICLE",
     "ELECTRICAL ROOM", "PLC-UPW-01 DO/AO RACK"),
    ("RIO-01", 285.0, 140.0, 24.0, 13.0, "REMOTE I/O PANEL",
     "PROCESS AREA - EAST", "ET200SP HA / PROFINET RING"),
    ("RIO-02", 105.0,  90.0, 24.0, 13.0, "REMOTE I/O PANEL",
     "PROCESS AREA - WEST", "ET200SP HA / PROFINET RING"),
]

# 케이블 트레이 경로 — 전기실에서 원격 판넬까지 (꺾은선)
TRAYS = [
    ("TR-01", [(146.0, 205.0), (168.0, 205.0), (168.0, 250.0),
               (285.0, 250.0), (285.0, 147.0)]),
    ("TR-02", [(146.0, 205.0), (168.0, 205.0), (168.0, 152.0),
               (105.0, 152.0), (105.0, 97.0)]),
]


# ─────────────────────────────────────────────────────────────
def grid_of(cx, cy):
    """중심 좌표 → 도면 그리드 표기 (예: B4)."""
    col = int((cx - GRID_X0) // GRID_W)
    row = int((cy - GRID_Y0) // GRID_H) + 1
    col = max(0, min(GRID_COLS - 1, col))
    row = max(1, min(GRID_ROWS, row))
    return "%s%d" % (chr(ord("A") + col), row)


def load_instruments(p):
    """계기 리스트는 config.INSTRUMENTS 한 곳에서만 읽는다."""
    if not os.path.exists(p):
        raise SystemExit("계기 리스트를 찾을 수 없습니다: %s" % p)
    ws = load_workbook(p, read_only=True, data_only=True).active
    rows = list(ws.iter_rows(values_only=True))
    hi = next(i for i, r in enumerate(rows)
              if r and "TAG" in [str(c).strip().upper() if c else "" for c in r])
    hdr = [str(c).strip().upper() if c else "" for c in rows[hi]]
    I = {h: i for i, h in enumerate(hdr) if h}
    need = ("TAG", "PANEL")
    for n in need:
        if n not in I:
            raise SystemExit("계기 리스트에 %s 열이 없습니다." % n)
    out = []
    for r in rows[hi + 1:]:
        if not r or not r[I["TAG"]]:
            continue
        out.append({h: (r[i] if i < len(r) else None) for h, i in I.items()})
    return out


def border(c, w, h, sheet_no, title, rev="0"):
    """도면 테두리 + 표제란 + 그리드 눈금."""
    c.setLineWidth(1.2)
    c.rect(10 * mm, 10 * mm, w - 20 * mm, h - 20 * mm)
    c.setLineWidth(0.4)
    c.rect(14 * mm, 14 * mm, w - 28 * mm, h - 28 * mm)

    # 그리드 눈금 (테두리 안쪽 가장자리)
    c.setFont("Helvetica", 5)
    for i in range(GRID_COLS):
        x = (GRID_X0 + i * GRID_W + GRID_W / 2) * mm
        c.drawCentredString(x, h - 19 * mm, chr(ord("A") + i))
        c.drawCentredString(x, 15.5 * mm, chr(ord("A") + i))
    for j in range(GRID_ROWS):
        y = (GRID_Y0 + j * GRID_H + GRID_H / 2) * mm
        c.drawCentredString(16 * mm, y, str(j + 1))
        c.drawCentredString(w - 16 * mm, y, str(j + 1))

    # 표제란
    tw, th = 95 * mm, 26 * mm
    tx, ty = w - 14 * mm - tw, 14 * mm
    c.setLineWidth(0.7)
    c.rect(tx, ty, tw, th)
    c.line(tx, ty + 13 * mm, tx + tw, ty + 13 * mm)
    c.line(tx, ty + 6.5 * mm, tx + tw, ty + 6.5 * mm)
    c.setFont("Helvetica-Bold", 8)
    c.drawString(tx + 3 * mm, ty + 20 * mm, PROJECT_TITLE)
    c.setFont("Helvetica", 6)
    c.drawString(tx + 3 * mm, ty + 16 * mm, MAKER)
    c.setFont("Helvetica-Bold", 7)
    c.drawString(tx + 3 * mm, ty + 8.8 * mm, title)
    c.setFont("Helvetica", 6)
    c.drawString(tx + 3 * mm, ty + 3 * mm, "SHEET No.  %s" % sheet_no)
    c.drawString(tx + 58 * mm, ty + 3 * mm, "PROJ. %s" % PROJECT_NO)
    c.drawString(tx + 82 * mm, ty + 3 * mm, "REV. %s" % rev)

    # 데모 고지
    c.setFont("Helvetica-Oblique", 5.5)
    c.drawString(16 * mm, h - 15.5 * mm,
                 "SYNTHETIC DRAWING FOR DEMONSTRATION ONLY - NOT AN ACTUAL "
                 "PLANT LAYOUT")


def build_arrangement(inst_rows, out_dir):
    """1페이지: 배치 평면도 / 2페이지: 판넬 일람표."""
    w, h = PAGE
    path = os.path.join(out_dir, "DEMO_ARRANGEMENT.pdf")
    c = canvas.Canvas(path, pagesize=PAGE)
    c.setTitle("Synthetic panel arrangement for demo")

    by_panel = defaultdict(list)
    for r in inst_rows:
        by_panel[str(r.get("PANEL") or "").strip()].append(r)

    drawn = {p[0] for p in PANELS}
    missing = sorted(k for k in by_panel if k and k not in drawn)
    if missing:
        print("[경고] 계기 리스트에 있으나 배치도에 정의되지 않은 판넬: %s"
              % ", ".join(missing))

    # ── 페이지 1 : 평면 배치도 ──────────────────────────────
    sheet1 = "%s%06d" % (ARR_PREFIX, ARR_START)
    border(c, w, h, sheet1, "PANEL LOCATION ARRANGEMENT  /  PLOT PLAN")

    # 그리드선 — 도면에 찍힌 그리드 표기를 눈으로 검증할 수 있게
    c.setStrokeGray(0.78)
    c.setLineWidth(0.25)
    c.setDash(0.8, 2.5)
    for i in range(1, GRID_COLS):
        x = (GRID_X0 + i * GRID_W) * mm
        c.line(x, GRID_Y0 * mm, x, (GRID_Y0 + GRID_ROWS * GRID_H) * mm)
    for j in range(1, GRID_ROWS):
        y = (GRID_Y0 + j * GRID_H) * mm
        c.line(GRID_X0 * mm, y, (GRID_X0 + GRID_COLS * GRID_W) * mm, y)
    c.setDash()
    c.setStrokeGray(0)

    # 구역
    for name, sub, x, y, aw, ah, indoor in AREAS:
        c.setLineWidth(0.9 if indoor else 0.5)
        if not indoor:
            c.setDash(3, 2)
        c.rect(x * mm, y * mm, aw * mm, ah * mm)
        c.setDash()
        c.setFont("Helvetica-Bold", 7)
        c.drawString((x + 3) * mm, (y + ah - 6) * mm, name)
        c.setFont("Helvetica", 5)
        c.drawString((x + 3) * mm, (y + ah - 10) * mm, sub)

    # 공정 설비 (배치 맥락용 — 조회 대상 아님)
    c.setStrokeGray(0.45)
    c.setFillGray(0.45)
    for name, ex, ey, ew, eh in EQUIPMENT:
        c.setLineWidth(0.4)
        c.rect((ex - ew / 2) * mm, (ey - eh / 2) * mm, ew * mm, eh * mm)
        c.setFont("Helvetica", 4.8)
        c.drawCentredString(ex * mm, ey * mm, name)
    c.setStrokeGray(0)
    c.setFillGray(0)

    # 케이블 트레이
    c.setLineWidth(0.5)
    c.setDash(1.5, 1.5)
    for name, pts in TRAYS:
        best, blen = None, -1
        for i in range(len(pts) - 1):
            c.line(pts[i][0] * mm, pts[i][1] * mm,
                   pts[i + 1][0] * mm, pts[i + 1][1] * mm)
            # 라벨은 가장 긴 수평 구간 위에 얹는다 (설비·판넬과 겹치지 않게)
            if abs(pts[i][1] - pts[i + 1][1]) < 0.1:
                d = abs(pts[i][0] - pts[i + 1][0])
                if d > blen:
                    blen, best = d, ((pts[i][0] + pts[i + 1][0]) / 2, pts[i][1])
        if best:
            c.setFont("Helvetica", 4.5)
            c.drawCentredString(best[0] * mm, (best[1] + 1.8) * mm,
                                "CABLE TRAY %s" % name)
    c.setDash()

    # 방위 + 축척
    nx, ny = 60.0, 256.0
    c.setLineWidth(0.7)
    c.line(nx * mm, (ny - 7) * mm, nx * mm, (ny + 7) * mm)
    c.line(nx * mm, (ny + 7) * mm, (nx - 2.5) * mm, (ny + 2) * mm)
    c.line(nx * mm, (ny + 7) * mm, (nx + 2.5) * mm, (ny + 2) * mm)
    c.setFont("Helvetica-Bold", 6)
    c.drawCentredString(nx * mm, (ny - 11) * mm, "N")
    sx = 90.0
    c.setLineWidth(0.6)
    c.line(sx * mm, ny * mm, (sx + 40) * mm, ny * mm)
    c.setFont("Helvetica", 5)
    for k in range(5):
        xx = (sx + k * 10) * mm
        c.line(xx, ny * mm, xx, (ny + 2) * mm)
    c.drawString(sx * mm, (ny - 4) * mm, "0")
    c.drawString((sx + 38) * mm, (ny - 4) * mm, "20 m")
    c.drawString(sx * mm, (ny + 4) * mm, "SCALE 1:500 (A3)")

    # 판넬
    for name, cx, cy, pw, pd, kind, area, note in PANELS:
        x, y = (cx - pw / 2), (cy - pd / 2)
        c.setLineWidth(1.1)
        c.rect(x * mm, y * mm, pw * mm, pd * mm)
        c.setLineWidth(0.35)
        c.rect((x + 1.2) * mm, (y + 1.2) * mm,
               (pw - 2.4) * mm, (pd - 2.4) * mm)
        # 정면(문) 방향 표시
        c.setLineWidth(0.8)
        c.line(x * mm, y * mm, (x + pw) * mm, y * mm)
        c.setFont("Helvetica-Bold", 8)
        c.drawCentredString(cx * mm, (cy + 1.0) * mm, name)
        c.setFont("Helvetica", 4.6)
        c.drawCentredString(cx * mm, (cy - 3.2) * mm, kind)
        c.setFont("Helvetica-Bold", 5.2)
        c.drawCentredString(cx * mm, (y - 4.2) * mm,
                            "GRID %s   (%d PT)" % (grid_of(cx, cy),
                                                   len(by_panel.get(name, []))))

    # 범례
    lx, ly = 192.0, 18.0
    c.setLineWidth(0.5)
    c.rect(lx * mm, ly * mm, 88 * mm, 22 * mm)
    c.setFont("Helvetica-Bold", 6)
    c.drawString((lx + 2) * mm, (ly + 17.5) * mm, "LEGEND")
    c.setFont("Helvetica", 5)
    c.rect((lx + 3) * mm, (ly + 11) * mm, 8 * mm, 4 * mm)
    c.drawString((lx + 13) * mm, (ly + 12.2) * mm,
                 "PANEL (bold line = front / door side)")
    c.setDash(1.5, 1.5)
    c.line((lx + 3) * mm, (ly + 7) * mm, (lx + 11) * mm, (ly + 7) * mm)
    c.setDash()
    c.drawString((lx + 13) * mm, (ly + 6) * mm, "CABLE TRAY ROUTE")
    c.drawString((lx + 3) * mm, (ly + 2.5) * mm,
                 "( n PT ) = number of instrument points wired to the panel")

    c.showPage()

    # ── 페이지 2 : 판넬 일람표 ──────────────────────────────
    sheet2 = "%s%06d" % (ARR_PREFIX, ARR_START + 1)
    border(c, w, h, sheet2, "PANEL LOCATION ARRANGEMENT  /  PANEL SCHEDULE")

    cols = [(30, "PANEL", 26), (56, "TYPE", 40), (96, "AREA", 52),
            (148, "GRID", 16), (164, "POINTS", 20), (184, "TB GROUPS", 96),
            (280, "REMARK", 100)]
    ty = 230.0
    c.setFont("Helvetica-Bold", 6.5)
    for x, label, _ in cols:
        c.drawString(x * mm, ty * mm, label)
    c.setLineWidth(0.6)
    c.line(30 * mm, (ty - 2) * mm, 380 * mm, (ty - 2) * mm)

    ty -= 8
    for name, cx, cy, pw, pd, kind, area, note in PANELS:
        items = by_panel.get(name, [])
        tbs = sorted({str(r.get("TERMINAL") or "").split("-")[0]
                      for r in items if r.get("TERMINAL")})
        c.setFont("Helvetica-Bold", 6.5)
        c.drawString(30 * mm, ty * mm, name)
        c.setFont("Helvetica", 6)
        c.drawString(56 * mm, ty * mm, kind)
        c.drawString(96 * mm, ty * mm, area)
        c.drawString(148 * mm, ty * mm, grid_of(cx, cy))
        c.drawString(164 * mm, ty * mm, str(len(items)))
        c.setFont("Helvetica", 5)
        txt = ", ".join(tbs)
        c.drawString(184 * mm, ty * mm, txt[:64])
        if len(txt) > 64:
            c.drawString(184 * mm, (ty - 3.2) * mm, txt[64:128])
        c.drawString(280 * mm, ty * mm, note)
        c.setLineWidth(0.2)
        c.line(30 * mm, (ty - 5.5) * mm, 380 * mm, (ty - 5.5) * mm)
        ty -= 13

    c.setFont("Helvetica-Oblique", 5.5)
    c.drawString(30 * mm, 46 * mm,
                 "POINTS / TB GROUPS are derived from the IO List. "
                 "They are not drafted values.")
    c.showPage()
    c.save()
    return path, {p[0]: (sheet1, 1) for p in PANELS}


def write_locations(inst_rows, out_dir, sheet_map):
    """도면과 같은 원천에서 조회용 CSV 를 떨군다."""
    by_panel = defaultdict(list)
    for r in inst_rows:
        by_panel[str(r.get("PANEL") or "").strip()].append(r)

    path = os.path.join(out_dir, "PANEL_LOCATIONS.csv")
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        wr = csv.writer(f)
        wr.writerow(["PANEL", "KIND", "AREA", "INDOOR", "GRID",
                     "WIDTH_MM", "DEPTH_MM", "POINTS",
                     "FILE", "SHEET_NO", "PAGE", "FIND", "REMARK"])
        indoor = {a[0]: a[6] for a in AREAS}
        for name, cx, cy, pw, pd, kind, area, note in PANELS:
            sheet_no, page = sheet_map.get(name, ("", 1))
            wr.writerow([name, kind, area,
                         "Y" if indoor.get(area) else "N",
                         grid_of(cx, cy),
                         int(pw * 10), int(pd * 10),
                         len(by_panel.get(name, [])),
                         "DEMO_ARRANGEMENT.pdf", sheet_no, page, name, note])
    return path


def main():
    ap = argparse.ArgumentParser(description="합성 판넬 배치도 생성")
    here = os.path.dirname(os.path.abspath(__file__))
    ap.add_argument("--src", default=None,
                    help="계기 리스트 파일 경로 (기본: config.INSTRUMENTS)")
    ap.add_argument("--out", default=None,
                    help="출력 폴더 (기본: data/drawings)")
    a = ap.parse_args()

    # 계기 리스트를 여기서 따로 찾지 않는다.
    #
    # 예전에는 이 폴더의 사본을 먼저 보고 없으면 config 로 넘어갔다.
    # 그 사본이 오래되면 배치도의 판넬별 점수(N PT)와 조회 결과가
    # 말없이 어긋난다. 실제로 계기 4점이 추가됐을 때 도면만 72점에
    # 머물러 있었다. 원천을 하나로 고정한다.
    sys.path.insert(0, os.path.dirname(here))
    import config
    src = a.src or config.INSTRUMENTS
    print("계기 리스트:", src)

    rows = load_instruments(src)
    # 도면은 drawings/, 조회용 CSV 는 data/ 루트.
    out_dwg = a.out or config.DRAWING_DIR
    os.makedirs(out_dwg, exist_ok=True)
    pdf, sheet_map = build_arrangement(rows, out_dwg)
    os.makedirs(config.DERIVED_DIR, exist_ok=True)
    csv_path = write_locations(rows, config.DERIVED_DIR, sheet_map)
    print("계기 %d건 / 판넬 %d개" % (len(rows), len(PANELS)))
    print("도면 :", pdf)
    print("위치 :", csv_path)


if __name__ == "__main__":
    main()

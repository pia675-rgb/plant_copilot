#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
make_interlock_list.py — 합성 인터락 리스트 생성

실제 인터락 리스트를 쓸 수 없으므로(도면과 같은 이유) 데모용으로 만든다.
다만 **깨끗하게 만들지 않는다.** 실제 서류의 다음 성질을 일부러 재현한다.

    · 제목 줄이 위에 있고 머리행이 3행쯤에서 시작
    · IL_NO 가 첫 행에만 있고 조건 행은 공란 (병합 셀 흉내)
    · 조건이 자연어 — "LIT-4003 HH (95%) 3sec delay", "PIT-2003 저압 < 0.8 bar"
    · 한글과 영문이 섞임
    · 단위 표기가 제각각 (%, %, bar, Mohm-cm)

깨끗한 CSV 를 만들어 놓고 파싱한다고 하면 데모로서 의미가 없다.
파싱의 어려움 자체가 이 기능의 기술적 내용이다.

산출:
    DEMO_OUTPUT_LIST.xlsx     출력 태그(밸브·펌프·기타) 22건
    DEMO_INTERLOCK_LIST.xlsx  인터락 리스트 (항목 30건 / 조건 행 80여 건)

사용:
    python -m data.make_interlock_list --out data
"""

import argparse
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ── 출력 태그 ────────────────────────────────────────────────
# (TAG, SERVICE, SYSTEM, TYPE, FAIL)
OUTPUTS = [
    ("P-1101A", "RAW WATER TRANSFER PUMP A", "PRETREATMENT", "PUMP", "FAIL STOP", "MCC"),
    ("P-1101B", "RAW WATER TRANSFER PUMP B", "PRETREATMENT", "PUMP", "FAIL STOP", "MCC"),
    ("XV-1102", "RAW WATER INLET VALVE", "PRETREATMENT", "ON-OFF VALVE", "FAIL CLOSE", "SOLENOID"),
    ("XV-1103", "MEDIA FILTER BACKWASH VALVE", "PRETREATMENT", "ON-OFF VALVE", "FAIL CLOSE", "SOLENOID"),
    ("P-2101A", "RO HIGH PRESSURE PUMP A", "RO SYSTEM", "PUMP", "FAIL STOP", "VFD"),
    ("P-2101B", "RO HIGH PRESSURE PUMP B", "RO SYSTEM", "PUMP", "FAIL STOP", "VFD"),
    ("XV-2102", "RO FEED VALVE", "RO SYSTEM", "ON-OFF VALVE", "FAIL CLOSE", "SOLENOID"),
    ("XV-2103", "RO CONCENTRATE DRAIN VALVE", "RO SYSTEM", "ON-OFF VALVE", "FAIL OPEN", "SOLENOID"),
    ("XV-2104", "RO PERMEATE DIVERT VALVE", "RO SYSTEM", "ON-OFF VALVE", "FAIL CLOSE", "SOLENOID"),
    ("P-3101", "CEDI FEED PUMP", "CEDI SYSTEM", "PUMP", "FAIL STOP", "VFD"),
    ("XV-3102", "CEDI INLET VALVE", "CEDI SYSTEM", "ON-OFF VALVE", "FAIL CLOSE", "SOLENOID"),
    ("XV-3103", "CEDI CONCENTRATE VALVE", "CEDI SYSTEM", "ON-OFF VALVE", "FAIL OPEN", "SOLENOID"),
    ("R-3104", "CEDI DC RECTIFIER", "CEDI SYSTEM", "RECTIFIER", "FAIL OFF", "PANEL"),
    ("XV-4101", "UPW TANK INLET VALVE", "UPW TANK", "ON-OFF VALVE", "FAIL CLOSE", "SOLENOID"),
    ("XV-4102", "UPW TANK DRAIN VALVE", "UPW TANK", "ON-OFF VALVE", "FAIL CLOSE", "SOLENOID"),
    ("P-5101A", "POLISHING CIRCULATION PUMP A", "POLISHING LOOP", "PUMP", "FAIL STOP", "VFD"),
    ("P-5101B", "POLISHING CIRCULATION PUMP B", "POLISHING LOOP", "PUMP", "FAIL STOP", "VFD"),
    ("UV-5102", "POLISHING UV TOC REDUCER", "POLISHING LOOP", "UV UNIT", "FAIL OFF", "PANEL"),
    ("XV-5103", "POLISHING LOOP INLET VALVE", "POLISHING LOOP", "ON-OFF VALVE", "FAIL CLOSE", "SOLENOID"),
    ("P-6101A", "DISTRIBUTION PUMP A", "DISTRIBUTION", "PUMP", "FAIL STOP", "VFD"),
    ("XV-6102", "POU SUPPLY VALVE", "DISTRIBUTION", "ON-OFF VALVE", "FAIL CLOSE", "SOLENOID"),
    ("P-7101", "RECLAIM TRANSFER PUMP", "RECLAIM", "PUMP", "FAIL STOP", "MCC"),
]

# ── 인터락 항목 ──────────────────────────────────────────────
# (IL_NO, OUTPUT_TAG, ACTION, KIND, LOGIC, RESET, BYPASSABLE, PRIORITY,
#  PLC_BLOCK, DWG_NO, SHEET, [(GROUP, COND_TEXT), ...], REMARK)
IL = [
    # ── UPW TANK 유입 밸브 — 데모의 주인공 ────────────────
    ("IL-4101-01", "XV-4101", "CLOSE", "INTERLOCK", "OR", "MANUAL", "N", 1,
     "FB_XV4101_ILK", "YI-Y01-C01-UW-LG-041", "3",
     [("G1", "LIT-4003 High High (95 %) 3sec delay"),
      ("G1", "AIT-4001 저항률 < 17 Mohm-cm, 10초 지연"),
      ("G1", "AIT-4002 TOC High High (5 ppb) 30sec"),
      ("G1", "E-STOP PB 동작 (EMERGENCY STOP)")],
     "안전 인터락. 상위 우선. 바이패스 불가."),
    ("IL-4101-02", "XV-4101", "OPEN", "PERMISSIVE", "AND", "AUTO", "Y", 2,
     "FB_XV4101_PERM", "YI-Y01-C01-UW-LG-041", "3",
     [("G1", "LIT-4003 < 90 %"),
      ("G1", "P-3101 운전 중 (RUN F/B)"),
      ("G1", "XV-3102 OPEN 확인"),
      ("G1", "AIT-3001 저항률 >= 17 Mohm-cm")],
     "기동 허가 조건. 전부 만족해야 열림 허용."),
    ("IL-4101-03", "XV-4101", "OPEN", "SEQUENCE", "AND", "AUTO", "Y", 3,
     "SFC_TANK_FILL", "YI-Y01-C01-UW-LG-041", "4",
     [("G1", "TANK FILL 시퀀스 STEP 3 진입"),
      ("G1", "LIT-4003 < 80 % (보충 개시점)")],
     "자동 보충 운전."),
    ("IL-4101-04", "XV-4101", "OPEN", "MANUAL", "AND", "MANUAL", "Y", 4,
     "HMI_XV4101", "YI-Y01-C01-UW-LG-041", "4",
     [("G1", "HMI 수동 모드 선택"),
      ("G1", "운전원 OPEN 명령")],
     "수동 조작. 인터락은 여전히 우선."),

    # ── UPW TANK 드레인 ──────────────────────────────────
    ("IL-4102-01", "XV-4102", "CLOSE", "INTERLOCK", "OR", "MANUAL", "N", 1,
     "FB_XV4102_ILK", "YI-Y01-C01-UW-LG-041", "5",
     [("G1", "LIT-4003 Low Low (10 %) 5sec"),
      ("G1", "P-5101A 또는 P-5101B 운전 중")],
     "저수위 드레인 방지."),
    ("IL-4102-02", "XV-4102", "OPEN", "PERMISSIVE", "AND", "MANUAL", "Y", 2,
     "FB_XV4102_PERM", "YI-Y01-C01-UW-LG-041", "5",
     [("G1", "LIT-4003 > 20 %"),
      ("G1", "DRAIN 허가 키스위치 ON")],
     ""),

    # ── RO 고압 펌프 A ───────────────────────────────────
    ("IL-2101A-01", "P-2101A", "STOP", "INTERLOCK", "OR", "MANUAL", "N", 1,
     "FB_P2101A_ILK", "YI-Y01-C01-UW-LG-021", "2",
     [("G1", "PIT-2003 Low Low < 0.8 bar, 5초 지연 (흡입 저압)"),
      ("G1", "FIT-2004 Low Low < 4.3 m3/h 10sec"),
      ("G1", "MOTOR OVERLOAD TRIP (49)"),
      ("G1", "E-STOP PB 동작")],
     "펌프 보호. 공회전·과부하 방지."),
    ("IL-2101A-02", "P-2101A", "START", "PERMISSIVE", "AND", "AUTO", "Y", 2,
     "FB_P2101A_PERM", "YI-Y01-C01-UW-LG-021", "2",
     [("G1", "XV-2102 OPEN 확인 (개도 F/B)"),
      ("G1", "PIT-2003 >= 1.0 bar"),
      ("G1", "MCC READY 신호 정상"),
      ("G1", "VFD FAULT 없음")],
     ""),
    ("IL-2101A-03", "P-2101A", "START", "SEQUENCE", "AND", "AUTO", "Y", 3,
     "SFC_RO_START", "YI-Y01-C01-UW-LG-021", "3",
     [("G1", "RO START 시퀀스 STEP 2"),
      ("G1", "P-2101B 정지 상태 (교대 운전)")],
     "A/B 교대 운전 로직."),

    # ── RO 급수 밸브 ─────────────────────────────────────
    ("IL-2102-01", "XV-2102", "CLOSE", "INTERLOCK", "OR", "MANUAL", "N", 1,
     "FB_XV2102_ILK", "YI-Y01-C01-UW-LG-021", "4",
     [("G1", "AIT-2002 전도도 High High > 90.27 uS/cm"),
      ("G1", "PIT-2003 High High > 8.83 bar 2sec"),
      ("G1", "E-STOP PB 동작")],
     ""),
    ("IL-2102-02", "XV-2102", "OPEN", "PERMISSIVE", "AND", "AUTO", "Y", 2,
     "FB_XV2102_PERM", "YI-Y01-C01-UW-LG-021", "4",
     [("G1", "LIT-2005 > 15 %"),
      ("G1", "AIT-2001 저항률 >= 17 Mohm-cm"),
      ("G1", "XV-2103 상태 정상")],
     ""),

    # ── RO 농축수 배출 (Fail Open) ───────────────────────
    ("IL-2103-01", "XV-2103", "OPEN", "INTERLOCK", "OR", "AUTO", "N", 1,
     "FB_XV2103_ILK", "YI-Y01-C01-UW-LG-021", "5",
     [("G1", "PIT-2003 High High > 8.83 bar"),
      ("G1", "P-2101A 및 P-2101B 모두 정지")],
     "Fail Open. 과압 시 강제 개방."),
    ("IL-2103-02", "XV-2103", "CLOSE", "PERMISSIVE", "AND", "AUTO", "Y", 2,
     "FB_XV2103_PERM", "YI-Y01-C01-UW-LG-021", "5",
     [("G1", "PIT-2003 < 7.5 bar"),
      ("G1", "RO 운전 중 (RUN MODE)")],
     ""),

    # ── RO 투과수 전환 밸브 ──────────────────────────────
    ("IL-2104-01", "XV-2104", "CLOSE", "INTERLOCK", "OR", "MANUAL", "N", 1,
     "FB_XV2104_ILK", "YI-Y01-C01-UW-LG-022", "1",
     [("G1", "AIT-2001 저항률 < 17 Mohm-cm 20sec"),
      ("G1", "AIT-2002 전도도 > 10.63 uS/cm")],
     "수질 미달 시 제품수 차단(divert)."),
    ("IL-2104-02", "XV-2104", "OPEN", "PERMISSIVE", "AND", "AUTO", "Y", 2,
     "FB_XV2104_PERM", "YI-Y01-C01-UW-LG-022", "1",
     [("G1", "AIT-2001 >= 17 Mohm-cm 연속 60초"),
      ("G1", "FIT-2004 >= 4.3 m3/h")],
     ""),

    # ── CEDI ────────────────────────────────────────────
    ("IL-3101-01", "P-3101", "STOP", "INTERLOCK", "OR", "MANUAL", "N", 1,
     "FB_P3101_ILK", "YI-Y01-C01-UW-LG-031", "2",
     [("G1", "PIT-3003 Low Low < 1.17 bar 5sec"),
      ("G1", "FIT-3004 Low Low < 3.71 m3/h 10sec"),
      ("G1", "MOTOR OVERLOAD TRIP (49)")],
     ""),
    ("IL-3101-02", "P-3101", "START", "PERMISSIVE", "AND", "AUTO", "Y", 2,
     "FB_P3101_PERM", "YI-Y01-C01-UW-LG-031", "2",
     [("G1", "XV-3102 OPEN 확인"),
      ("G1", "LIT-2005 > 20 %"),
      ("G1", "VFD FAULT 없음")],
     ""),
    ("IL-3104-01", "R-3104", "OFF", "INTERLOCK", "OR", "MANUAL", "N", 1,
     "FB_R3104_ILK", "YI-Y01-C01-UW-LG-031", "6",
     [("G1", "P-3101 정지"),
      ("G1", "FIT-3004 < 3.71 m3/h (무통수 통전 방지)"),
      ("G1", "TIT-1004 High > 68.84 degC"),
      ("G1", "정류기 자체 FAULT")],
     "무통수 통전은 스택을 파손시킴. 최우선 인터락."),
    ("IL-3104-02", "R-3104", "ON", "PERMISSIVE", "AND", "AUTO", "Y", 2,
     "FB_R3104_PERM", "YI-Y01-C01-UW-LG-031", "6",
     [("G1", "P-3101 운전 중"),
      ("G1", "FIT-3004 >= 3.71 m3/h 연속 30초"),
      ("G1", "XV-3103 OPEN 확인")],
     ""),
    ("IL-3102-01", "XV-3102", "CLOSE", "INTERLOCK", "OR", "MANUAL", "N", 1,
     "FB_XV3102_ILK", "YI-Y01-C01-UW-LG-031", "3",
     [("G1", "AIT-3002 전도도 High High > 80.05 uS/cm"),
      ("G1", "E-STOP PB 동작")],
     ""),
    ("IL-3102-02", "XV-3102", "OPEN", "PERMISSIVE", "AND", "AUTO", "Y", 2,
     "FB_XV3102_PERM", "YI-Y01-C01-UW-LG-031", "3",
     [("G1", "AIT-2001 >= 17 Mohm-cm"),
      ("G1", "LIT-2005 > 20 %")],
     ""),

    # ── 폴리싱 루프 ─────────────────────────────────────
    ("IL-5101A-01", "P-5101A", "STOP", "INTERLOCK", "OR", "MANUAL", "N", 1,
     "FB_P5101A_ILK", "YI-Y01-C01-UW-LG-051", "2",
     [("G1", "LIT-4003 Low Low (10 %) 5sec (탱크 저수위)"),
      ("G1", "PIT-5004 Low Low < 0.9 bar 5sec"),
      ("G1", "MOTOR OVERLOAD TRIP (49)"),
      ("G1", "E-STOP PB 동작")],
     "탱크 저수위 연동. 캐비테이션 방지."),
    ("IL-5101A-02", "P-5101A", "START", "PERMISSIVE", "AND", "AUTO", "Y", 2,
     "FB_P5101A_PERM", "YI-Y01-C01-UW-LG-051", "2",
     [("G1", "LIT-4003 > 25 %"),
      ("G1", "XV-5103 OPEN 확인"),
      ("G1", "MCC READY 신호 정상"),
      ("G1", "VFD FAULT 없음")],
     ""),
    ("IL-5102-01", "UV-5102", "OFF", "INTERLOCK", "OR", "MANUAL", "N", 1,
     "FB_UV5102_ILK", "YI-Y01-C01-UW-LG-051", "4",
     [("G1", "P-5101A 및 P-5101B 모두 정지 (무통수)"),
      ("G1", "TIT-5005 High High > 60 degC"),
      ("G1", "UV 램프 FAULT 신호")],
     "무통수 시 UV 점등 금지 — 관내 승온."),
    ("IL-5102-02", "UV-5102", "ON", "PERMISSIVE", "AND", "AUTO", "Y", 2,
     "FB_UV5102_PERM", "YI-Y01-C01-UW-LG-051", "4",
     [("G1", "P-5101A 또는 P-5101B 운전 중"),
      ("G1", "FIT-5006 >= 5 m3/h 연속 20초")],
     ""),
    ("IL-5103-01", "XV-5103", "CLOSE", "INTERLOCK", "OR", "MANUAL", "N", 1,
     "FB_XV5103_ILK", "YI-Y01-C01-UW-LG-051", "3",
     [("G1", "AIT-5001 저항률 < 17 Mohm-cm 30sec"),
      ("G1", "AIT-5003 TOC > 5 ppb 30sec"),
      ("G1", "E-STOP PB 동작")],
     ""),
    ("IL-5103-02", "XV-5103", "OPEN", "PERMISSIVE", "AND", "AUTO", "Y", 2,
     "FB_XV5103_PERM", "YI-Y01-C01-UW-LG-051", "3",
     [("G1", "LIT-4003 > 25 %"),
      ("G1", "AIT-4001 >= 17 Mohm-cm")],
     ""),

    # ── 분배 ────────────────────────────────────────────
    ("IL-6101A-01", "P-6101A", "STOP", "INTERLOCK", "OR", "MANUAL", "N", 1,
     "FB_P6101A_ILK", "YI-Y01-C01-UW-LG-061", "2",
     [("G1", "LIT-4003 Low Low (10 %)"),
      ("G1", "PIT-6003 Low Low < 1.5 bar 5sec"),
      ("G1", "MOTOR OVERLOAD TRIP (49)")],
     ""),
    ("IL-6102-01", "XV-6102", "CLOSE", "INTERLOCK", "OR", "MANUAL", "N", 1,
     "FB_XV6102_ILK", "YI-Y01-C01-UW-LG-061", "4",
     [("G1", "AIT-6002 저항률 < 17 Mohm-cm 30sec"),
      ("G1", "AIT-6004 TOC High High > 5 ppb"),
      ("G1", "E-STOP PB 동작")],
     "POU 공급 차단. FAB 직결이라 최우선."),
    ("IL-6102-02", "XV-6102", "OPEN", "PERMISSIVE", "AND", "AUTO", "N", 2,
     "FB_XV6102_PERM", "YI-Y01-C01-UW-LG-061", "4",
     [("G1", "P-6101A 운전 중"),
      ("G1", "AIT-6002 >= 17 Mohm-cm 연속 120초"),
      ("G1", "PIT-6003 >= 2.0 bar"),
      ("G1", "FAB 측 수요 신호 (DEMAND)")],
     "바이패스 불가 — 품질 직결."),

    # ── 전처리 / 회수 ───────────────────────────────────
    ("IL-1101A-01", "P-1101A", "STOP", "INTERLOCK", "OR", "MANUAL", "N", 1,
     "FB_P1101A_ILK", "YI-Y01-C01-UW-LG-011", "2",
     [("G1", "PIT-1002 Low Low < 1.31 bar 5sec"),
      ("G1", "MOTOR OVERLOAD TRIP (49)"),
      ("G1", "E-STOP PB 동작")],
     ""),
    ("IL-1102-01", "XV-1102", "CLOSE", "INTERLOCK", "OR", "MANUAL", "N", 1,
     "FB_XV1102_ILK", "YI-Y01-C01-UW-LG-011", "3",
     [("G1", "AIT-1001 전도도 High High > 88.94 uS/cm"),
      ("G1", "TIT-1004 High High > 68.84 degC")],
     ""),
    ("IL-7101-01", "P-7101", "STOP", "INTERLOCK", "OR", "MANUAL", "N", 1,
     "FB_P7101_ILK", "YI-Y01-C01-UW-LG-071", "2",
     [("G1", "LIT-7003 Low Low < 15 %"),
      ("G1", "MOTOR OVERLOAD TRIP (49)")],
     ""),
]

HDR_OUT = ["NO", "TAG", "SERVICE", "SYSTEM", "TYPE", "FAIL POSITION",
           "PLC", "REMARK"]
HDR_IL = ["IL NO", "OUTPUT TAG", "ACTION", "KIND", "LOGIC", "GROUP",
          "CONDITION", "RESET", "BYPASS", "PRIORITY", "PLC BLOCK",
          "DWG No.", "SHEET", "REMARK"]

NOTE = ("데모용 가상 데이터입니다. 실제 프로젝트 자료가 아닙니다. "
        "CONDITION 열은 자연어로 작성된 실제 서류 형태를 재현한 것입니다.")

THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEAD_FILL = PatternFill("solid", fgColor="D9E2EC")
ILK_FILL = PatternFill("solid", fgColor="FDE2E2")


def style_header(ws, ncol, row):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True, size=9)
        cell.fill = HEAD_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)


def write_outputs(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "OUTPUT LIST"
    ws["A1"] = NOTE
    ws["A1"].font = Font(size=9, italic=True, color="C00000")
    ws.append([])
    ws.append(HDR_OUT)
    style_header(ws, len(HDR_OUT), 3)
    for i, (tag, svc, sysm, typ, fail, drive) in enumerate(OUTPUTS, 1):
        ws.append([i, tag, svc, sysm, typ, fail, drive, "PLC-UPW-01", ""])
    for col, w in zip("ABCDEFGHI", (5, 12, 34, 16, 14, 14, 11, 13, 20)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"
    wb.save(path)
    return len(OUTPUTS)


def write_interlocks(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "INTERLOCK LIST"
    ws["A1"] = NOTE
    ws["A1"].font = Font(size=9, italic=True, color="C00000")
    ws.append([])
    ws.append(HDR_IL)
    style_header(ws, len(HDR_IL), 3)

    # 출력 태그 → 고장 위치. REMARK 에 "* Valve Action : ..." 로 붙인다.
    # 실물 인터락 리스트가 그 자리에 적기 때문이다 — 별도 열을 만들면
    # 실물과 모양이 달라지고, 파서도 두 벌이 된다.
    fail_of = {o[0]: o[4] for o in OUTPUTS}

    rowcount = 0
    for (ilno, tag, action, kind, logic, reset, byp, prio,
         block, dwg, sheet, conds, remark) in IL:
        fail = fail_of.get(tag)
        if fail:
            va = "* Valve Action : %s" % fail.title()
            remark = ("%s %s" % (remark, va)).strip() if remark else va
        first = True
        for grp, text in conds:
            # 실제 서류처럼 첫 행에만 항목 정보를 쓰고 나머지는 공란으로 둔다.
            ws.append([
                ilno if first else None,
                tag if first else None,
                action if first else None,
                kind if first else None,
                logic if first else None,
                grp,
                text,
                reset if first else None,
                byp if first else None,
                prio if first else None,
                block if first else None,
                dwg if first else None,
                sheet if first else None,
                remark if first else None,
            ])
            rowcount += 1
            r = ws.max_row
            for c in range(1, len(HDR_IL) + 1):
                ws.cell(row=r, column=c).border = BORDER
                ws.cell(row=r, column=c).font = Font(size=9)
                ws.cell(row=r, column=c).alignment = Alignment(
                    vertical="center", wrap_text=(c == 7))
            if kind == "INTERLOCK":
                ws.cell(row=r, column=4).fill = ILK_FILL
            first = False

    for col, w in zip(["A", "B", "C", "D", "E", "F", "G", "H", "I", "J",
                       "K", "L", "M", "N"],
                      (13, 12, 8, 12, 7, 7, 46, 9, 8, 8, 18, 24, 7, 34)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"
    wb.save(path)
    return len(IL), rowcount


def main():
    ap = argparse.ArgumentParser(description="합성 인터락 리스트 생성")
    # 기본 출력은 demo/ 다. data/ 는 사용자가 넣는 자료만 두는 자리라
    # 생성기가 그리로 쓰면 안 된다.
    ap.add_argument("--out", default=os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "demo"))
    args = ap.parse_args()
    os.makedirs(args.out, exist_ok=True)

    p1 = os.path.join(args.out, "DEMO_OUTPUT_LIST.xlsx")
    n = write_outputs(p1)
    print("출력 태그 %d건 → %s" % (n, p1))

    p2 = os.path.join(args.out, "DEMO_INTERLOCK_LIST.xlsx")
    n_il, n_row = write_interlocks(p2)
    print("인터락 항목 %d건 / 조건 행 %d건 → %s" % (n_il, n_row, p2))


if __name__ == "__main__":
    main()

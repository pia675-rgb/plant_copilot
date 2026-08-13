#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
make_tb_list.py — 데모 TB 결선표 생성 (mastertool 격자 양식)

TB List 는 사용자가 넣는 자료다. 데모에는 그 자리를 채울 파일이 없어서
단자(TERMINAL) 정보가 부속 데이터에만 있었다. 부속 데이터는 프로젝트
산출물이 아니므로, 데모에서도 **실물과 같은 경로**로 단자가 오게 한다.

IO List 의 (판넬·카드·채널)과 계기 리스트를 읽어 mastertool 이 만드는
격자 배치 결선표와 같은 모양으로 쓴다.

    ■ PANEL : CUB-B
    DI32                 AI4-4W
    TB 1-1    D01-S03    TB 2-1    D01-S08
    N011  I4.0  AIT-1001 IW528+ IW528  TIT-4004
    ...

단자 번호를 **여기서 새로 만들지 않는다.** IO List 에 이미 있는 값을
쓰고, 없으면 그 채널은 비운다 — 지어낸 단자 번호로 도면을 하이라이트하면
없는 자리를 가리키게 된다.

    python -m tools.make_tb_list

의존성: openpyxl
"""

import argparse
import os
import re
import sys
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import config  # noqa: E402

BLOCK_WIDTH = 5              # 블록 하나가 차지하는 열 수 (리더와 같은 값)
GAP = 1                      # 블록 사이 빈 열
THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEAD_FILL = PatternFill("solid", fgColor="D9E2EC")

NOTE = ("데모용 가상 데이터입니다. 실제 프로젝트 자료가 아닙니다. "
        "mastertool TB 결선표 양식을 재현한 것입니다.")


def _s(v):
    return str(v).strip() if v is not None else ""


def _card_kind(io_type, points):
    """IO TYPE + 채널 수 → 블록 머리 표기 (DI32 · AI4-4W 형태)."""
    t = (io_type or "").upper()
    grp = next((g for g in ("AI", "AO", "DI", "DO") if g in t), "AI")
    size = 32 if grp in ("DI", "DO") else (16 if points > 8 else
                                           (8 if points > 4 else 4))
    return "%s%d" % (grp, size) if grp in ("DI", "DO") else \
        "%s%d-4W" % (grp, size)


def collect(points):
    """
    TAG → 배선 을 (판넬 → 카드 → 채널) 로 묶는다.

    배선이 없는 점은 결선표에 실을 자리가 없으므로 뺀다.
    """
    by_panel = defaultdict(lambda: defaultdict(list))
    for tag, r in points.items():
        panel = _s(r.get("PANEL"))
        slot = _s(r.get("SLOT"))
        if not panel or not slot:
            continue
        key = (_s(r.get("PN(DP)")), _s(r.get("RACK")), slot)
        by_panel[panel][key].append({
            "tag": tag,
            "ch": _s(r.get("CH")),
            "terminal": _s(r.get("TERMINAL")),
            "io_type": _s(r.get("IO TYPE")) or _s(r.get("SIGNAL")),
        })
    for panel in by_panel:
        for key in by_panel[panel]:
            by_panel[panel][key].sort(
                key=lambda x: int(x["ch"]) if x["ch"].isdigit() else 0)
    return by_panel


def build(by_panel, out_path):
    wb = Workbook()
    first = True
    for panel in sorted(by_panel):
        title = re.sub(r"[\\/*?:\[\]]", "_", panel)[:31]
        ws = wb.active if first else wb.create_sheet(title)
        if first:
            ws.title = title
            first = False
        ws.cell(1, 1, NOTE).font = Font(size=9, italic=True, color="C00000")
        ws.cell(1, 2, "■ PANEL : %s" % panel).font = Font(bold=True, size=10)

        col = 2
        tb_no = defaultdict(int)     # 카드 종류별 TB 번호
        for n, (key, chans) in enumerate(sorted(by_panel[panel].items()), 1):
            pn, rack, slot = key
            kind = _card_kind(chans[0]["io_type"], len(chans))
            grp = kind[:2]
            tb_no[grp] += 1
            tb = "TB %d-%d" % ({"DI": 1, "AI": 2, "DO": 3, "AO": 4}.get(grp, 5),
                               tb_no[grp])
            ref = "D%02d-S%02d" % (int(rack or 0), int(slot or 0))

            c = ws.cell(2, col, kind)
            c.font = Font(bold=True, size=9)
            c.fill = HEAD_FILL
            ws.cell(3, col, tb).font = Font(size=9)
            ws.cell(3, col + 2, ref).font = Font(size=9, color="808080")

            for i, ch in enumerate(chans):
                r = 4 + i
                if grp in ("DI", "DO"):
                    # [단자블록] [주소] [ ] [태그]
                    ws.cell(r, col, "N%03d" % (int(slot or 0) * 10 + 1))
                    ws.cell(r, col + 1, "%s%s.%s"
                            % ("I" if grp == "DI" else "Q", slot,
                               ch["ch"] or "0"))
                    ws.cell(r, col + 3, ch["tag"])
                else:
                    # [ ] [단자 +] [주소] [태그]  — 아날로그는 +/- 두 행
                    base = "IW%d" % (500 + int(slot or 0) * 10
                                     + (int(ch["ch"]) if ch["ch"].isdigit()
                                        else 0) * 2)
                    ws.cell(r, col + 1, base + "+")
                    ws.cell(r, col + 2, base)
                    ws.cell(r, col + 3, ch["tag"])
                # 단자 번호는 IO List 에 있는 값을 그대로 쓴다. 없으면 비운다.
                if ch["terminal"]:
                    ws.cell(r, col + 4, ch["terminal"])
                for cc in range(col, col + BLOCK_WIDTH):
                    ws.cell(r, cc).font = Font(size=8)
                    ws.cell(r, cc).border = BORDER
                    ws.cell(r, cc).alignment = Alignment(vertical="center")

            for cc in range(col, col + BLOCK_WIDTH):
                ws.column_dimensions[get_column_letter(cc)].width = \
                    14 if cc == col + 3 else 10
                ws.cell(2, cc).border = BORDER
                ws.cell(3, cc).border = BORDER
            col += BLOCK_WIDTH + GAP
        ws.freeze_panes = ws.cell(4, 2)
    wb.save(out_path)


def main():
    ap = argparse.ArgumentParser(description="데모 TB 결선표 생성")
    ap.add_argument("--out", default=None,
                    help="기본: data/TB_LIST.xlsx")
    a = ap.parse_args()

    from ingest.lists import load_points
    # 부속 데이터를 함께 읽는다. 여기서 단자 번호를 만들지 않고 옮기기만
    # 하므로, 원천이 부속이든 실물이든 결과는 같다.
    points = load_points(config.IO_LIST, getattr(config, "INSTRUMENT_SPECS", None) or getattr(config, "INSTRUMENT_SPEC", None),
                         None, None)
    by_panel = collect(points)
    out = a.out or config.TB_LIST
    build(by_panel, out)

    n_ch = sum(len(v) for p in by_panel.values() for v in p.values())
    n_tb = sum(len(p) for p in by_panel.values())
    n_term = sum(1 for p in by_panel.values() for v in p.values()
                 for x in v if x["terminal"])
    print("판넬 %d · TB 블록 %d · 채널 %d (단자 있는 채널 %d)"
          % (len(by_panel), n_tb, n_ch, n_term))
    print("→", out)
    if n_term < n_ch:
        print("※ 단자 번호가 없는 채널 %d점은 비워 두었습니다 — 지어내지 "
              "않습니다." % (n_ch - n_term))


if __name__ == "__main__":
    main()

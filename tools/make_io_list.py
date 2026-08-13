#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
make_io_list.py — 표준 IO List + 계기 리스트 생성 (두 문서)

표준 컬럼 24종은 mastertool(데스크톱)과 MAXIS(MCP 서버)가 이미 쓰는
체계다. 여기서 따로 정의하지 않고 **그 표기를 그대로 따른다.**

주의 — 5번 열은 **PN(DP)** 다. "DP(PN)" 은 구 표기이고 io_tools_core.py 의
STD_ALIASES 에 하위 호환으로만 남아 있다. 구 표기로 헤더를 쓰면 완전일치
매핑에서 어긋난다.

── 문서를 둘로 나눈 이유 ─────────────────────────────────────

**IO List 에는 표준 24종만 넣는다.** 한때 copilot 이 쓰는 계기 정보를
25번 열부터 붙여 뒀는데, 실제 프로젝트의 IO List 에는 그런 열이 없다.
제조사·모델·레인지·고장모드·알람 설정 같은 것은 **계기 리스트(Instrument
List)** 에 있는 정보다. 두 문서를 한 파일에 섞으면 실물과 모양이 달라져
그대로 갈아 끼울 수 없다.

    IO_LIST.xlsx          배선 — mastertool/MAXIS 표준 24종, 그것만
    INSTRUMENT_LIST.xlsx  사양 — 제조사·레인지·고장모드·알람·매뉴얼,
                                밸브·펌프의 고장 위치·구동 방식 포함

두 문서는 **TAG 로 조인**한다. 실제 프로젝트에서 사람이 하는 것과 같다.
조인은 ingest/lists.py 한 곳에서만 한다.

── 출력도 IO 점이다 ──────────────────────────────────────────

한때 출력(밸브·펌프)을 DEMO_OUTPUT_LIST.xlsx 라는 별도 문서로 두었다.
실제 프로젝트에 그런 문서는 없다. 출력은 **IO List 의 DO/AO 점**이고,
고장 위치·구동 방식은 계기 리스트에 있다.

그래서 출력 24점을 IO List 에 DO/AO 로 싣는다. 다만 랙·슬롯·채널을
지어낼 수 없으므로 **배선 열은 비운다.** 그러면 "IO List 에 있으나 배선
미정" 상태가 되고, 실물에서 채워야 할 칸이 정확히 드러난다.

── 채우지 않는 열 ────────────────────────────────────────────

ADD·PRG·PG·BIT·BYTE·SIGNAL TYPE1/2·POWER SOURCE·INST. PANEL·P&ID TAG 는
데모 원천에 없다. **비워 둔다.** 그럴듯한 값을 지어 넣으면 나중에 실물과
구분되지 않는다. MAXIS 검증에서 '필수값 누락' 으로 잡히는 것이 정상이다.

    python tools/make_io_list.py

의존성: openpyxl
"""

import argparse
import csv
import os
import sys
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import config  # noqa: E402

# ── mastertool / MAXIS 표준 24종 (io_tools_core.STANDARD_ORDER) ──
STANDARD_ORDER = [
    "INDEX", "PLC", "PANEL", "LOCATION", "PN(DP)", "RACK", "SLOT", "CH",
    "UNIT", "IO TYPE", "ADD", "TAG", "DESCRIPTION",
    "SIGNAL TYPE1", "SIGNAL TYPE2", "POWER SOURCE", "INST. PANEL",
    "PRG", "PG", "BIT", "BYTE", "DWG No.", "P&ID TAG", "REMARK",
]

# ── 계기 리스트 (실물 양식) ─────────────────────────────────
# 2단 머리(그룹 → 하위 항목). 배선(랙·슬롯·채널)은 여기 없다 —
# IO List 에서 받아간다.
#
#   (그룹명, [하위 항목])   하위가 비면 단일 열(두 행 병합)
INSTRUMENT_HEAD = [
    ("TAG NO.", []), ("DESCRIPTION", []), ("Q'TY", []), ("SENSOR TYPE", []),
    ("MATERIAL", ["ELEMENT", "BODY"]),
    ("SCALE RANGE", ["MIN", "MAX", "UNIT"]),
    ("DISPLAY", []), ("POWER", []), ("OUTPUT SIGNAL", []),
    ("CONNECTION", ["TYPE", "SIZE", "MATERIAL"]),
    ("LINE", ["SIZE", "MATERIAL"]),
    ("FLUID", ["NAME", "CONDITION"]),
    ("MODEL", []), ("MAKER", []), ("LOCATION", []), ("REMARKS", []),
]

# 실물 양식의 열 → 데모 원천의 열. 없는 것은 비운다.
INSTRUMENT_FILL = {
    "TAG NO.": "TAG",
    "DESCRIPTION": "SERVICE",
    "SENSOR TYPE": "MEAS TYPE",
    "SCALE RANGE|MIN": "RANGE MIN",
    "SCALE RANGE|MAX": "RANGE MAX",
    "SCALE RANGE|UNIT": "UNIT",
    "OUTPUT SIGNAL": "SIGNAL",
    "MODEL": "MODEL",
    "MAKER": "MAKER",
    "REMARKS": "REMARK",
}

# ── copilot 부속 데이터 (프로젝트 산출물 아님) ───────────────
# 실물 계기 리스트에도 IO List 에도 없는데 copilot 이 쓰는 항목이다.
# 실제 프로젝트에서는 각각 다른 문서에 있다(계기 사양서·알람 설정
# 리스트·TB 리스트·문서 관리대장). 어디서 받아올지 정해지기 전까지
# 별도 파일로 분리해 둔다 — 실물 양식에 섞어 넣으면 그 양식이
# 실물과 달라진다.
ATTR_NOTICE = ("copilot 부속 데이터입니다 — 프로젝트 산출물이 아닙니다. "
               "실물 계기 리스트·IO List 에 없는 항목만 모았습니다. "
               "실제로 어느 문서에서 받아와야 하는지는 SOURCE 시트를 "
               "보십시오.")

# FAULT MODE·ALARM L/H 는 뺐다.
#   FAULT MODE  실물 Instrument List 에 없다. 관행으로 가정하면 리스트에
#               없는 값으로 인터락 성립을 판정하게 되므로 기능째 제거했다.
#   ALARM L/H   알람 설정값은 업로드 대상 문서에 없다. SCALE RANGE 만 쓴다.
# TERMINAL 은 TB List 가 원본이다. 다만 TB List 가 없거나 태그가 겹치지
# 않으면 단자 조회가 통째로 비므로, 데모 값은 여기 남겨 두고 TB List 가
# 있을 때 그쪽이 덮어쓰게 한다.
# 부속 데이터(TAG_ATTRIBUTES.xlsx)는 없앴다.
#
# 네 항목이 모두 업로드 자료에서 오게 됐다.
#   TERMINAL       TB List
#   TYPE           계기 리스트 SENSOR TYPE
#   FAIL POSITION  인터락 리스트 "* Valve Action"
#   MANUAL FILE    MODEL ↔ 매뉴얼 파일명 대조 (ingest/manual_match.py)
#
# 내가 만든 열이 하나라도 남아 있으면 데모에서는 채워지고 실물에서는
# 비는 항목이 생긴다. 그 차이는 화면에 드러나지 않는다.

WIDTH = {"INDEX": 6, "PLC": 13, "PANEL": 9, "LOCATION": 26, "PN(DP)": 8,
         "RACK": 6, "SLOT": 6, "CH": 5, "UNIT": 9, "IO TYPE": 9, "ADD": 10,
         "TAG": 12, "DESCRIPTION": 34, "DWG No.": 14, "P&ID TAG": 12,
         "REMARK": 34, "SIGNAL": 13,
         "TERMINAL": 10, "MAKER": 17, "MODEL": 10,
         "MEAS TYPE": 14, "MANUAL FILE": 22, "POINT TYPE": 11}

# ── 스테이션·랙 번호 (가정) ─────────────────────────────────
# 원천의 RACK 은 전 행 0 이라 카드를 구분하지 못한다. 표준 양식에는
# PN(DP)(PROFINET 스테이션) 열이 있으므로 중앙과 원격을 나눠 적는다.
#   중앙 큐비클  PN(DP) 공란, RACK 으로 구분
#   원격 스테이션 PN(DP) 부여, RACK 0
# **확정값이 아니다.** 실제 번호로 이 표만 고치면 아래가 전부 따라온다.
STATION_MAP = {
    "CUB-A": {"pn": "", "rack": 0},
    "CUB-B": {"pn": "", "rack": 1},
    "CUB-C": {"pn": "", "rack": 2},
    "CUB-D": {"pn": "", "rack": 3},      # 출력(DO/AO) 전용 큐비클
    "RIO-01": {"pn": 11, "rack": 0},
    "RIO-02": {"pn": 12, "rack": 0},
}
STATION_NOTE = ("PN(DP)·RACK 은 원천의 RACK 이 전 행 0 이라 부여한 "
                "가정값입니다 (중앙 큐비클은 RACK 으로 구분, 원격 "
                "스테이션은 PN(DP) 부여). 실물 IO List 로 교체 시 실제 "
                "값을 사용하십시오.")

OUT_COLUMNS = ["INDEX", "TAG", "DESCRIPTION", "POINT TYPE", "TYPE",
               "FAIL POSITION", "PLC", "REMARK"]

NOTICE = ("데모용 가상 데이터입니다. 실제 프로젝트 자료가 아닙니다. "
          "컬럼은 mastertool/MAXIS 표준 24종(STANDARD_ORDER)뿐입니다. "
          "계기 사양은 INSTRUMENT_LIST.xlsx 에 있고 TAG 로 조인합니다. "
          "노란색 DWG No. 열은 도면 매핑으로 채워지는 칸입니다.")
INST_NOTICE = ("데모용 가상 데이터입니다. 실제 프로젝트 자료가 아닙니다. "
               "계기 사양 문서입니다 — 배선(랙·슬롯·채널·판넬)은 "
               "IO_LIST.xlsx 에 있고 TAG 로 조인합니다.")
BLANK_NOTE = ("ADD·PRG·PG·BIT·BYTE·SIGNAL TYPE1/2·POWER SOURCE·"
              "INST. PANEL·P&ID TAG 는 데모 원천에 없어 비워 둡니다 — "
              "값을 지어내지 않습니다.")

HEAD_ROW = 3
YELLOW = PatternFill("solid", fgColor="FFF2CC")
GREY = PatternFill("solid", fgColor="EDEDED")
BLUE = PatternFill("solid", fgColor="DDEBF7")


def _s(v):
    return str(v).strip() if v is not None else ""


def read_sheet(path):
    ws = load_workbook(path, read_only=True, data_only=True).active
    rows = list(ws.iter_rows(values_only=True))
    hi = next(i for i, r in enumerate(rows)
              if r and "TAG" in [_s(c).upper() for c in r])
    hdr = [_s(c).upper() for c in rows[hi]]
    idx = {h: i for i, h in enumerate(hdr) if h}
    out = []
    for r in rows[hi + 1:]:
        if not r or not r[idx["TAG"]]:
            continue
        out.append({h: (r[i] if i < len(r) else None) for h, i in idx.items()})
    return out


def read_locations():
    """배치도가 만든 판넬 위치. 있으면 LOCATION 열에 넣는다."""
    p = getattr(config, "PANEL_LOCATIONS", "")
    if not p or not os.path.exists(p):
        return {}
    out = {}
    with open(p, newline="", encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            name = (r.get("PANEL") or "").strip()
            if name:
                out[name] = "%s / %s" % (r.get("AREA") or "",
                                         r.get("GRID") or "")
    return out


def output_io_type(sig):
    """
    출력의 IO 종류.

    예전에는 DRIVE 열(SOLENOID / MCC / AO-POSITIONER)에서 읽었는데,
    그 열은 어느 실물 문서에도 없다. 조절 신호(4-20mA)면 AO, 아니면 DO 로
    본다. 판정 근거가 없으면 비운다.
    """
    t = (sig or "").upper()
    if "4-20" in t or "AO" in t or "POSITION" in t:
        return "AO"
    return "DO" if t else "DO"


def io_group(signal):
    """SIGNAL 표기 → IO TYPE 그룹. 채널 수는 원천에 없으므로 붙이지 않는다."""
    t = (signal or "").strip().upper()
    for g in ("AI", "AO", "DI", "DO"):
        if t.endswith(" " + g) or t.startswith(g + " ") or (" %s " % g) in t:
            return g
    return ""


def build_io_list(inst_rows, out_rows, locs, out_path):
    """IO List — 표준 24종만. 확장 열을 붙이지 않는다."""
    wb = Workbook()
    ws = wb.active
    ws.title = "IO LIST"
    ws.cell(1, 1, NOTICE).font = Font(bold=True, color="C00000")
    ws.cell(2, 1, STATION_NOTE + " " + BLANK_NOTE).font = Font(
        italic=True, size=9, color="808080")

    for c, name in enumerate(STANDARD_ORDER, start=1):
        cell = ws.cell(HEAD_ROW, c, name)
        cell.font = Font(bold=True)
        cell.fill = YELLOW if name == "DWG No." else GREY
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(c)].width = WIDTH.get(name, 12)

    unknown = set()
    for i, r in enumerate(inst_rows, start=1):
        panel = _s(r.get("PANEL"))
        st = STATION_MAP.get(panel)
        if panel and st is None:
            unknown.add(panel)
        vals = {
            "INDEX": i,
            "PLC": _s(r.get("PLC")),
            "PANEL": panel,
            "LOCATION": locs.get(panel, ""),
            "PN(DP)": st["pn"] if st else "",
            "RACK": st["rack"] if st else r.get("RACK"),
            "SLOT": r.get("SLOT"),
            "CH": r.get("CH"),
            "UNIT": _s(r.get("UNIT")),
            "IO TYPE": io_group(_s(r.get("SIGNAL"))),
            "TAG": _s(r.get("TAG")),
            "DESCRIPTION": _s(r.get("SERVICE")),
            "DWG No.": _s(r.get("DWG NO.")) or _s(r.get("DWG NO")),
            "REMARK": _s(r.get("REMARK")),
        }
        for c, name in enumerate(STANDARD_ORDER, start=1):
            cell = ws.cell(HEAD_ROW + i, c, vals.get(name))
            if name == "DWG No.":
                cell.fill = YELLOW
    # 출력 — DO/AO 점.
    #
    # 원천에 배선(PANEL·RACK·SLOT·CH)이 적혀 있으면 그대로 싣고, 없으면
    # 비운다. 지어내지 않는다. 예전에는 무조건 비웠는데, 그러면 엑셀에서
    # 손으로 채운 배선이 다음 재생성에서 조용히 사라진다. 실제로 CUB-D
    # 24점을 그렇게 잃을 뻔했다.
    base = len(inst_rows)
    for j, r in enumerate(out_rows, start=1):
        panel = _s(r.get("PANEL"))
        st = STATION_MAP.get(panel) if panel else None
        if panel and st is None:
            unknown.add(panel)
        vals = {
            "INDEX": base + j,
            "PLC": _s(r.get("PLC")),
            "PANEL": panel,
            "LOCATION": locs.get(panel, ""),
            "PN(DP)": (st["pn"] if st else _s(r.get("PN(DP)"))),
            "RACK": (st["rack"] if st else r.get("RACK")),
            "SLOT": r.get("SLOT"),
            "CH": r.get("CH"),
            "IO TYPE": output_io_type(_s(r.get("SIGNAL")) or _s(r.get("TYPE"))),
            "TAG": _s(r.get("TAG")),
            "DESCRIPTION": _s(r.get("SERVICE")),
            "REMARK": (_s(r.get("REMARK")) if panel
                       else "배선 미정 — 랙·슬롯·채널 확인 필요"),
        }
        for c, name in enumerate(STANDARD_ORDER, start=1):
            cell = ws.cell(HEAD_ROW + base + j, c, vals.get(name))
            if name == "DWG No.":
                cell.fill = YELLOW
    ws.freeze_panes = ws.cell(HEAD_ROW + 1, 4)

    _provenance(wb, len(inst_rows) + len(out_rows), "IO LIST",
                " | ".join(STANDARD_ORDER))
    wb.save(out_path)
    return unknown


def _flat_head():
    """(그룹, 하위) → 평탄한 열 이름 목록. 하위가 있으면 '그룹|하위'."""
    out = []
    for grp, subs in INSTRUMENT_HEAD:
        if subs:
            out += ["%s|%s" % (grp, sc) for sc in subs]
        else:
            out.append(grp)
    return out


def build_instrument_list(inst_rows, out_rows, out_path):
    """
    계기 리스트 — 실물 양식(2단 머리).

    배선(랙·슬롯·채널)은 여기 없다. IO List 에서 TAG 로 받아간다.
    원천에 없는 열(재질·접속·배관·유체·전원·표시)은 비운다.
    """
    from openpyxl.utils import get_column_letter as _col

    wb = Workbook()
    ws = wb.active
    ws.title = "INSTRUMENT LIST"
    ws.cell(1, 1, INST_NOTICE).font = Font(bold=True, color="C00000")

    r1, r2 = HEAD_ROW, HEAD_ROW + 1
    c = 1
    for grp, subs in INSTRUMENT_HEAD:
        if subs:
            ws.merge_cells(start_row=r1, start_column=c,
                           end_row=r1, end_column=c + len(subs) - 1)
            cell = ws.cell(r1, c, grp)
            cell.font = Font(bold=True)
            cell.fill = GREY
            cell.alignment = Alignment(horizontal="center",
                                       vertical="center")
            for k, sc in enumerate(subs):
                sub = ws.cell(r2, c + k, sc)
                sub.font = Font(bold=True, size=9)
                sub.fill = GREY
                sub.alignment = Alignment(horizontal="center")
                ws.column_dimensions[_col(c + k)].width = WIDTH.get(sc, 11)
            c += len(subs)
        else:
            ws.merge_cells(start_row=r1, start_column=c,
                           end_row=r2, end_column=c)
            cell = ws.cell(r1, c, grp)
            cell.font = Font(bold=True)
            cell.fill = GREY
            cell.alignment = Alignment(horizontal="center",
                                       vertical="center")
            ws.column_dimensions[_col(c)].width = WIDTH.get(grp, 13)
            c += 1

    flat = _flat_head()
    # 밸브·펌프의 종류(ON-OFF VALVE · PUMP)는 계기의 SENSOR TYPE 자리에
    # 들어간다. 출력 원천은 그 값을 TYPE 열에 갖고 있으므로 옮겨 준다.
    merged = []
    for r in list(inst_rows) + list(out_rows):
        r = dict(r)
        if not r.get("MEAS TYPE") and r.get("TYPE"):
            r["MEAS TYPE"] = r["TYPE"]
        merged.append(r)

    for i, r in enumerate(merged, start=1):
        for k, name in enumerate(flat, start=1):
            src = INSTRUMENT_FILL.get(name)
            v = r.get(src) if src else None
            ws.cell(r2 + i, k, _s(v) if isinstance(v, str) else v)
    ws.freeze_panes = ws.cell(r2 + 1, 3)

    _provenance(wb, len(merged), "INSTRUMENT LIST",
                " | ".join(flat))
    wb.save(out_path)


def _provenance(wb, rows, kind, columns):
    """어느 원천에서 몇 행을 읽어 만들었는지. selfcheck 가 이걸 대조한다."""
    ws = wb.create_sheet("PROVENANCE")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 78
    src_in = os.path.join(config.SOURCE_DIR, "DEMO_INSTRUMENT_LIST.xlsx")
    stamp = [
        ("GENERATED", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("GENERATOR", "tools/make_io_list.py"),
        ("DOCUMENT", kind),
        ("COLUMNS", columns),
        ("SOURCE INPUT", os.path.basename(src_in)),
        ("SOURCE ROWS", rows),
    ]
    if kind == "IO LIST":
        stamp += [("STANDARD",
                   "mastertool / MAXIS STANDARD_ORDER 24종 — 확장 열 없음"),
                  ("STATION RULE", STATION_NOTE),
                  ("BLANK COLUMNS", BLANK_NOTE)]
    for i, (k, v) in enumerate(stamp, start=1):
        ws.cell(i, 1, k).font = Font(bold=True)
        ws.cell(i, 2, v)


def main():
    ap = argparse.ArgumentParser(description="표준 IO List + 계기 리스트 생성")
    here = os.path.dirname(os.path.abspath(__file__))
    ap.add_argument("--out-io", default=None)
    ap.add_argument("--out-inst", default=None)
    ap.add_argument("--src-out", default=None,
                    help="흡수할 출력 원천 (기본: v1 demo_data)")
    a = ap.parse_args()

    # 생성물은 정해진 자리로 나간다.
    #   IO List·계기 리스트  → data/   (사용자가 넣는 자료와 같은 자리)
    #   부속 데이터           → derived/ (도구가 만드는 것)
    out_io = a.out_io or config.IO_LIST
    out_inst = a.out_inst or config.INSTRUMENT_SPEC

    src_in = os.path.join(config.SOURCE_DIR, "DEMO_INSTRUMENT_LIST.xlsx")
    print("원천 :", src_in)
    inst = read_sheet(src_in)
    # 원천은 v1 demo_data 에 둔다. v2 data/ 에는 생성물만 남긴다 —
    # 원천 사본을 v2 에 두면 원천이 바뀌었을 때 조용히 어긋난다.
    src_out = a.src_out or os.path.join(config.SOURCE_DIR,
                                        "DEMO_OUTPUT_LIST.xlsx")
    outs = read_sheet(src_out) if os.path.isfile(src_out) else []
    print("출력 원천 :", src_out)
    if not outs:
        print("[경고] 출력 원천이 없어 DO/AO 점 없이 만듭니다 — 인터락 "
              "출력 태그가 IO List 에서 통째로 빕니다.")
    locs = read_locations()
    if not locs:
        print("[안내] PANEL_LOCATIONS.csv 가 없어 LOCATION 을 비웁니다.")

    unknown = build_io_list(inst, outs, locs, out_io)
    build_instrument_list(inst, outs, out_inst)
    if unknown:
        print("[경고] STATION_MAP 에 없는 판넬: %s" % ", ".join(sorted(unknown)))
    print("IO List      입력 %d + 출력 %d = %d점 → %s"
          % (len(inst), len(outs), len(inst) + len(outs), out_io))
    print("계기 리스트   %d점 → %s" % (len(inst) + len(outs), out_inst))
    if outs:
        nowire = sum(1 for r in outs if not _s(r.get("PANEL")))
        if nowire:
            print("※ 출력 %d점 중 %d점은 배선이 비어 있습니다 — 실물에서 "
                  "채워야 할 칸입니다." % (len(outs), nowire))
        else:
            print("※ 출력 %d점 모두 배선이 채워져 있습니다." % len(outs))
    print("※ PN(DP)·RACK 은 가정값입니다. STATION_MAP 을 실제 값으로 "
          "고치십시오.")


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
make_maintenance_history.py — 데모용 가상 보수 이력 생성

Plant Maintenance Copilot 기획서의 차별화 항목
"매뉴얼은 A라 하는데 우리 현장 이력은 B였다" 를 성립시키는 데이터.

- error_codes.json 의 실제 코드에 걸어서 생성 (매뉴얼과 조인 가능)
- DEMO_INSTRUMENT_LIST 의 실제 태그에 배정
- 매뉴얼 일치도를 일치 / 부분일치 / 불일치 세 등급으로 부여

주의: 전량 합성 데이터입니다. 실제 보수 이력이 아닙니다.
매뉴얼 원문은 복사하지 않고 code_ref 로 참조만 겁니다 (원본은 error_codes.json).

사용:
    python make_maintenance_history.py --src demo_data --out demo_data

의존성: openpyxl
"""

import argparse
import datetime as dt
import json
import os
import random

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

MATCH_FULL = "일치"
MATCH_PART = "부분일치"
MATCH_NONE = "불일치"

# ─────────────────────────────────────────────────────────────
# 시나리오 — (기기, 코드ref, 증상, 초동조치, 실제원인, 조치, 일치도, 소요분, 부품)
# 불일치·부분일치 건이 이 데이터셋의 존재 이유다.
# ─────────────────────────────────────────────────────────────
SCENARIOS = [
    # ── Sievers M9e ────────────────────────────────────────
    ("M9e", "M9E-10084", "시료 유량 미검출, 분석 정지",
     "매뉴얼대로 시료 라인 밸브 개도 확인 (이상 없음)",
     "시료 라인 인라인 필터 막힘",
     "인라인 필터 교체 후 라인 플러싱, 분석 재개",
     MATCH_PART, 75, "인라인 필터 0.45um"),

    ("M9e", "M9E-10084", "시료 유량 미검출, 분석 정지",
     "필터 교체했으나 동일 알람 재발",
     "유량센서 자체 오염 — 시료측 정상이었음",
     "유량센서 분해 세정 및 재조립",
     MATCH_NONE, 180, "-"),

    ("M9e", "M9E-500", "UV 램프 잔여 수명 15일 미만 경고",
     "매뉴얼대로 신규 램프 발주",
     "램프 수명 도래 (정상 소모)",
     "정기 PM 시 램프 교체, 카운터 리셋",
     MATCH_FULL, 45, "UV 램프"),

    ("M9e", "M9E-300", "산 잔량 10% 미만 경고",
     "매뉴얼대로 산 카트리지 교체 준비",
     "산 잔량 정상 — 산 라인 에어락으로 소모량 오산정",
     "산 라인 배기 후 잔량 재산정, 카트리지 유지",
     MATCH_NONE, 60, "-"),

    ("M9e", "M9E-300", "산 잔량 10% 미만 경고",
     "잔량 확인 후 카트리지 발주",
     "산 카트리지 소진 (정상 소모)",
     "카트리지 교체 및 프라이밍",
     MATCH_FULL, 40, "산 카트리지"),

    ("M9e", "M9E-3101", "IC 채널 DI 유량 제한 감지",
     "매뉴얼대로 진단 모드 수행",
     "DI 카트리지 수명 만료로 차압 상승",
     "DI 카트리지 교체 후 유량 회복 확인",
     MATCH_PART, 90, "DI 카트리지"),

    ("M9e", "M9E-2403", "산 시린지 기포 감지",
     "매뉴얼대로 산 레벨 확인 및 라인 플러시 (재발)",
     "시린지 플런저 씰 경화로 미세 누기",
     "시린지 어셈블리 교체",
     MATCH_NONE, 150, "산 시린지 ASSY"),

    ("M9e", "M9E-800", "시료 전도도 아날로그 판독값 범위 초과",
     "센서 교정 상태 점검",
     "계기 정상 — 상류 열교환기 트러블로 시료 온도 급변",
     "공정팀 이관, 열교환기 조치 후 자연 해소",
     MATCH_NONE, 30, "-"),

    ("M9e", "M9E-7200", "TC 시료 유로 공기 혼입",
     "매뉴얼대로 시료 충분량 및 프로토콜 확인",
     "시료 라인 배기 불충분",
     "라인 배기 재실시 후 정상화",
     MATCH_FULL, 35, "-"),

    ("M9e", "M9E-602", "TC 시료 펌프 튜브 잔여 수명 15일 미만",
     "매뉴얼대로 신규 튜브 발주",
     "튜브 수명 도래 (정상 소모)",
     "PM 시 튜브 교체",
     MATCH_FULL, 50, "펌프 튜브 세트"),

    ("M9e", "M9E-10047", "8시간 이상 산화제 미사용 알림",
     "장비 상태 확인",
     "대기 모드 정상 동작 — 조치 불요",
     "오경보로 판정, 알람 임계 재검토 요청",
     MATCH_PART, 15, "-"),

    ("M9e", "M9E-5503", "시료 밸브 전류 저하",
     "매뉴얼대로 밸브 진단 수행",
     "밸브 정상 — 커넥터 접촉 불량",
     "커넥터 재체결 및 접점 세정",
     MATCH_NONE, 55, "-"),

    ("M9e", "M9E-3100", "DI 저수조 레벨 저하",
     "매뉴얼대로 DI 저수조 보충",
     "저수조 소모 (정상)",
     "DI수 보충 후 정상화",
     MATCH_FULL, 25, "-"),

    ("M9e", "M9E-5700", "산화제 시린지 위치 이상",
     "진단 모드 수행 후 재기동",
     "시린지 홈 센서 위치 틀어짐",
     "센서 위치 재조정 및 원점 복귀",
     MATCH_PART, 120, "-"),

    # ── Mettler Toledo M300 ────────────────────────────────
    ("M300", "M300-Cond-Cell-open", "Cond Cell open 알람",
     "매뉴얼대로 배선 및 셀 상태 확인",
     "센서 케이블 단선 (트레이 통과부 손상)",
     "케이블 교체 및 트레이 보호대 추가",
     MATCH_FULL, 140, "센서 케이블"),

    ("M300", "M300-Cond-Cell-open", "Cond Cell open 알람",
     "배선 점검 (이상 없음)",
     "상류 밸브 차단으로 셀 건조",
     "밸브 복구 후 셀 재침수, 자연 해소",
     MATCH_FULL, 45, "-"),

    ("M300", "M300-Cond-Cell-shorted", "Cond Cell shorted 알람",
     "매뉴얼대로 센서·케이블 단락 점검",
     "커넥터 침수 (판넬 결로)",
     "커넥터 건조·재결선, 판넬 히터 설치 요청",
     MATCH_PART, 165, "커넥터 세트"),

    ("M300", "M300-Watchdog-time-out", "Watchdog time-out 알람",
     "재기동 후 정상 복구",
     "판넬 24V 순간 전압 강하",
     "전원 품질 측정 및 UPS 회로 재배분",
     MATCH_PART, 200, "-"),

    ("M300", "M300-Cond-Cell-open", "Cond Cell open 알람",
     "배선 및 셀 점검",
     "셀 상수 설정값 오입력으로 오판정",
     "셀 상수 재설정 및 교정",
     MATCH_NONE, 70, "-"),

    ("M300", "M300-Watchdog-time-out", "Watchdog time-out 알람",
     "재기동",
     "펌웨어 이슈 — 벤더 확인 후 업데이트 권고",
     "펌웨어 업데이트 적용",
     MATCH_PART, 240, "-"),

    # ── Siemens ET200SP HA AI 16xI ────────────────────────
    ("AI 16xI 2-wire HART HA", "ET200SP-6H", "채널 단선 진단 (Wire break)",
     "매뉴얼대로 계기-모듈 간 배선 점검",
     "단자대 나사 풀림 (진동)",
     "전 채널 단자 토크 재점검 및 체결",
     MATCH_PART, 95, "-"),

    ("AI 16xI 2-wire HART HA", "ET200SP-105H", "L+ 단락 진단",
     "매뉴얼대로 프로세스 배선 점검",
     "케이블 피복 손상으로 센서전원 단락",
     "손상부 절단 후 재결선",
     MATCH_FULL, 130, "케이블"),

    ("AI 16xI 2-wire HART HA", "ET200SP-11H", "공급 전압 상실 진단",
     "매뉴얼대로 L+ 단자 배선 확인",
     "판넬 24V 파워서플라이 퓨즈 단선",
     "퓨즈 교체 및 부하 재계산",
     MATCH_PART, 85, "퓨즈 2A"),

    ("AI 16xI 2-wire HART HA", "ET200SP-8H", "하한 위반 진단 (Low limit violated)",
     "매뉴얼대로 계기-모듈 튜닝 확인",
     "M300 알람 전류가 3.6mA로 설정되어 있어 자기진단 신호가 하한 위반으로 보인 것",
     "M300 알람 전류를 22mA로 변경, 상하한 구분 가능해짐",
     MATCH_NONE, 110, "-"),

    ("AI 16xI 2-wire HART HA", "ET200SP-10EH", "센서 전원 단락·과부하 진단",
     "매뉴얼대로 배선 점검",
     "2선식 트랜스미터 극성 반대 결선",
     "극성 정정 후 정상화",
     MATCH_PART, 60, "-"),

    ("AI 16xI 2-wire HART HA", "ET200SP-6H", "채널 단선 진단 (Wire break)",
     "배선 점검 (이상 없음)",
     "미사용 채널이 Disable 되지 않아 단선으로 진단됨",
     "미사용 채널 Disable 처리",
     MATCH_NONE, 40, "-"),

    ("AI 16xI 2-wire HART HA", "ET200SP-7H", "상한 위반 진단 (High limit violated)",
     "계기 출력 확인",
     "계기 자기진단 22mA 출력 (계기측 알람)",
     "계기 매뉴얼 조회 후 계기측 조치로 이관",
     MATCH_FULL, 50, "-"),

    ("M9e", "M9E-10125", "iOS 유량센서 시료 유량 미검출",
     "지연 시간 내 유량 회복 대기",
     "상류 필터 교체 작업 중 일시 차단",
     "작업 완료 후 자동 복구, 조치 불요",
     MATCH_PART, 20, "-"),

    ("M300", "M300-Cond-Cell-shorted", "Cond Cell shorted 알람",
     "센서 점검",
     "센서 수명 만료 (전극 열화)",
     "센서 교체 및 교정",
     MATCH_PART, 155, "전도도 센서"),

    ("M9e", "M9E-1203", "IC 온도 아날로그 판독값 범위 초과",
     "매뉴얼대로 진단 수행",
     "온도 센서 리드선 접촉 불량",
     "리드선 재결선",
     MATCH_PART, 100, "-"),
]

TECHS = ["K.M.", "J.H.", "S.Y.", "P.C.", "L.D."]


def load_tags(xlsx_path):
    """Instrument List 에서 모델별 태그 목록을 뽑는다."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    hdr_i = next(i for i, r in enumerate(rows)
                 if r and "TAG" in [str(c).strip().upper() if c else "" for c in r])
    hdr = [str(c).strip().upper() if c else "" for c in rows[hdr_i]]
    ti, mi, si = hdr.index("TAG"), hdr.index("MODEL"), hdr.index("SYSTEM")
    by_model = {}
    for r in rows[hdr_i + 1:]:
        if not r or not r[ti]:
            continue
        by_model.setdefault(str(r[mi]), []).append((str(r[ti]), str(r[si])))
    return by_model


def build(args):
    rnd = random.Random(args.seed)
    src = args.src

    codes = {}
    ec = os.path.join(src, "error_codes.json")
    if os.path.exists(ec):
        for x in json.load(open(ec, encoding="utf-8")):
            codes[x["id"]] = x
    else:
        print("주의: error_codes.json 없음 — code_ref 검증 생략")

    tags = load_tags(os.path.join(src, "DEMO_INSTRUMENT_LIST.xlsx"))
    # ET200SP 는 계기가 아니라 카드이므로 아무 태그에나 붙는다
    all_tags = [t for v in tags.values() for t in v]

    scen = list(SCENARIOS)
    rnd.shuffle(scen)
    scen = scen[:args.count]

    today = dt.date(2026, 8, 1)
    recs = []
    missing = []
    for i, (dev, ref, symptom, first, cause, action, match, mins, part) in enumerate(scen, 1):
        if codes and ref not in codes:
            missing.append(ref)
        pool = tags.get(dev) or all_tags
        # 카드(ET200SP) 진단이라도 원인 문구가 특정 계기를 지목하면 그 계기 태그에 붙인다
        for model in tags:
            if model in cause or model in action:
                pool = tags[model]
                break
        tag, system = rnd.choice(pool)
        day = today - dt.timedelta(days=rnd.randint(20, 540))
        recs.append({
            "wo_no": "WO-%d-%03d" % (day.year, i),
            "date": day.isoformat(),
            "tag": tag,
            "system": system,
            "device": dev,
            "symptom": symptom,
            "code_ref": ref,
            "first_action": first,
            "root_cause": cause,
            "action_taken": action,
            "manual_match": match,
            "duration_min": mins,
            "parts": part,
            "tech": rnd.choice(TECHS),
        })
    recs.sort(key=lambda r: r["date"])

    os.makedirs(args.out, exist_ok=True)
    jp = os.path.join(args.out, "maintenance_history.json")
    with open(jp, "w", encoding="utf-8") as f:
        json.dump(recs, f, ensure_ascii=False, indent=2)

    # XLSX
    HDR = ["WO No.", "발생일", "TAG", "계통", "기기", "증상", "매뉴얼 코드",
           "초동 조치", "실제 원인", "조치 내용", "매뉴얼 일치도",
           "소요(분)", "사용 부품", "담당"]
    wb = Workbook()
    ws = wb.active
    ws.title = "보수 이력"
    ws["A1"] = ("데모용 합성 데이터입니다. 실제 보수 이력이 아닙니다. "
                "매뉴얼 원문은 error_codes.json 을 매뉴얼 코드로 조인해 참조하세요.")
    ws["A1"].font = Font(name="Arial", size=9, italic=True, color="808080")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HDR))
    ws.append([])
    ws.append(HDR)
    hrow = ws.max_row

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fills = {MATCH_FULL: "E2EFDA", MATCH_PART: "FFF2CC", MATCH_NONE: "FCE4D6"}

    for c in range(1, len(HDR) + 1):
        cell = ws.cell(row=hrow, column=c)
        cell.font = Font(name="Arial", size=9, bold=True)
        cell.fill = PatternFill("solid", fgColor="D9D9D9")
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)

    mcol = HDR.index("매뉴얼 일치도") + 1
    for r in recs:
        ws.append([r["wo_no"], r["date"], r["tag"], r["system"], r["device"],
                   r["symptom"], r["code_ref"], r["first_action"],
                   r["root_cause"], r["action_taken"], r["manual_match"],
                   r["duration_min"], r["parts"], r["tech"]])
        rr = ws.max_row
        for c in range(1, len(HDR) + 1):
            cell = ws.cell(row=rr, column=c)
            cell.font = Font(name="Arial", size=9)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True,
                                       horizontal="left" if c in (6, 8, 9, 10)
                                       else "center")
        ws.cell(row=rr, column=mcol).fill = PatternFill(
            "solid", fgColor=fills[r["manual_match"]])

    for i, w in enumerate([12, 11, 12, 15, 20, 26, 20, 30, 34, 32, 11, 8, 16, 8], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=hrow + 1, column=3)
    ws.auto_filter.ref = "A%d:%s%d" % (hrow, get_column_letter(len(HDR)), ws.max_row)
    xp = os.path.join(args.out, "maintenance_history.xlsx")
    wb.save(xp)

    dist = {}
    for r in recs:
        dist[r["manual_match"]] = dist.get(r["manual_match"], 0) + 1
    print("생성 완료: %d건" % len(recs))
    print("  XLSX : %s" % xp)
    print("  JSON : %s" % jp)
    print("  일치도: " + ", ".join("%s %d건(%.0f%%)"
                                % (k, v, 100 * v / len(recs)) for k, v in dist.items()))
    if missing:
        print("  경고: error_codes.json 에 없는 code_ref %d종 → %s"
              % (len(set(missing)), ", ".join(sorted(set(missing)))))
    else:
        print("  code_ref 검증: 전건 매뉴얼 코드와 조인 가능")


def main():
    ap = argparse.ArgumentParser(description="가상 보수 이력 생성")
    ap.add_argument("--src", default="demo_data",
                    help="error_codes.json / DEMO_INSTRUMENT_LIST.xlsx 위치")
    ap.add_argument("--out", default="demo_data")
    ap.add_argument("--count", type=int, default=30)
    ap.add_argument("--seed", type=int, default=20260805)
    build(ap.parse_args())


if __name__ == "__main__":
    main()

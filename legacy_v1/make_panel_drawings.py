#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
make_panel_drawings.py — 합성 SCHEMATIC(결선도) / OUTLINE(판넬 배치도) 생성

P&ID 만으로는 정비원이 필요한 걸 다 못 봅니다.

    P&ID       공정상 어디에 붙어 있는 계기인가
    SCHEMATIC  IO 카드 채널부터 로컬 계기까지 어떻게 결선되어 있나
    OUTLINE    그 단자대가 판넬 어디에 박혀 있나

Instrument List 를 읽어 두 도면을 만들고, 기존 P&ID 색인과 합쳐
통합 색인(drawings_index.csv)을 출력합니다.

사용:
    python make_panel_drawings.py --src demo_data --out demo_data

의존성: reportlab, openpyxl
"""

import argparse
import csv
import os
import random
from collections import defaultdict

from openpyxl import load_workbook
from reportlab.lib.pagesizes import A3, landscape
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas

PAGE = landscape(A3)
PROJECT_TITLE = "DEMO WATER PLANT Ph1"
PROJECT_NO = "99001D"
MAKER = "DEMO ENGINEERING CO., LTD."
SCH_PREFIX = "DM-PNT1C01-UW-E30-"
OUT_PREFIX = "DM-PNT1C01-UW-E10-"
SCH_START = 20001
OUT_START = 30001

LOOPS_PER_SHEET = 5


# ─────────────────────────────────────────────────────────────
def load_instruments(src):
    p = os.path.join(src, "DEMO_INSTRUMENT_LIST.xlsx")
    ws = load_workbook(p, read_only=True, data_only=True).active
    rows = list(ws.iter_rows(values_only=True))
    hi = next(i for i, r in enumerate(rows)
              if r and "TAG" in [str(c).strip().upper() if c else "" for c in r])
    hdr = [str(c).strip().upper() if c else "" for c in rows[hi]]
    out = []
    for r in rows[hi + 1:]:
        if not r or not r[hdr.index("TAG")]:
            continue
        out.append({hdr[i]: r[i] for i in range(len(hdr)) if hdr[i]})
    return out


def title_block(c, w, h, sheet_no, title_lines, rev):
    bw, bh = 105 * mm, 40 * mm
    x0, y0 = w - 12 * mm - bw, 12 * mm
    c.setLineWidth(0.8)
    c.rect(x0, y0, bw, bh)
    c.setLineWidth(0.4)
    c.line(x0, y0 + bh - 5 * mm, x0 + bw, y0 + bh - 5 * mm)
    c.setFont("Helvetica-Bold", 5)
    c.drawString(x0 + 2 * mm, y0 + bh - 3.5 * mm, "REV  B    ISSUED FOR APPROVAL"
                                                  "    26-05-20    DEM")

    ty = y0 + bh - 11 * mm
    c.setFont("Helvetica-Bold", 6)
    c.drawString(x0 + 2 * mm, ty, "DWG. TITLE")
    c.setFont("Helvetica", 7)
    for i, line in enumerate(title_lines[:-1]):
        c.drawString(x0 + 24 * mm, ty - i * 3.4 * mm, line)
    c.setFont("Helvetica-Bold", 8)
    c.drawString(x0 + 24 * mm, ty - len(title_lines[:-1]) * 3.6 * mm,
                 title_lines[-1])

    py = y0 + 17 * mm
    c.setFont("Helvetica", 5.5)
    for i, lab in enumerate(["PROJECT TITLE", "PROJECT NO.", "MAKER"]):
        c.drawString(x0 + 2 * mm, py - i * 4 * mm, lab)
    c.setFont("Helvetica-Bold", 6)
    for i, val in enumerate([PROJECT_TITLE, PROJECT_NO, MAKER]):
        c.drawString(x0 + 26 * mm, py - i * 4 * mm, val)

    sy = y0 + 3 * mm
    c.setFont("Helvetica-Bold", 6)
    c.drawString(x0 + 2 * mm, sy, "SHEET NO.")
    c.setFont("Helvetica-Bold", 9)
    c.drawString(x0 + 26 * mm, sy, sheet_no)
    c.setFont("Helvetica-Bold", 6)
    c.drawString(x0 + 88 * mm, sy, "REV")
    c.setFont("Helvetica-Bold", 9)
    c.drawString(x0 + 96 * mm, sy, rev)


def border(c, w, h):
    c.setLineWidth(1.2)
    c.rect(10 * mm, 10 * mm, w - 20 * mm, h - 20 * mm)
    c.setLineWidth(0.4)
    c.rect(12 * mm, 12 * mm, w - 24 * mm, h - 24 * mm)


# ─────────────────────────────────────────────────────────────
# SCHEMATIC — 루프 결선도
# ─────────────────────────────────────────────────────────────
def draw_loop(c, x0, y, w, rec, rnd):
    """한 루프를 가로 띠 하나로 그린다. 좌 = 현장, 우 = 판넬."""
    tag = str(rec["TAG"])
    term = str(rec["TERMINAL"])          # 예: TB8-35
    tb, tn = (term.split("-") + ["1"])[:2]
    tn = int(tn)
    panel = str(rec["PANEL"])
    slot, ch = rec["SLOT"], rec["CH"]

    c.setLineWidth(0.4)
    c.setFont("Helvetica", 6)

    # ── 현장 계기 (2선식)
    fx = x0 + 6 * mm
    c.circle(fx + 9 * mm, y, 9 * mm)
    c.line(fx, y, fx + 18 * mm, y)
    c.setFont("Helvetica", 5.4)
    c.drawCentredString(fx + 9 * mm, y + 2 * mm, "XMTR")
    c.drawCentredString(fx + 9 * mm, y - 4.2 * mm, tag)
    c.setFont("Helvetica", 5)
    c.drawCentredString(fx + 9 * mm, y - 13 * mm, str(rec["SERVICE"])[:26])

    # ── 현장 접속함
    jx = fx + 34 * mm
    c.rect(jx, y - 11 * mm, 26 * mm, 22 * mm)
    c.drawCentredString(jx + 13 * mm, y + 7 * mm, "JB-%02d" % (1 + tn % 6))
    for k, lab in enumerate(("1", "2")):
        yy = y + (3 - k * 6) * mm
        c.circle(jx + 6 * mm, yy, 1.2 * mm)
        c.circle(jx + 20 * mm, yy, 1.2 * mm)
        c.line(jx + 7.2 * mm, yy, jx + 18.8 * mm, yy)
        c.drawString(jx + 9 * mm, yy + 1.2 * mm, lab)

    # ── 판넬 단자대
    px = jx + 62 * mm
    c.rect(px, y - 11 * mm, 30 * mm, 22 * mm)
    c.setFont("Helvetica-Bold", 5.6)
    c.drawCentredString(px + 15 * mm, y + 7 * mm, "%s  %s" % (panel, tb))
    c.setFont("Helvetica", 5)
    for k in (0, 1):
        yy = y + (3 - k * 6) * mm
        c.circle(px + 7 * mm, yy, 1.2 * mm)
        c.circle(px + 23 * mm, yy, 1.2 * mm)
        c.line(px + 8.2 * mm, yy, px + 21.8 * mm, yy)
        c.drawString(px + 10.5 * mm, yy + 1.2 * mm, "%s-%d" % (tb, tn + k))

    # ── IO 카드
    ix = px + 52 * mm
    c.rect(ix, y - 11 * mm, 40 * mm, 22 * mm)
    c.setFont("Helvetica-Bold", 5.6)
    c.drawCentredString(ix + 20 * mm, y + 7 * mm, "AI 16xI 2-wire HART")
    c.setFont("Helvetica", 5)
    c.drawCentredString(ix + 20 * mm, y + 2 * mm,
                        "%s  SLOT %s" % (rec["PLC"], slot))
    c.drawCentredString(ix + 20 * mm, y - 3 * mm,
                        "CH%s   I%s+ / UV%s" % (ch, ch, ch))
    c.drawCentredString(ix + 20 * mm, y - 8.5 * mm, "4-20mA  2-WIRE")

    # ── 케이블 구간
    c.setLineWidth(0.6)
    for k in (0, 1):
        yy = y + (3 - k * 6) * mm
        c.line(fx + 18 * mm, y, jx + 6 * mm, yy) if k == 0 else None
        c.line(jx + 20 * mm, yy, px + 7 * mm, yy)
        c.line(px + 23 * mm, yy, ix, yy)
    c.line(fx + 18 * mm, y, jx + 6 * mm, y - 3 * mm)

    c.setFont("Helvetica", 4.8)
    c.drawCentredString((jx + 20 * mm + px + 7 * mm) / 2, y + 5 * mm,
                        "C-%04d  2C x 1.5SQ  SHIELDED" % (1000 + tn * 7))
    c.setFont("Helvetica", 5)
    c.drawString(px + 32 * mm, y + 1 * mm, "+")
    c.drawString(px + 32 * mm, y - 5 * mm, "-")


def build_schematic(recs, out_dir, rnd):
    w, h = PAGE
    path = os.path.join(out_dir, "DEMO_SCHEMATIC.pdf")
    c = canvas.Canvas(path, pagesize=PAGE)
    c.setTitle("Synthetic loop schematic for demo")
    index = []

    pages = [recs[i:i + LOOPS_PER_SHEET]
             for i in range(0, len(recs), LOOPS_PER_SHEET)]
    for p, group in enumerate(pages):
        sheet_no = "%s%06d" % (SCH_PREFIX, SCH_START + p)
        border(c, w, h)
        c.setFont("Helvetica-Bold", 8)
        c.drawString(22 * mm, h - 22 * mm,
                     "LOOP WIRING DIAGRAM   /   %s" % group[0]["SYSTEM"])
        c.setFont("Helvetica", 5.5)
        c.drawString(22 * mm, h - 27 * mm,
                     "FIELD INSTRUMENT        JUNCTION BOX          "
                     "PANEL TERMINAL BLOCK          I/O MODULE")

        for k, rec in enumerate(group):
            y = h - 52 * mm - k * 44 * mm
            draw_loop(c, 18 * mm, y, w, rec, rnd)
            c.setLineWidth(0.3)
            c.setDash(1, 2)
            c.line(20 * mm, y - 20 * mm, w - 125 * mm, y - 20 * mm)
            c.setDash()
            index.append({"TAG": str(rec["TAG"]), "TYPE": "SCHEMATIC",
                          "SHEET_NO": sheet_no, "FILE": "DEMO_SCHEMATIC.pdf",
                          "PAGE": p + 1, "FIND": str(rec["TAG"])})

        title_block(c, w, h, sheet_no,
                    ["DEMO UPW SYSTEM", "LOOP WIRING DIAGRAM FOR",
                     str(group[0]["SYSTEM"])], "B")
        c.showPage()
    c.save()
    return path, index


# ─────────────────────────────────────────────────────────────
# OUTLINE — 판넬 배치도
# ─────────────────────────────────────────────────────────────
def build_outline(recs, out_dir, rnd):
    w, h = PAGE
    path = os.path.join(out_dir, "DEMO_OUTLINE.pdf")
    c = canvas.Canvas(path, pagesize=PAGE)
    c.setTitle("Synthetic panel outline for demo")
    index = []

    by_panel = defaultdict(list)
    for r in recs:
        by_panel[str(r["PANEL"])].append(r)

    for p, (panel, items) in enumerate(sorted(by_panel.items())):
        sheet_no = "%s%06d" % (OUT_PREFIX, OUT_START + p)
        border(c, w, h)
        c.setFont("Helvetica-Bold", 8)
        c.drawString(22 * mm, h - 22 * mm,
                     "PANEL GENERAL ARRANGEMENT   /   %s" % panel)

        # 판넬 정면도
        px, py = 30 * mm, 40 * mm
        pw, ph = 120 * mm, 200 * mm
        c.setLineWidth(1.0)
        c.rect(px, py, pw, ph)
        c.setLineWidth(0.4)
        c.rect(px + 4 * mm, py + 4 * mm, pw - 8 * mm, ph - 8 * mm)
        c.setFont("Helvetica", 5.5)
        c.drawCentredString(px + pw / 2, py + ph + 4 * mm,
                            "FRONT VIEW   W800 x H2000 x D600")
        c.drawCentredString(px + pw / 2, py - 6 * mm, "SCALE 1:10")

        # 상단 기기 레일
        ry = py + ph - 24 * mm
        c.setLineWidth(0.6)
        c.rect(px + 10 * mm, ry, pw - 20 * mm, 14 * mm)
        c.setFont("Helvetica", 5)
        for k, lab in enumerate(["PS", "IM", "AI01", "AI02", "AI03", "AI04"]):
            bx = px + 12 * mm + k * 16 * mm
            if bx + 14 * mm > px + pw - 10 * mm:
                break
            c.rect(bx, ry + 1.5 * mm, 14 * mm, 11 * mm)
            c.drawCentredString(bx + 7 * mm, ry + 5.5 * mm, lab)
        c.setFont("Helvetica", 5)
        c.drawString(px + 10 * mm, ry + 16 * mm, "DIN RAIL 1  —  I/O MODULES")

        # 단자대 레일 — TB1..TB9
        tbs = sorted({str(r["TERMINAL"]).split("-")[0] for r in items})
        base = ry - 16 * mm
        pos = {}
        for k, tb in enumerate(tbs):
            ty = base - k * 15 * mm
            if ty < py + 12 * mm:
                break
            c.setLineWidth(0.6)
            c.rect(px + 10 * mm, ty, pw - 20 * mm, 10 * mm)
            c.setFont("Helvetica-Bold", 6)
            c.drawString(px + 12 * mm, ty + 3.4 * mm, tb)
            c.setFont("Helvetica", 4.6)
            for j in range(1, 21):
                xx = px + 24 * mm + (j - 1) * 3.9 * mm
                if xx > px + pw - 12 * mm:
                    break
                c.line(xx, ty, xx, ty + 10 * mm)
                if j % 5 == 0:
                    c.drawCentredString(xx - 2 * mm, ty + 3.4 * mm, str(j))
            pos[tb] = ty

        # 단자 배정표 (우측)
        tx = px + pw + 18 * mm
        c.setFont("Helvetica-Bold", 6)
        c.drawString(tx, py + ph - 4 * mm, "TERMINAL ASSIGNMENT")
        c.setFont("Helvetica", 5.2)
        hdr_y = py + ph - 10 * mm
        c.drawString(tx, hdr_y, "TERMINAL")
        c.drawString(tx + 26 * mm, hdr_y, "TAG")
        c.drawString(tx + 54 * mm, hdr_y, "SERVICE")
        c.setLineWidth(0.3)
        c.line(tx, hdr_y - 1.5 * mm, tx + 120 * mm, hdr_y - 1.5 * mm)
        for k, r in enumerate(sorted(items, key=lambda x: str(x["TERMINAL"]))):
            yy = hdr_y - 6 * mm - k * 4.4 * mm
            if yy < py + 6 * mm:
                break
            c.drawString(tx, yy, str(r["TERMINAL"]))
            c.drawString(tx + 26 * mm, yy, str(r["TAG"]))
            c.drawString(tx + 54 * mm, yy, str(r["SERVICE"])[:34])

        title_block(c, w, h, sheet_no,
                    ["DEMO UPW SYSTEM", "PANEL GENERAL ARRANGEMENT FOR",
                     panel], "A")
        c.showPage()

        for r in items:
            index.append({"TAG": str(r["TAG"]), "TYPE": "OUTLINE",
                          "SHEET_NO": sheet_no, "FILE": "DEMO_OUTLINE.pdf",
                          "PAGE": p + 1,
                          "FIND": "%s|%s" % (str(r["TERMINAL"]).split("-")[0],
                                             str(r["TERMINAL"]))})
    c.save()
    return path, index


# ─────────────────────────────────────────────────────────────
def load_pid_index(src):
    p = os.path.join(src, "DEMO_PID_truth.csv")
    out = []
    if not os.path.exists(p):
        return out
    with open(p, newline="", encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            out.append({"TAG": r["TAG"], "TYPE": "P&ID",
                        "SHEET_NO": r["SHEET_NO"], "FILE": "DEMO_PID.pdf",
                        "PAGE": int(r["PAGE"]), "FIND": r["TAG"]})
    return out


def main():
    ap = argparse.ArgumentParser(description="SCHEMATIC / OUTLINE 도면 생성")
    ap.add_argument("--src", default="demo_data")
    ap.add_argument("--out", default="demo_data")
    ap.add_argument("--seed", type=int, default=20260806)
    args = ap.parse_args()

    rnd = random.Random(args.seed)
    recs = load_instruments(args.src)
    os.makedirs(args.out, exist_ok=True)

    sch_path, sch_idx = build_schematic(recs, args.out, rnd)
    out_path, out_idx = build_outline(recs, args.out, rnd)
    pid_idx = load_pid_index(args.src)

    index = pid_idx + sch_idx + out_idx
    idx_path = os.path.join(args.out, "drawings_index.csv")
    with open(idx_path, "w", newline="", encoding="utf-8-sig") as f:
        wr = csv.DictWriter(f, fieldnames=["TAG", "TYPE", "SHEET_NO",
                                           "FILE", "PAGE", "FIND"])
        wr.writeheader()
        wr.writerows(sorted(index, key=lambda r: (r["TAG"], r["TYPE"])))

    n = {}
    for r in index:
        n[r["TYPE"]] = n.get(r["TYPE"], 0) + 1
    print("생성 완료")
    print("  SCHEMATIC : %s" % sch_path)
    print("  OUTLINE   : %s" % out_path)
    print("  통합 색인 : %s  (%d행)" % (idx_path, len(index)))
    print("  유형별    : %s" % ", ".join("%s %d" % kv for kv in n.items()))


if __name__ == "__main__":
    main()

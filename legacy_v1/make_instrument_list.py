#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
make_instrument_list.py — 데모용 가상 Instrument List 생성

Plant Maintenance Copilot 프로토타입의 기준 데이터.
여기서 나온 TAG를 make_fake_pid.py 에 넣으면 도면과 리스트가 맞물린다.

    python make_instrument_list.py
    python make_fake_pid.py --tags-from DEMO_INSTRUMENT_LIST.xlsx --tag-col TAG

의존성: openpyxl
"""

import argparse
import csv
import os
import random

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────
# 계통 정의 — (코드, 계통명, 루프그룹 목록)
# ─────────────────────────────────────────────────────────────
SYSTEMS = [
    ("1", "PRETREATMENT",   ["PT-TRAIN-A", "PT-TRAIN-B"]),
    ("2", "RO SYSTEM",      ["RO-TRAIN-A", "RO-TRAIN-B"]),
    ("3", "CEDI SYSTEM",    ["CEDI-STACK-1"]),
    ("4", "UPW TANK",       ["TANK-01"]),
    ("5", "POLISHING LOOP", ["POL-LOOP-1"]),
    ("6", "DISTRIBUTION",   ["DIST-LOOP-A", "DIST-LOOP-B"]),
    ("7", "RECLAIM",        ["RCL-01"]),
]

# 계기 유형 — (함수코드, 서비스, 메이커, 모델, 레인지, 단위, 결함거동, 매뉴얼)
M300 = ("METTLER TOLEDO", "M300", "Upscale 22mA (self-diag)", "MT_M300_OM.pdf")
M9E = ("VEOLIA SIEVERS", "M9e", "Upscale 22mA (self-diag)", "SIEVERS_M9e_OM.pdf")
GEN = ("GENERIC", "TX-2000", "Downscale <3.6mA (NAMUR)", "GENERIC_TX_OM.pdf")

TYPES = [
    ("AIT", "RESISTIVITY",      M300, 0.0, 18.2, "Mohm-cm"),
    ("AIT", "CONDUCTIVITY",     M300, 0.0, 100.0, "uS/cm"),
    ("AIT", "pH",               M300, 0.0, 14.0, "pH"),
    ("AIT", "DISSOLVED OXYGEN", M300, 0.0, 200.0, "ppb"),
    ("AIT", "TOC",              M9E,  0.0, 50.0, "ppb"),
    ("PIT", "PRESSURE",         GEN,  0.0, 10.0, "bar"),
    ("FIT", "FLOW",             GEN,  0.0, 50.0, "m3/h"),
    ("LIT", "LEVEL",            GEN,  0.0, 100.0, "%"),
    ("TIT", "TEMPERATURE",      GEN,  0.0, 80.0, "degC"),
]

# 계통별로 들어갈 계기 유형 (인덱스)
SYSTEM_TYPES = {
    "1": [1, 5, 6, 8],
    "2": [0, 1, 5, 6, 7],
    "3": [0, 1, 5, 6],
    "4": [0, 4, 7, 8],
    "5": [0, 3, 4, 5, 6],
    "6": [0, 4, 5, 6, 8],
    "7": [1, 2, 5, 7],
}

PANELS = ["CUB-A", "CUB-B", "CUB-C", "RIO-01", "RIO-02"]

HEADERS = ["NO", "TAG", "SERVICE", "SYSTEM", "LOOP GROUP", "MAKER", "MODEL",
           "MEAS TYPE", "RANGE MIN", "RANGE MAX", "UNIT", "SIGNAL",
           "FAULT MODE", "PLC", "RACK", "SLOT", "CH", "PANEL", "TERMINAL",
           "DWG No.", "ALARM L", "ALARM H", "MANUAL FILE", "REMARK"]


def build_rows(target, rnd):
    rows = []
    seq = {}
    while len(rows) < target:
        for code, sysname, groups in SYSTEMS:
            if len(rows) >= target:
                break
            for ti in SYSTEM_TYPES[code]:
                if len(rows) >= target:
                    break
                func, service, vendor, rmin, rmax, unit = TYPES[ti]
                maker, model, fault, manual = vendor
                seq[code] = seq.get(code, 0) + 1
                tag = "%s-%s%03d" % (func, code, seq[code])
                group = rnd.choice(groups)

                # 알람 설정값 — 레인지의 10% / 90% 근처
                span = rmax - rmin
                alo = round(rmin + span * rnd.uniform(0.05, 0.15), 2)
                ahi = round(rmin + span * rnd.uniform(0.80, 0.92), 2)
                if service in ("RESISTIVITY",):
                    alo, ahi = 17.0, ""          # 저항률은 하한만 의미 있음
                if service in ("TOC",):
                    alo, ahi = "", 5.0           # TOC는 상한만

                rack = 0
                slot = rnd.randint(1, 8)
                ch = rnd.randint(0, 7)
                rows.append([
                    len(rows) + 1, tag, "%s %s" % (sysname, service), sysname,
                    group, maker, model, service, rmin, rmax, unit,
                    "4-20mA AI", fault,
                    "PLC-UPW-01", rack, slot, ch,
                    rnd.choice(PANELS),
                    "TB%d-%d" % (rnd.randint(1, 9), rnd.randint(1, 40)),
                    "", alo, ahi, manual, "",
                ])
    return rows


def write_xlsx(rows, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "INSTRUMENT LIST"

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="D9D9D9")
    fill_me = PatternFill("solid", fgColor="FFF2CC")

    # 범례
    ws["A1"] = ("데모용 가상 데이터입니다. 실제 프로젝트 자료가 아닙니다. "
                "노란색 DWG No. 열은 도면 매핑으로 채워지는 칸입니다.")
    ws["A1"].font = Font(name="Arial", size=9, italic=True, color="808080")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))

    ws.append([])
    ws.append(HEADERS)
    hrow = ws.max_row
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=hrow, column=c)
        cell.font = Font(name="Arial", size=9, bold=True)
        cell.fill = head_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)

    dwg_col = HEADERS.index("DWG No.") + 1
    for r in rows:
        ws.append(r)
        rr = ws.max_row
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=rr, column=c)
            cell.font = Font(name="Arial", size=9)
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if c != 3 else "left",
                                       vertical="center")
        ws.cell(row=rr, column=dwg_col).fill = fill_me

    widths = [5, 12, 26, 16, 14, 17, 10, 16, 10, 10, 10, 11, 24, 12, 6, 6, 5,
              9, 10, 24, 9, 9, 22, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = ws.cell(row=hrow + 1, column=3)
    ws.auto_filter.ref = "A%d:%s%d" % (hrow, get_column_letter(len(HEADERS)),
                                       ws.max_row)
    wb.save(path)


def main():
    ap = argparse.ArgumentParser(description="가상 Instrument List 생성")
    ap.add_argument("--count", type=int, default=72, help="계기 점수 (기본 72)")
    ap.add_argument("--outdir", default=".", help="출력 폴더")
    ap.add_argument("--name", default="DEMO_INSTRUMENT_LIST", help="파일명")
    ap.add_argument("--seed", type=int, default=20260805)
    args = ap.parse_args()

    rnd = random.Random(args.seed)
    rows = build_rows(args.count, rnd)
    os.makedirs(args.outdir, exist_ok=True)

    xlsx = os.path.join(args.outdir, args.name + ".xlsx")
    write_xlsx(rows, xlsx)

    csvp = os.path.join(args.outdir, args.name + ".csv")
    with open(csvp, "w", newline="", encoding="utf-8-sig") as f:
        wr = csv.writer(f)
        wr.writerow(HEADERS)
        wr.writerows(rows)

    models = {}
    for r in rows:
        models[r[6]] = models.get(r[6], 0) + 1
    print("생성 완료: %d점" % len(rows))
    print("  XLSX : %s" % xlsx)
    print("  CSV  : %s" % csvp)
    print("  모델별: %s" % ", ".join("%s %d점" % kv for kv in models.items()))


if __name__ == "__main__":
    main()

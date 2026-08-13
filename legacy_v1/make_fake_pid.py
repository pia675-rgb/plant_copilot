#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
make_fake_pid.py — 합성(가짜) P&ID 도면 PDF 생성기

실제 고객 도면 대신 데모/검증에 쓸 벡터 PDF를 만든다.
- 도각(우측 하단)에 SHEET NO. / PROJECT NO. / REV / 개정이력 포함
- 계기 버블: 원 안 2줄 (상단 = 기능·루프 표기, 하단 = IO List TAG)
- 본문 노이즈: SPARE, FG, 단자번호, EPLAN 상호참조(/010141.12:D), 참조도면번호
- 정답지(ground truth) CSV 및 매핑 입력용 XLSX 동시 생성

사용 예:
    python make_fake_pid.py                          # 기본값으로 생성
    python make_fake_pid.py --sheets 12 --per-sheet 20
    python make_fake_pid.py --tags-from myio.xlsx --tag-col TAG
    python make_fake_pid.py --bubble-style mixed --dup-across 3

의존성: reportlab, openpyxl (XLSX 출력 시)
"""

import argparse
import csv
import os
import random
import string

from reportlab.lib.pagesizes import A3, landscape
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas

# ─────────────────────────────────────────────────────────────
# 가상 프로젝트 정보 — 실제 프로젝트와 겹치지 않게 지어낸 값
# ─────────────────────────────────────────────────────────────
PROJECT_TITLE = "DEMO WATER PLANT Ph1"
PROJECT_NO = "99001D"
MAKER = "DEMO ENGINEERING CO., LTD."
SHEET_PREFIX = "DM-PNT1C01-UW-A20-"
SHEET_START = 10001          # 마지막 6자리: 010001 부터
TITLE_HEAD = ("DEMO UPW SYSTEM",
              "PIPING & INSTRUMENT DIAGRAM FOR")

AREA_NAMES = ["PRETREATMENT", "RO SYSTEM", "CEDI SYSTEM", "POLISHING LOOP",
              "UPW TANK AREA", "DISTRIBUTION LOOP", "WASTE COLLECTION",
              "CHEMICAL DOSING", "DEGASSER AREA", "UV STERILIZER"]

FUNC_CODES = ["LIA", "PIT", "FIC", "TIT", "PSH", "LSL", "AIT", "FIT",
              "PIC", "LIC", "TSH", "FSL", "PDT", "CIT"]

EQUIP_NAMES = ["FEED PUMP", "RO SKID", "CEDI STACK", "UPW TANK", "UF UNIT",
               "TRANSFER PUMP", "CARTRIDGE FILTER", "DEGASSER", "UV UNIT"]

NOISE_WORDS = ["SPARE", "FG", "N.C.", "N.O.", "TYP.", "NOTE 1", "HOLD",
               "CONT'D", "BY OTHERS", "FUTURE"]

PAGE = landscape(A3)   # 420 x 297 mm


# ─────────────────────────────────────────────────────────────
# 태그 생성 / 로드
# ─────────────────────────────────────────────────────────────
def gen_tags(n, rnd):
    """P1511B 형식 — 영문1 + 숫자4 + 영문1 (구분자 없음)."""
    out, seen = [], set()
    heads = list("PFLTAC")
    tails = list("ABCD")
    while len(out) < n:
        t = "%s%04d%s" % (rnd.choice(heads), rnd.randint(1000, 9999),
                          rnd.choice(tails))
        if t not in seen:
            seen.add(t)
            out.append(t)
    return out


def load_tags(path, col_name):
    """엑셀/CSV에서 TAG 컬럼을 읽어온다."""
    tags = []
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = None
        idx = None
        for row in rows:
            if header is None:
                header = [str(c).strip().upper() if c is not None else ""
                          for c in row]
                if col_name.upper() in header:
                    idx = header.index(col_name.upper())
                    continue
                header = None
                continue
            if idx is None or idx >= len(row):
                continue
            v = row[idx]
            if v is None:
                continue
            v = str(v).strip()
            if v and v.upper() not in ("SPARE", "-", "NONE", "NAN"):
                tags.append(v)
        wb.close()
    else:
        with open(path, newline="", encoding="utf-8-sig") as f:
            for r in csv.DictReader(f):
                v = (r.get(col_name) or "").strip()
                if v and v.upper() not in ("SPARE", "-", "NONE", "NAN"):
                    tags.append(v)
    # 중복 제거 (순서 유지)
    seen, out = set(), []
    for t in tags:
        if t not in seen:
            seen.add(t)
            out.append(t)
    return out


# ─────────────────────────────────────────────────────────────
# 도면 요소 그리기
# ─────────────────────────────────────────────────────────────
def draw_border(c, w, h):
    c.setLineWidth(1.2)
    c.rect(10 * mm, 10 * mm, w - 20 * mm, h - 20 * mm)
    c.setLineWidth(0.4)
    c.rect(12 * mm, 12 * mm, w - 24 * mm, h - 24 * mm)


def draw_title_block(c, w, h, sheet_no, area, rev, rnd):
    """우측 하단 도각. SHEET NO. 라벨과 값을 같은 높이에 둔다."""
    bw, bh = 105 * mm, 48 * mm
    x0, y0 = w - 12 * mm - bw, 12 * mm
    c.setLineWidth(0.8)
    c.rect(x0, y0, bw, bh)

    # 개정 이력 표 (상단 3행)
    rowh = 5 * mm
    c.setLineWidth(0.4)
    for i in range(4):
        c.line(x0, y0 + bh - i * rowh, x0 + bw, y0 + bh - i * rowh)
    for cx in (18, 70, 88):
        c.line(x0 + cx * mm, y0 + bh - 3 * rowh, x0 + cx * mm, y0 + bh)
    c.setFont("Helvetica-Bold", 5)
    c.drawString(x0 + 2 * mm, y0 + bh - rowh + 1.6 * mm, "REV")
    c.drawString(x0 + 20 * mm, y0 + bh - rowh + 1.6 * mm, "DESCRIPTION")
    c.drawString(x0 + 72 * mm, y0 + bh - rowh + 1.6 * mm, "DATE")
    c.drawString(x0 + 90 * mm, y0 + bh - rowh + 1.6 * mm, "APP'D")
    c.setFont("Helvetica", 5)
    hist = [("A", "ISSUED FOR REVIEW", "26-03-11"),
            ("B", "ISSUED FOR APPROVAL", "26-05-20")]
    for i, (r, d, dt) in enumerate(hist):
        yy = y0 + bh - (i + 2) * rowh + 1.6 * mm
        c.drawString(x0 + 2 * mm, yy, r)
        c.drawString(x0 + 20 * mm, yy, d)
        c.drawString(x0 + 72 * mm, yy, dt)
        c.drawString(x0 + 90 * mm, yy, "DEM")

    # 도면명 (공통 머리말 + 개별 내용)
    ty = y0 + bh - 3 * rowh - 5 * mm
    c.setFont("Helvetica-Bold", 6)
    c.drawString(x0 + 2 * mm, ty, "DWG. TITLE")
    c.setFont("Helvetica", 7)
    for i, line in enumerate(TITLE_HEAD):
        c.drawString(x0 + 22 * mm, ty - i * 3.4 * mm, line)
    c.setFont("Helvetica-Bold", 8)
    c.drawString(x0 + 22 * mm, ty - 2 * 3.6 * mm, area)

    # 프로젝트 정보
    py = y0 + 14 * mm
    c.setFont("Helvetica", 5.5)
    c.drawString(x0 + 2 * mm, py, "PROJECT TITLE")
    c.drawString(x0 + 2 * mm, py - 4 * mm, "PROJECT NO.")
    c.drawString(x0 + 2 * mm, py - 8 * mm, "MAKER")
    c.setFont("Helvetica-Bold", 6)
    c.drawString(x0 + 26 * mm, py, PROJECT_TITLE)
    c.drawString(x0 + 26 * mm, py - 4 * mm, PROJECT_NO)
    c.drawString(x0 + 26 * mm, py - 8 * mm, MAKER)

    # SHEET NO. — 라벨과 값을 같은 baseline 에 (추출기 우선순위 로직용)
    sy = y0 + 3 * mm
    c.setFont("Helvetica-Bold", 6)
    c.drawString(x0 + 2 * mm, sy, "SHEET NO.")
    c.setFont("Helvetica-Bold", 9)
    c.drawString(x0 + 26 * mm, sy, sheet_no)
    c.setFont("Helvetica-Bold", 6)
    c.drawString(x0 + 88 * mm, sy, "REV")
    c.setFont("Helvetica-Bold", 9)
    c.drawString(x0 + 96 * mm, sy, rev)

    # ISSUE 스탬프
    c.setLineWidth(0.6)
    c.rect(x0 - 26 * mm, y0 + 2 * mm, 24 * mm, 10 * mm)
    c.setFont("Helvetica-Bold", 7)
    c.drawCentredString(x0 - 14 * mm, y0 + 8 * mm, "ISSUE")
    c.setFont("Helvetica", 6)
    c.drawCentredString(x0 - 14 * mm, y0 + 4 * mm, "FOR APPROVAL")


def draw_process(c, w, h, rnd):
    """배관/설비 실루엣 — 도면처럼 보이게 하는 배경."""
    c.setLineWidth(0.9)
    for i in range(4):
        y = (60 + i * 48) * mm / 1.0
        y = 60 * mm + i * 45 * mm
        if y > h - 40 * mm:
            continue
        c.line(20 * mm, y, w - 130 * mm, y)
    c.setLineWidth(0.5)
    for i in range(6):
        x = 40 * mm + i * 45 * mm
        if x > w - 135 * mm:
            continue
        c.line(x, 55 * mm, x, h - 30 * mm)

    c.setLineWidth(0.8)
    for i in range(3):
        x = 30 * mm + i * 90 * mm
        if x > w - 160 * mm:
            continue
        c.rect(x, 30 * mm, 34 * mm, 18 * mm)
        c.setFont("Helvetica", 6)
        c.drawCentredString(x + 17 * mm, 38 * mm, rnd.choice(EQUIP_NAMES))


def draw_bubble(c, x, y, func, tag, style, rnd):
    """계기 버블 — 원 + 가로선, 상단 기능표기 / 하단 TAG."""
    r = 7.5 * mm
    c.setLineWidth(0.7)
    c.circle(x, y, r)
    c.setLineWidth(0.5)
    c.line(x - r, y, x + r, y)
    # 계기 → 배관 연결 점선
    c.setDash(2, 2)
    c.line(x, y - r, x, y - r - 6 * mm)
    c.setDash()

    if style == "joined":
        c.setFont("Helvetica", 4.4)
        c.drawCentredString(x, y - 2.2 * mm, func + tag)
    elif style == "split":
        # 문자런/숫자런 경계에서 쪼개져 추출되는 케이스 재현
        c.setFont("Helvetica", 5.6)
        c.drawCentredString(x, y + 1.6 * mm, func)
        head, tail = tag[:-1], tag[-1:]
        c.drawString(x - 6.4 * mm, y - 4.2 * mm, head)
        c.drawString(x + 3.6 * mm, y - 4.2 * mm, tail)
    else:  # two_line (기본)
        c.setFont("Helvetica", 5.6)
        c.drawCentredString(x, y + 1.6 * mm, func)
        c.drawCentredString(x, y - 4.2 * mm, tag)


def _noise_spot(w, h, rnd):
    """버블 격자(가로 40mm / 세로 45mm 간격) 사이 여백에만 놓는다."""
    x = 32 * mm + rnd.randint(0, 6) * 40 * mm + rnd.uniform(14, 26) * mm
    y = 68 * mm + rnd.randint(0, 3) * 45 * mm + rnd.uniform(16, 30) * mm
    x = min(x, w - 142 * mm)
    y = min(y, h - 26 * mm)
    return x, y


def draw_noise(c, w, h, page_no, rnd):
    """태그 추출 시 걸러져야 할 텍스트들."""
    c.setFont("Helvetica", 5)
    for _ in range(10):
        c.drawString(*_noise_spot(w, h, rnd), text=rnd.choice(NOISE_WORDS))
    # 단자번호
    for _ in range(6):
        c.drawString(*_noise_spot(w, h, rnd),
                     text="TB%d-%d" % (rnd.randint(1, 9), rnd.randint(1, 40)))
    # EPLAN 상호참조
    for _ in range(5):
        c.drawString(*_noise_spot(w, h, rnd),
                     text="/%06d.%d:%s" % (rnd.randint(10001, 10099),
                                           rnd.randint(1, 20),
                                           rnd.choice("ABCDEF")))
    # 본문 참조 도면번호 — 도각 SHEET NO. 와 헷갈리게 하는 미끼
    c.setFont("Helvetica", 5.5)
    c.drawString(22 * mm, h - 22 * mm,
                 "REF. DWG : %s%06d" % (SHEET_PREFIX, SHEET_START + 90))


# ─────────────────────────────────────────────────────────────
# 메인
# ─────────────────────────────────────────────────────────────
def build(args):
    rnd = random.Random(args.seed)
    w, h = PAGE

    n_needed = args.sheets * args.per_sheet
    if args.tags_from:
        tags = load_tags(args.tags_from, args.tag_col)
        if not tags:
            raise SystemExit("태그를 찾지 못했습니다: %s (컬럼 %s)"
                             % (args.tags_from, args.tag_col))
        if len(tags) > n_needed:
            tags = rnd.sample(tags, n_needed)
        print("태그 %d종 로드: %s" % (len(tags), args.tags_from))
    else:
        tags = gen_tags(n_needed, rnd)
        print("태그 %d종 자동 생성" % len(tags))

    # 시트별 배분 — 한 태그는 한 시트에만 (정상 케이스)
    rnd.shuffle(tags)
    per = max(1, len(tags) // args.sheets)
    sheets = [tags[i * per:(i + 1) * per] for i in range(args.sheets)]
    for i, t in enumerate(tags[args.sheets * per:]):
        sheets[i % args.sheets].append(t)

    # 일부러 여러 시트에 걸치게 심는 케이스 (복수 검출 테스트용)
    dup_tags = []
    if args.dup_across > 0 and args.sheets >= 2:
        pool = [t for s in sheets for t in s]
        dup_tags = rnd.sample(pool, min(args.dup_across, len(pool)))
        for t in dup_tags:
            home = next(i for i, s in enumerate(sheets) if t in s)
            other = rnd.choice([i for i in range(args.sheets) if i != home])
            sheets[other].append(t)

    os.makedirs(args.outdir, exist_ok=True)
    pdf_path = os.path.join(args.outdir, args.name + ".pdf")
    c = canvas.Canvas(pdf_path, pagesize=PAGE)
    c.setTitle("Synthetic P&ID for demo")
    c.setAuthor("demo generator")

    truth = []   # (tag, sheet_no, page, occurrences)
    for p, page_tags in enumerate(sheets):
        sheet_no = "%s%06d" % (SHEET_PREFIX, SHEET_START + p)
        area = AREA_NAMES[p % len(AREA_NAMES)]
        rev = rnd.choice(["A", "B", "C"])

        draw_border(c, w, h)
        draw_process(c, w, h, rnd)
        draw_noise(c, w, h, p + 1, rnd)

        # 버블 배치 — 느슨한 격자 + 흔들기
        cols, rows = 7, 4
        slots = [(i, j) for i in range(cols) for j in range(rows)]
        rnd.shuffle(slots)
        placed = {}
        for k, tag in enumerate(page_tags):
            reps = 2 if rnd.random() < args.repeat_ratio else 1  # 같은 시트 내 중복은 정상
            for _ in range(reps):
                if not slots:
                    break
                i, j = slots.pop()
                x = 32 * mm + i * 40 * mm + rnd.uniform(-4, 4) * mm
                y = 68 * mm + j * 45 * mm + rnd.uniform(-4, 4) * mm
                if x > w - 140 * mm:
                    x = w - 145 * mm
                style = args.bubble_style
                if style == "mixed":
                    style = rnd.choices(["two_line", "joined", "split"],
                                        weights=[7, 2, 1])[0]
                draw_bubble(c, x, y, rnd.choice(FUNC_CODES) + str(rnd.randint(1, 9)),
                            tag, style, rnd)
                placed[tag] = placed.get(tag, 0) + 1

        draw_title_block(c, w, h, sheet_no, area, rev, rnd)
        c.showPage()

        for tag, n in placed.items():
            truth.append((tag, sheet_no, p + 1, n))

    c.save()

    # 정답지 CSV
    csv_path = os.path.join(args.outdir, args.name + "_truth.csv")
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        wr = csv.writer(f)
        wr.writerow(["TAG", "SHEET_NO", "PAGE", "OCCURRENCES"])
        wr.writerows(sorted(truth))

    # 매핑 입력용 XLSX (DWG No. 비움)
    xlsx_path = ""
    if args.xlsx:
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "IO LIST"
            ws.append(["INDEX", "TAG", "DESCRIPTION", "DWG No.", "P&ID TAG"])
            uniq = sorted({t for t, _, _, _ in truth})
            for i, t in enumerate(uniq, 1):
                ws.append([i, t, "DEMO INSTRUMENT %s" % t, "", t])
            xlsx_path = os.path.join(args.outdir, args.name + "_taglist.xlsx")
            wb.save(xlsx_path)
        except ImportError:
            print("openpyxl 미설치 — XLSX 생성 생략")

    uniq_tags = {t for t, _, _, _ in truth}
    print("\n생성 완료")
    print("  PDF        : %s  (%d페이지)" % (pdf_path, len(sheets)))
    print("  정답지 CSV : %s" % csv_path)
    if xlsx_path:
        print("  태그 XLSX  : %s" % xlsx_path)
    print("  고유 태그  : %d종 / 버블 배치 %d건"
          % (len(uniq_tags), sum(n for _, _, _, n in truth)))
    if dup_tags:
        print("  복수 시트 태그 (검출 테스트용): %s" % ", ".join(sorted(dup_tags)))


def main():
    ap = argparse.ArgumentParser(description="합성 P&ID 도면 PDF 생성")
    ap.add_argument("--sheets", type=int, default=8, help="도면 매수 (기본 8)")
    ap.add_argument("--per-sheet", type=int, default=16,
                    help="시트당 태그 수 (기본 16)")
    ap.add_argument("--tags-from", default="",
                    help="실제 태그를 쓸 IO List 파일 (xlsx/csv)")
    ap.add_argument("--tag-col", default="TAG", help="태그 컬럼명 (기본 TAG)")
    ap.add_argument("--bubble-style", default="two_line",
                    choices=["two_line", "joined", "split", "mixed"],
                    help="버블 텍스트 형태 (기본 two_line, 검증엔 mixed)")
    ap.add_argument("--dup-across", type=int, default=2,
                    help="일부러 여러 시트에 심을 태그 수 (기본 2)")
    ap.add_argument("--repeat-ratio", type=float, default=0.25,
                    help="같은 시트 내 중복 배치 비율 (기본 0.25)")
    ap.add_argument("--outdir", default="fake_pid", help="출력 폴더")
    ap.add_argument("--name", default="DEMO_PID", help="출력 파일명 접두")
    ap.add_argument("--no-xlsx", dest="xlsx", action="store_false",
                    help="태그 XLSX 생성 생략")
    ap.add_argument("--seed", type=int, default=20260805, help="난수 시드")
    ap.set_defaults(xlsx=True)
    build(ap.parse_args())


if __name__ == "__main__":
    main()

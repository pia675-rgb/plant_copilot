#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
copilot_core.py — Plant Maintenance Copilot 검색·대비 엔진

기획서 3주 MVP 중 "알람 입력 → 근거 인용 조치 제안" 부분.
임베딩 없이 코드 매칭 + 키워드 스코어링으로 동작한다 (306+30건 규모에 충분).

핵심 원칙: 매뉴얼(권위)과 현장 이력(경험)을 절대 섞지 않는다.
          둘이 어긋나면 어긋난다고 표시한다.

입력 데이터 (모두 demo_data/):
    DEMO_INSTRUMENT_LIST.xlsx  태그 ↔ 모델 ↔ 판넬 ↔ 단자 (색인의 축)
    error_codes.json           매뉴얼 코드표 306건
    maintenance_history.json   현장 보수 이력 30건
    DEMO_PID_truth.csv         태그 → 도면번호 색인

사용:
    python copilot_core.py --tag AIT-4002 --alarm "acid container low"
    python copilot_core.py --tag AIT-5001 --alarm "셀 단선"
    python copilot_core.py --code M9E-10084
"""

import argparse
import csv
import json
import math
import os
import re
from collections import Counter, defaultdict

from openpyxl import load_workbook

# 한글 알람 입력을 영문 매뉴얼 용어로 넘겨주는 최소 사전
KO_EN = {
    "단선": "wire break open", "단락": "short circuit", "전원": "supply voltage",
    "유량": "flow", "압력": "pressure", "온도": "temperature",
    "레벨": "level reservoir", "전도도": "conductivity", "저항률": "resistivity",
    "기포": "bubble air", "누기": "leak bubble", "램프": "lamp uv",
    "산": "acid", "산화제": "oxidizer", "시약": "reagent",
    "펌프": "pump", "밸브": "valve", "튜브": "tube",
    "교정": "calibration", "셀": "cell", "센서": "sensor",
    "시린지": "syringe", "시료": "sample", "필터": "filter",
    "하한": "low limit underflow", "상한": "high limit overflow",
    "과부하": "overload", "수명": "life remaining", "막힘": "restricted",
    "워치독": "watchdog", "타임아웃": "time out timeout", "재기동": "restart",
    "드리프트": "drift unstable", "오염": "contamination dirty",
    "단자": "terminal", "채널": "channel", "모듈": "module",
    "과온": "overtemperature", "결로": "condensation",
}

CARD = "AI 16xI 2-wire HART HA"

STOP = set("the a an of is are be to for in on at and or with by from this that "
           "has have been not no if it its as detected please follow".split())


def tok(s):
    s = (s or "").lower()
    for ko, en in KO_EN.items():
        if ko in s:
            s += " " + en
    return [w for w in re.findall(r"[a-z0-9]+", s)
            if w not in STOP and len(w) > 1]


class Copilot:
    def __init__(self, src):
        self.src = src
        self.instruments = self._load_instruments()
        self.codes = json.load(open(os.path.join(src, "error_codes.json"),
                                    encoding="utf-8"))
        self.history = json.load(open(os.path.join(src, "maintenance_history.json"),
                                      encoding="utf-8"))
        self.drawings = self._load_drawings()
        self._build_idf()

    # ── 색인 로드 ────────────────────────────────────────
    def _load_instruments(self):
        p = os.path.join(self.src, "DEMO_INSTRUMENT_LIST.xlsx")
        ws = load_workbook(p, read_only=True, data_only=True).active
        rows = list(ws.iter_rows(values_only=True))
        hi = next(i for i, r in enumerate(rows)
                  if r and "TAG" in [str(c).strip().upper() if c else "" for c in r])
        hdr = [str(c).strip().upper() if c else "" for c in rows[hi]]
        out = {}
        for r in rows[hi + 1:]:
            if not r or not r[hdr.index("TAG")]:
                continue
            rec = {hdr[i]: r[i] for i in range(len(hdr)) if hdr[i]}
            out[str(rec["TAG"])] = rec
        return out

    def _load_drawings(self):
        """통합 색인(drawings_index.csv)이 있으면 그것을, 없으면 P&ID 색인만."""
        idx = defaultdict(list)
        uni = os.path.join(self.src, "drawings_index.csv")
        if os.path.exists(uni):
            with open(uni, newline="", encoding="utf-8-sig") as f:
                for r in csv.DictReader(f):
                    idx[r["TAG"]].append({
                        "type": r["TYPE"], "sheet_no": r["SHEET_NO"],
                        "file": r["FILE"], "page": int(r["PAGE"]),
                        "find": r.get("FIND") or r["TAG"]})
            for tag in idx:
                idx[tag].sort(key=lambda d: {"P&ID": 0, "SCHEMATIC": 1,
                                             "OUTLINE": 2}.get(d["type"], 9))
            return idx

        p = os.path.join(self.src, "DEMO_PID_truth.csv")
        if os.path.exists(p):
            with open(p, newline="", encoding="utf-8-sig") as f:
                for r in csv.DictReader(f):
                    idx[r["TAG"]].append({
                        "type": "P&ID", "sheet_no": r["SHEET_NO"],
                        "file": "DEMO_PID.pdf", "page": int(r["PAGE"]),
                        "find": r["TAG"]})
        return idx

    def _build_idf(self):
        df = Counter()
        self._code_tokens = {}
        for c in self.codes:
            t = set(tok(c["name"] + " " + c["description"]))
            self._code_tokens[c["id"]] = t
            df.update(t)
        n = max(1, len(self.codes))
        self._idf = {w: math.log(1 + n / (1 + v)) for w, v in df.items()}

    # ── 매뉴얼 검색 ──────────────────────────────────────
    def search_manual(self, query, device=None, top=5):
        q = set(tok(query))
        qraw = (query or "").upper().replace(" ", "")
        # 코드는 토큰 단위로만 일치시킨다 (부분문자열 매칭 금지)
        qcodes = set(re.findall(r"[0-9A-Fa-f]{1,5}H?", query or ""))
        qcodes = {c.upper() for c in qcodes}
        hits = []
        for c in self.codes:
            score = 0.0
            # 코드 직접 지정 — 정비원이 "10123 떴어요" 하는 경우
            if c["code"] and c["code"].upper() in qcodes:
                score += 100
            if c["id"].upper() in qraw:
                score += 100
            overlap = q & self._code_tokens[c["id"]]
            score += sum(self._idf.get(w, 0) for w in overlap) * 6
            # 이름 칸에 걸린 키워드는 설명에만 걸린 것보다 무겁게
            name_hit = q & set(tok(c["name"]))
            score += sum(self._idf.get(w, 0) for w in name_hit) * 4
            if not c["name"]:
                score -= 3
            if score <= 0:
                continue
            if device:
                if c["device"] == device:
                    score += 12
                elif c["device"] != CARD:
                    continue   # 이 태그의 기기가 아닌 벤더 매뉴얼은 아예 배제
            if score <= 0:
                continue
            hits.append((score, c))

        def code_num(c):
            return int(c["code"]) if (c["code"] or "").isdigit() else 10 ** 9

        # 동점이면 코드 번호가 작은 쪽(대표 코드)이 먼저
        hits.sort(key=lambda x: (-x[0], code_num(x[1]), x[1]["id"]))
        # 공유 설명 코드는 대표 1건만
        seen_desc, out = set(), []
        for s, c in hits:
            key = (c["device"], c["description"][:60])
            if c.get("shared_description") and key in seen_desc:
                continue
            seen_desc.add(key)
            out.append((round(s, 1), c))
            if len(out) >= top:
                break
        return out

    # ── 현장 이력 검색 ───────────────────────────────────
    def search_history(self, tag=None, code_refs=(), device=None, top=6):
        refs = set(code_refs)
        hits = []
        for h in self.history:
            score = 0.0
            why = []
            if tag and h["tag"] == tag:
                score += 60
                why.append("동일 태그")
            if h["code_ref"] and h["code_ref"] in refs:
                score += 45
                why.append("동일 매뉴얼 코드")
            if device and h["device"] == device and score == 0:
                score += 12
                why.append("동일 기종")
            if score <= 0:
                continue
            hits.append((round(score, 1), why, h))
        hits.sort(key=lambda x: (-x[0], x[2]["date"]), reverse=False)
        hits.sort(key=lambda x: -x[0])
        return hits[:top]

    # ── 통합 응답 ────────────────────────────────────────
    def answer(self, tag=None, alarm="", code=None):
        inst = self.instruments.get(tag) if tag else None
        device = str(inst["MODEL"]) if inst else None
        query = alarm or ""
        if code:
            query = (query + " " + code).strip()

        manual = self.search_manual(query, device=device)
        refs = [c["id"] for _, c in manual]
        hist = self.search_history(tag=tag, code_refs=refs, device=device)

        # 매뉴얼 ↔ 현장 대비 집계
        cmp_counts = Counter(h["manual_match"] for _, _, h in hist)
        divergent = [h for _, _, h in hist if h["manual_match"] == "불일치"]

        loc = None
        if tag and tag in self.drawings:
            d = self.drawings[tag][0]
            loc = {
                "dwg_no": d["sheet_no"],
                "pdf_page": d["page"],
                "drawings": self.drawings[tag],
                "panel": inst.get("PANEL") if inst else None,
                "terminal": inst.get("TERMINAL") if inst else None,
                "plc": inst.get("PLC") if inst else None,
                "slot_ch": "%s / CH%s" % (inst.get("SLOT"), inst.get("CH"))
                           if inst else None,
            }

        return {
            "tag": tag,
            "instrument": {
                "model": device,
                "maker": inst.get("MAKER") if inst else None,
                "service": inst.get("SERVICE") if inst else None,
                "system": inst.get("SYSTEM") if inst else None,
                "loop_group": inst.get("LOOP GROUP") if inst else None,
                "fault_mode": inst.get("FAULT MODE") if inst else None,
                "manual_file": inst.get("MANUAL FILE") if inst else None,
            } if inst else None,
            "manual": [{
                "score": s, "id": c["id"], "code": c["code"],
                "severity": c["severity"], "name": c["name"],
                "description": c["description"], "remedy": c["remedy"],
                "cite": "%s p.%d (%s)" % (c["source"]["file"],
                                          c["source"]["pdf_page"],
                                          c["source"]["section"]),
            } for s, c in manual],
            "history": [{
                "score": s, "why": why, "wo_no": h["wo_no"], "date": h["date"],
                "tag": h["tag"], "symptom": h["symptom"],
                "root_cause": h["root_cause"], "action": h["action_taken"],
                "match": h["manual_match"], "duration_min": h["duration_min"],
                "parts": h["parts"],
            } for s, why, h in hist],
            "comparison": {
                "counts": dict(cmp_counts),
                "divergent_count": len(divergent),
                "note": ("현장 이력이 매뉴얼과 다른 원인을 지목한 사례가 %d건 있습니다."
                         % len(divergent)) if divergent else
                        ("현장 이력이 매뉴얼 설명과 대체로 일치합니다."
                         if hist else "관련 현장 이력이 없습니다."),
            },
            "location": loc,
        }


# ─────────────────────────────────────────────────────────────
def render(a):
    L = []
    P = L.append
    P("=" * 74)
    P("TAG %s" % (a["tag"] or "-"))
    if a["instrument"]:
        i = a["instrument"]
        P("  %s %s | %s | %s / %s" % (i["maker"], i["model"], i["service"],
                                      i["system"], i["loop_group"]))
        P("  결함 거동: %s" % i["fault_mode"])
    P("")
    P("[매뉴얼 — 벤더 문서 근거]")
    if not a["manual"]:
        P("  해당 없음")
    for m in a["manual"]:
        P("  · %-14s %-9s %s" % (m["code"] or m["id"], m["severity"], m["name"]))
        P("      %s" % m["description"][:150])
        if m["remedy"]:
            P("      조치: %s" % m["remedy"][:130])
        P("      근거: %s" % m["cite"])
    P("")
    P("[현장 이력 — 우리 경험]")
    direct = [h for h in a["history"] if "동일 매뉴얼 코드" in h["why"]]
    other = [h for h in a["history"] if "동일 매뉴얼 코드" not in h["why"]]
    if not a["history"]:
        P("  해당 없음")
    for label, group in (("직접 관련", direct), ("이 태그의 다른 이력", other)):
        if not group:
            continue
        P("  < %s >" % label)
        for h in group:
            P("  · %s %s (%s) [%s]" % (h["date"], h["wo_no"], h["tag"], h["match"]))
            P("      증상   : %s" % h["symptom"])
            P("      실제원인: %s" % h["root_cause"])
            P("      조치   : %s  (%d분, 부품 %s)" % (h["action"],
                                                    h["duration_min"], h["parts"]))
    P("")
    c = a["comparison"]
    P("[대비] %s" % c["note"])
    if c["counts"]:
        P("       일치도 분포: %s" % ", ".join("%s %d건" % kv
                                             for kv in c["counts"].items()))
    if a["location"]:
        l = a["location"]
        P("")
        P("[위치] 도면 %s (PDF %dp) | 판넬 %s 단자 %s | %s 슬롯 %s"
          % (l["dwg_no"], l["pdf_page"], l["panel"], l["terminal"],
             l["plc"], l["slot_ch"]))
    P("=" * 74)
    return "\n".join(L)


def main():
    ap = argparse.ArgumentParser(description="Plant Maintenance Copilot 조회")
    ap.add_argument("--src", default="demo_data")
    ap.add_argument("--tag", default=None)
    ap.add_argument("--alarm", default="", help="알람 문구 (한글 가능)")
    ap.add_argument("--code", default=None, help="계기 화면에 뜬 코드")
    ap.add_argument("--json", action="store_true", help="JSON 출력")
    args = ap.parse_args()

    cp = Copilot(args.src)
    a = cp.answer(tag=args.tag, alarm=args.alarm, code=args.code)
    if args.json:
        print(json.dumps(a, ensure_ascii=False, indent=2))
    else:
        print(render(a))


if __name__ == "__main__":
    main()
